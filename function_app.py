import json
from math import sqrt, cos
import re
import azure.functions as func
import os, logging, uuid
from azure.storage.blob import BlobServiceClient
from azure.identity import DefaultAzureCredential
import urllib.request
import pandas as pd
import xml.etree.ElementTree as ET
import defusedxml.ElementTree as dET

# from loguru import logger as LOGGER
import traceback
import openpyxl

import math

import socket
print(socket.gethostname())

import pprint
# from dictsearch.search import iterate_dictionary


# import matplotlib.pyplot as plt
# import numpy as np



MAX_REAL_FLOORS = 10

# https://ksnmagicplanfunc3e54b9.file.core.windows.net/attachment/Survey Portal Excel Sheet_Export_Template.xlsx


def cart_distance(p1 : tuple[float, float], p2 : tuple[float, float]) -> float:
    (x1, y1) = p1
    (x2, y2) = p2
    return sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)

def create_table(dict : dict[str, list[float]], headers : list,
                  do_not_sum : list[str] = [], 
                  styling: str = "", colour_table : bool = False) -> str:
    
    output = f'<table {styling}><tr>'
    
    for header in headers:
        output += f'<th>{header}</th>'
    output += '</tr>'
    
    for i, key in enumerate(dict):
        if colour_table:
            output += f'<tr><td><font color="{key[:len(key)-2]}"><b>Colour {i}</b></font></td>'
        else:
            output += f'<tr><td>{key}</td>'
        for elem in dict[key]:
            output += f'<td>{round(elem, 2)}</td>'
        if do_not_sum != ['All']:
            if key in do_not_sum:
                output += '<td>N/A</td></tr>'
            else:
                output += f'<td>{round(sum(dict[key]), 2)}</td></tr>'
    
    output += '</table>'
    return output




def create_table_text(dict, headers : list,
                  do_not_sum : list[str] = [], 
                  styling: str = "", colour_table : bool = False
                  , order_list = []) -> str:
    try:
        
        # print(dict)
        
        
        
        
        output = f'<table {styling}><tr>'
        
        for header in headers:
            output += f'<th>{header}</th>'
        output += '</tr>'
        
        if len(order_list) != 0:
            for item in order_list:
                if item.isupper():
                    output += f'<tr><td><strong>{item}</strong></td>'
                else:
                    output += f'<tr><td>{item}</td>'
                value = dict[item] if item in dict.keys() else ''
                if (type(value) == bool and value == True):
                    value = "Yes"
                if (type(value) == bool and value == False):
                    value = "No"
                output += f'<td>{value}</td>'
                # print(item, value)
        else:
            for i, key in enumerate(dict):
                # print(key, dict[key])
                if key.isupper():
                    output += f'<tr><td><strong>{key}</strong></td>'
                elif colour_table:
                    output += f'<tr><td><font color="{key[:len(key)-2]}"><b>Colour {i}</b></font></td>'
                else:
                    output += f'<tr><td>{key}</td>'
                output += f'<td>{dict[key]}</td>'

        output += '</table>'
        
        
    except Exception as ex:
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # output = "Line " + str(exc_tb.tb_lineno) + ": " + exc_type 
        
        output = str(ex)
        output = traceback.format_exc()
        # LOGGER.info('Exception : ' + str(traceback.format_exc()))
        print(output)
        
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print(exc_type, fname, exc_tb.tb_lineno)
    finally:
        return output
    return output

def roof_general(json_val_dict):
    json_val_dict["Room in Roof"] = False
    json_val_dict["Suitable for Insulation *"] = False
    json_val_dict["Roof not suitable details*"] = ''
    json_val_dict["Notes (Roof)"] = ''
    
    if "sfi" not in json_val_dict.keys():
        json_val_dict["sfi"] = []
    if "sfi_dict" not in json_val_dict.keys():
        json_val_dict["sfi_dict"] = {}
    
    
    
    for n in range(1, 5):
        if f"Roof Type {n} Suitable for Insulation" in json_val_dict.keys():
            # print(f"Roof Type {n} Suitable for Insulation", json_val_dict[f"Roof Type {n} Suitable for Insulation"])
            if json_val_dict[f"Roof Type {n} Suitable for Insulation"] == True:
                json_val_dict["Suitable for Insulation *"] = True
                json_val_dict["sfi"].append(n)
                # json_val_dict["sfi_dict"][n] = 300

        if f"Roof type {n} Suitable for Insulation*" in json_val_dict.keys():
            # print(f"Roof type {n} Suitable for Insulation*", json_val_dict[f"Roof type {n} Suitable for Insulation*"])
            if json_val_dict[f"Roof type {n} Suitable for Insulation*"] == True:
                json_val_dict["Suitable for Insulation *"] = True
                json_val_dict["sfi"].append(n)
                # json_val_dict["sfi_dict"][n] = 300

        if f"Roof Type {n} Suitable for Insulation*" in json_val_dict.keys():
            # print(f"Roof Type {n} Suitable for Insulation*", json_val_dict[f"Roof Type {n} Suitable for Insulation*"])
            if json_val_dict[f"Roof Type {n} Suitable for Insulation*"] == True:
                json_val_dict["Suitable for Insulation *"] = True
                json_val_dict["sfi"].append(n)
                # json_val_dict["sfi_dict"][n] = 300

        if f"Roof Type {n} Sloping Ceiling Suitable for Insulation*" in json_val_dict.keys():
            # print(f"Roof Type {n} Sloping Ceiling Suitable for Insulation*", json_val_dict[f"Roof Type {n} Sloping Ceiling Suitable for Insulation*"])
            if json_val_dict[f"Roof Type {n} Sloping Ceiling Suitable for Insulation*"] == True:
                json_val_dict["Suitable for Insulation *"] = True
                json_val_dict["sfi"].append(str(n) + 's')
                # json_val_dict["sfi_dict"][str(n) + 's'] = 300

    # print('sfi: ', json_val_dict["sfi"])
    # print('sfi: ', json_val_dict["sfi_dict"])

    for n in range(1, 5):
        # Suitable?
        if n not in json_val_dict["sfi"]:
            continue
        
        # Existing? (Thickness)
        e = 0
        if f"Roof {n} Thickness (mm)*" in json_val_dict.keys():
            e = json_val_dict[f"Roof {n} Thickness (mm)*"]
        elif f"Roof {n} Thickness (mm)" in json_val_dict.keys():
            e = json_val_dict[f"Roof {n} Thickness (mm)"]
        # print(f"Roof {n} Thickness (mm)", ": ",  e)
        
        # value: Area - add to appropriate dict entry
        # print(f"roof_{n}_area", ": ", json_val_dict[f"roof_{n}_area"])
        for t in [100, 150, 200, 250, 300]:
            if int(e) + int(t) >= 300:
                # print(f"need to add roof_{n}_area to dict entry {t}")
                key = str(t)
                if key not in json_val_dict["sfi_dict"].keys():
                    json_val_dict["sfi_dict"][key] = json_val_dict[f"roof_{n}_area"]
                else:
                    json_val_dict["sfi_dict"][key] += json_val_dict[f"roof_{n}_area"]
                break


    for n in range(1, 5):
        if json_val_dict["Suitable for Insulation *"] == False:
            if f"Roof Type {n} Not Suitable Details" in json_val_dict.keys():
                # print('n', ':', n, json_val_dict[f"Roof Type {n} Not Suitable Details"])
                json_val_dict["Roof not suitable details*"] += f"Roof Type {n} Not Suitable Details: "
                json_val_dict["Roof not suitable details*"] += json_val_dict[f"Roof Type {n} Not Suitable Details"]
                json_val_dict["Roof not suitable details*"] += "<BR>"
            if f"Roof Type {n} Sloping Ceiling Not Suitable Details*" in json_val_dict.keys():
                # print('n', ':', n, json_val_dict[f"Roof Type {n} Sloping Ceiling Not Suitable Details*"])
                json_val_dict["Roof not suitable details*"] += f"Roof Type {n} Sloping Ceiling Not Suitable Details: "
                json_val_dict["Roof not suitable details*"] += json_val_dict[f"Roof Type {n} Sloping Ceiling Not Suitable Details*"]
                json_val_dict["Roof not suitable details*"] += "<BR>"
        else:
            json_val_dict["Roof not suitable details*"] = 'N/A'
    # print('json_val_dict["Roof not suitable details*"]: ', json_val_dict["Roof not suitable details*"])
    
    for n in range(1, 5):
        if f"Notes (Roof Type {n})" in json_val_dict.keys():
            if json_val_dict[f"Notes (Roof Type {n})"] not in json_val_dict["Notes (Roof)"]:
                json_val_dict["Notes (Roof)"] += f"Notes (Roof Type {n}): "
                json_val_dict["Notes (Roof)"] += json_val_dict[f"Notes (Roof Type {n})"]
                json_val_dict["Notes (Roof)"] += "<BR>"
        if f"Notes (Roof Type {n})*" in json_val_dict.keys():
            if json_val_dict[f"Notes (Roof Type {n})*"] not in json_val_dict["Notes (Roof)"]:
                json_val_dict["Notes (Roof)"] += f"Notes (Roof Type {n})*: "
                json_val_dict["Notes (Roof)"] += json_val_dict[f"Notes (Roof Type {n})*"]
                json_val_dict["Notes (Roof)"] += "<BR>"
    # print('json_val_dict["Notes (Roof)"]: ', json_val_dict["Notes (Roof)"])
    
    for n in range(1, 5):
        if f"Roof {n} Type*" in json_val_dict.keys():
            # print(f"Roof {n} Type*", json_val_dict[f"Roof {n} Type*"])
            if json_val_dict[f"Roof {n} Type*"] == "Dormer / room in roof":
                json_val_dict["Room in Roof"] = True
        if f"Roof {n} Type" in json_val_dict.keys():
            # print(f"Roof {n} Type", json_val_dict[f"Roof {n} Type"])
            if json_val_dict[f"Roof {n} Type"] == "Dormer / room in roof":
                json_val_dict["Room in Roof"] = True
    



def walls_general(json_val_dict):

    json_val_dict["Is the property suitable for wall insulation? *"] = False
    json_val_dict["No wall insulation details *"] = ''
    json_val_dict["Notes (Walls)"] = ''
    
    if "sfwi" not in json_val_dict.keys():
        json_val_dict["sfwi"] = []
    if "sfwi_dict" not in json_val_dict.keys():
        json_val_dict["sfwi_dict"] = {}
    
    
    
    
    for n in range(1, 5):
        if f"Is wall type {n} suitable for wall insulation?" in json_val_dict.keys():
            # print(f"Is wall type {n} suitable for wall insulation?", json_val_dict[f"Is wall type {n} suitable for wall insulation?"])
            if json_val_dict[f"Is wall type {n} suitable for wall insulation?"] == True:
                json_val_dict["Is the property suitable for wall insulation? *"] = True
                json_val_dict["sfwi"].append(n)
                # json_val_dict["sfwi_dict"][n] = 300

        if f"Is wall type {n} suitable for wall insulation?*" in json_val_dict.keys():
            # print(f"Is wall type {n} suitable for wall insulation?*", json_val_dict[f"Is wall type {n} suitable for wall insulation?*"])
            if json_val_dict[f"Is wall type {n} suitable for wall insulation?*"] == True:
                json_val_dict["Is the property suitable for wall insulation? *"] = True
                json_val_dict["sfwi"].append(n)
                # json_val_dict["sfwi_dict"][n] = 300


    # print('sfwi: ', json_val_dict["sfwi"])
    # print('sfwi: ', json_val_dict["sfwi_dict"])



    for n in range(1, 5):
        if json_val_dict["Is the property suitable for wall insulation? *"] == False:
            if f"No wall type {n} insulation details" in json_val_dict.keys():
                # print('n', ':', n, json_val_dict[f"No wall type {n} insulation details"])
                json_val_dict["No wall insulation details *"] += f"No wall type {n} insulation details: "
                json_val_dict["No wall insulation details *"] += json_val_dict[f"No wall type {n} insulation details"]
                json_val_dict["No wall insulation details *"] += "<BR>"
        else:
            json_val_dict["No wall insulation details *"] = 'N/A'
    # print('json_val_dict["No wall insulation details *"]: ', json_val_dict["No wall insulation details *"])
    
    for n in range(1, 5):
        if f"Notes (Wall type {n} Walls)" in json_val_dict.keys():
            if json_val_dict[f"Notes (Wall type {n} Walls)"] not in json_val_dict["Notes (Walls)"]:
                json_val_dict["Notes (Walls)"] += f"Notes (Wall type {n} Walls): "
                json_val_dict["Notes (Walls)"] += json_val_dict[f"Notes (Wall type {n} Walls)"]
                json_val_dict["Notes (Walls)"] += "<BR>"
        if f"Notes (Wall type {n} Walls)*" in json_val_dict.keys():
            if json_val_dict[f"Notes (Wall type {n} Walls)*"] not in json_val_dict["Notes (Walls)"]:
                json_val_dict["Notes (Walls)"] += f"Notes (Wall type {n} Walls)*: "
                json_val_dict["Notes (Walls)"] += json_val_dict[f"Notes (Wall type {n} Walls)*"]
                json_val_dict["Notes (Walls)"] += "<BR>"
    # print('json_val_dict["Notes (Walls)"]: ', json_val_dict["Notes (Walls)"])
    
    if json_val_dict["Is the property suitable for wall insulation? *"] == False:
        json_val_dict["No wall insulation details *"] += json_val_dict["Notes (Walls)"]

def is_point_in_line_segment(x1, y1, a1, b1, a2, b2, epsilon=0.001, zeta=0.0):
    # print('checking if (' + str(x1) + ',' + str(y1) + ') is contained in (' + str(a1) + ',' + str(b1) + ') -> (' + str(a2) + ',' + str(b2) + ')')
    
    
    cp = (y1 - b1) * (a2 - a1) - (x1 - a1) * (b2 - b1)
    if abs(cp) > epsilon:
        # print('abs(cp)', ':', abs(cp), ' > ', 'epsilon', ':', epsilon)
        return False
    
    dp = (x1 - a1) * (a2 - a1) + (y1 - b1) * (b2 - b1)
    if dp < (0 - zeta):
        # print('dp', ':', dp, ' < ', '0 - zeta', ':', 0 - zeta)
        return False
    
    slba = (a2 - a1) * (a2 - a1) + (b2 - b1) * (b2 - b1) # pythagoras
    if dp > slba:
        # print('dp', ':', dp, ' > ', 'slba', ':', slba)
        return False
    
    return True



def linear_subset(x1, y1, x2, y2, a1, b1, a2, b2, epsilon=0.001, zeta=0.0):
    
    # print('zeta', ':', zeta)
    
    if not is_point_in_line_segment(x1, y1, a1, b1, a2, b2, epsilon, zeta):
        return False
    
    if not is_point_in_line_segment(x2, y2, a1, b1, a2, b2, epsilon, zeta):
        return False
    
    return True



def XML_2_dict(root, t = "floor"):
    try:
        # d = {}
        xml_ref_dict = {}
        nwa_dict = {}
        obj_dict = {}
        xml_val_dict = {}
        
        id = root.get('id')
        plan_name = root.get('name')
        if plan_name[-1] == ' ':
            plan_name = plan_name[:-1]
        print('id', ':', id)
        print('plan_name', ':', plan_name)
        xml_val_dict['Application ID'] = plan_name
        xml_val_dict['id'] = id
        xml_val_dict['plan_name'] = plan_name
        
        xml_val_dict['Client Address'] = ''
        address_fields = ['street', 'city', 'province', 'country', 'postalCode']
        for af in address_fields:
            f = root.get(af)
            if f is not None:
                xml_val_dict['Client Address'] = (xml_val_dict['Client Address'] + ', ' + str(f)) if xml_val_dict['Client Address'] != '' else str(f)
        xml_val_dict['Eircode'] = root.get('postalCode')
        
        
        # date = root.find('values/value[@key="date"]').text
        # xml_val_dict['Survey Date'] = date
        
        
        MagicPlan_2_SEAI_dict = {"date": "Survey Date", "qf.34d66ce4q3": "rating_type", "qf.34d66ce4q4": "rating_purpose", "author": "Surveyor"}
        
        # note MagicPlan also has a separate "Surveyor" field ("qf.34d66ce4q1") but "author" is the one used for SEAI survey purposes
        
        
        values = root.findall('values/value')
        for value in values:
            k = value.attrib["key"]
            for mpk in MagicPlan_2_SEAI_dict:
                if k == mpk:
                    xml_val_dict[MagicPlan_2_SEAI_dict[mpk]] = value.text
                    # print(MagicPlan_2_SEAI_dict[mpk], ':', xml_val_dict[MagicPlan_2_SEAI_dict[mpk]])
        
        
        
        
        # w = {}
        wd_list = ['634004d284d12@edit:0063fa41-fa2d-4493-9f86-dcd0263e8108'
                    , '634004d284d12@edit:0ecdca7d-a4c3-4692-893a-89e6eaa76e74'
                    , '634004d284d12@edit:28960da1-84f6-4f3b-a446-7c72b9febe9f'
                    , '634004d284d12@edit:28b0fb8c-47a4-4d9e-8ce5-2b35a1a0404e'
                    , '634004d284d12@edit:2b72a58f-7380-4b6c-9d74-667f937a9b57'
                    , '634004d284d12@edit:32b043c7-432a-409f-972d-a75b386b1789'
                    , '634004d284d12@edit:60194a47-84ce-414b-8368-69ec53167111'
                    , '634004d284d12@edit:6976cc78-3a2e-4935-99c6-6aff8011be8a'
                    , '634004d284d12@edit:735122f1-ab8b-47e8-b5ca-d4ec4d492f1c'
                    , '634004d284d12@edit:7d851726-6ff6-48f7-8371-9ea09bd5179f'
                    , '634004d284d12@edit:7f6101da-4b6d-4c31-9293-d59552aeff3a'
                    , '634004d284d12@edit:a9a0a953-0fd3-4733-b161-de4f08fe5d49'
                    , '634004d284d12@edit:e6026a1e-3089-4fe7-9ec4-8504b001eb2e'
                    , '634004d284d12@edit:fc02c0c5-d9d8-4679-8a77-dc75edf7f592'
                    , 'arcdoor'
                    , 'doorbypass'
                    , 'doorbypassglass'
                    , 'doordoublefolding'
                    , 'doordoublehinged'
                    , 'doordoublesliding'
                    , 'doorfolding'
                    , 'doorfrench'
                    , 'doorgarage'
                    , 'doorglass'
                    , 'doorhinged'
                    , 'doorpocket'
                    , 'doorsliding'
                    , 'doorslidingglass'
                    , 'doorswing'
                    , 'doorwithwindow'
                    , 'windowarched'
                    , 'windowawning'
                    , 'windowbay'
                    , 'windowbow'
                    , 'windowcasement'
                    , 'windowfixed'
                    , 'windowfrench'
                    , 'windowhopper'
                    , 'windowhung'
                    , 'windowsliding'
                    , 'windowtrapezoid'
                    , 'windowtriangle'
                    , 'windowtskylight1'
                    , 'windowtskylight2'
                    , 'windowtskylight3'
                    ]
        xml_ref_dict['habitable_rooms'] = []
        xml_ref_dict['wet_rooms'] = []
        xml_ref_dict['exclude_rooms'] = []
        # xml_ref_dict['include_rooms'] = []
        xml_ref_dict['exclude_room_types'] = ['Attic', 'Balcony', 'Storage', 'Patio', 'Deck', 'Porch', 'Cellar', 'Garage', 'Furnace Room', 'Outbuilding', 'Unfinished Basement', 'Workshop']
        
        xml_ref_dict['habitable_room_types'] = ['Kitchen', 'Dining Room', 'Living Room', 'Bedroom', 'Primary Bedroom', "Children's Bedroom", 'Study', 'Music Room']
        xml_ref_dict['wet_room_types'] = ['Kitchen', 'Bathroom', 'Half Bathroom', 'Laundry Room', 'Toilet', 'Primary Bathroom']

        
        
        
        
        
        
        floors = root.findall('interiorRoomPoints/floor')
        for floor in floors:
            ft = floor.get('floorType')
            xml_ref_dict[floor.get('floorType')] = floor.get('uid')
            xml_ref_dict[floor.get('uid')] = floor.get('floorType')
            nwa_dict[ft] = {}
            
            for room in floor.findall('floorRoom'):
                if room.get('type') not in xml_ref_dict.keys():
                    xml_ref_dict[room.get('type')] = []
                xml_ref_dict[room.get('type')].append(room.get('uid'))
                xml_ref_dict[room.get('uid')] = room.get('type')
                # print(room.get('type'))
                if room.get('type') in xml_ref_dict['habitable_room_types']:
                    xml_ref_dict['habitable_rooms'].append(room.get('uid'))
                    xml_ref_dict['habitable_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid'))
                
                if room.get('type') in xml_ref_dict['wet_room_types']:
                    xml_ref_dict['wet_rooms'].append(room.get('uid'))
                    xml_ref_dict['wet_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid'))
                
                if room.get('type') in xml_ref_dict['exclude_room_types']:
                    xml_ref_dict['exclude_rooms'].append(room.get('uid'))
                    xml_ref_dict['exclude_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid') + " (" + room.get('area') + ")")
                # else:
                    # xml_ref_dict['include_rooms'].append(room.get('uid'))
                    # xml_ref_dict['include_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid') + " (" + room.get('area') + ")")
                
                # print('exclude_rooms', ':', xml_ref_dict['exclude_rooms'])
                
                for value in room.findall('values/value'):
                    key = value.get('key')
                    # print(key)
                    if key == "qcustomfield.2979903aq1": # Include?
                        # print(room.get('type'))
                        floor_area_include = value.text
                        # print('floor_area_include', ':', floor_area_include)
                        # if floor_area_include == '0':
                            # xml_ref_dict['exclude_rooms'].append(room.get('uid'))
                        if floor_area_include == '1':
                            if room.get('uid') in xml_ref_dict['exclude_rooms']:
                                xml_ref_dict['exclude_rooms'].remove(room.get('uid'))
                                xml_ref_dict['exclude_rooms'].remove('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid') + " (" + room.get('area') + ")")
                                # print(xml_ref_dict['exclude_rooms'])
                                # print(room.get('type'))
                
                # print('exclude_rooms', ':', xml_ref_dict['exclude_rooms'])
                
                rt = room.get('type') + ' (' + room.get('uid') + ')'
                x = {}
                room_x = room.get('x')
                room_y = room.get('y')
                w_index = 0
                for point in room.findall('point'):
                    w_index += 1
                    # uid = point.get('uid')
                    x[w_index] = {}
                    for value in point.findall('values/value'):
                        if value.get('key') in ['qf.c52807ebq1', 'qf.bdbaf056q1', 'qf.c52807ebq1']:
                            x[w_index]['type'] = value.text
                    # if 'type' not in list(x[w_index].keys()):
                        # x.pop(w_index)
                        # continue
                    x[w_index]['uid'] = point.get('uid')
                    x[w_index]['x1'] = float(point.get('snappedX')) + float(room_x)
                    x[w_index]['y1'] = -float(point.get('snappedY')) - float(room_y)
                    x[w_index]['h'] = point.get('height')
                    for value in point.findall('values/value'):
                        if value.get('key') == "loadBearingWall":
                            # print("loadBearingWall", ':', value.text)
                            x[w_index]['loadBearingWall'] = value.text
                # print('ft', ':', ft)
                # print('rt', ':', rt)
                # print('x', ':', x)
                # print('len(x)', ':', len(x))
                
                        
                w_index = 0
                for wall in x:
                    w_index += 1
                    # print(list(x[1].keys()))
                    if w_index + 1 in list(x.keys()):
                        x[w_index]['x2'] = x[w_index + 1]['x1']
                        x[w_index]['y2'] = x[w_index + 1]['y1']
                    else:
                        x[w_index]['x2'] = x[1]['x1']
                        x[w_index]['y2'] = x[1]['y1']
                    x[w_index]['l'] = cart_distance((x[w_index]['x1'], x[w_index]['y1']), (x[w_index]['x2'], x[w_index]['y2']))
                    x[w_index]['a'] = float(x[w_index]['l']) * float(x[w_index]['h'])
                
                y = {}
                for wall in x:
                    uid = x[wall]['uid']
                    y[uid] = {}
                    if 'type' in list(x[wall].keys()):
                        y[uid]['type'] = x[wall]['type']
                    y[uid]['x1'] = x[wall]['x1']
                    y[uid]['y1'] = x[wall]['y1']
                    y[uid]['x2'] = x[wall]['x2']
                    y[uid]['y2'] = x[wall]['y2']
                    y[uid]['h'] = x[wall]['h']
                    y[uid]['l'] = x[wall]['l']
                    y[uid]['a'] = x[wall]['a']
                    # print(list(x[wall].keys()))
                    if 'loadBearingWall' in list(x[wall].keys()):
                        y[uid]['loadBearingWall'] = x[wall]['loadBearingWall']
                        
                
                # print('len(y)', ':', len(y))
                # print('y', ':', y)
                # print('adding wall dict y for room ' + rt + ' to nwa_dict')
                nwa_dict[ft][rt] = y
                
        # print('nwa_dict', ':', nwa_dict)
        
        # print("xml_ref_dict['exclude_rooms']", ':', str(xml_ref_dict['exclude_rooms']))
        # print("xml_ref_dict['include_rooms']", ':', str(xml_ref_dict['include_rooms']))
        
        
        
        
        # Create Object Dictionary 
            # - first get list of all objects on each floor
            # - then add any additional details available from "exploded" section (linked via "id" e.g. "W-1-5")
        floors = root.findall('floor')
        for floor in floors:
            ft = floor.get('floorType')
            
            o = {}
            
            for p in floor.findall('symbolInstance'):
                # print("p.get('uid')", ':', p.get('uid'))
                if p.get('symbol') in wd_list:
                    id = p.get('id')
                    o[id] = {}
                    o[id]['uid'] = p.get('uid')
                    o[id]['symbol'] = p.get('symbol')
            # print('o', ':', o)
                
            
            
            for p in floor.findall('exploded/door'):
                si = p.get('symbolInstance')
                # print('si', ':', si)
                if si in list(o.keys()):
                # o[si] = {}
                # o[si]['symbolInstance'] = window.get('symbolInstance')
                    o[si]['x1'] = p.get('x1')
                    o[si]['y1'] = -float(p.get('y1'))
                    o[si]['x2'] = p.get('x2')
                    o[si]['y2'] = -float(p.get('y2'))
                    o[si]['w'] = p.get('width')
                    o[si]['d'] = p.get('depth')
                    o[si]['h'] = p.get('height')
                    o[si]['a'] = float(o[si]['w']) * float(o[si]['h'])
            
            for p in floor.findall('exploded/window'):
                # o_index += 1
                si = p.get('symbolInstance')
                # print('si', ':', si)
                if si in list(o.keys()):
                # o[si] = {}
                # o[si]['symbolInstance'] = window.get('symbolInstance')
                    o[si]['x1'] = p.get('x1')
                    o[si]['y1'] = -float(p.get('y1'))
                    o[si]['x2'] = p.get('x2')
                    o[si]['y2'] = -float(p.get('y2'))
                    o[si]['w'] = p.get('width')
                    o[si]['d'] = p.get('depth')
                    o[si]['h'] = p.get('height')
                    o[si]['a'] = float(o[si]['w']) * float(o[si]['h'])
            
            
            
            
            # print('o', ':', o)
            obj_dict[ft] = o
            
            for room in floor.findall('floorRoom'):
                rt = room.get('type') + ' (' + room.get('uid') + ')'
                
                w = {}
                room_x = room.get('x')
                room_y = room.get('y')
                w_index = 0
                for point in room.findall('point'): # get (x3, y3)
                    w_index += 1
                    w[w_index] = {}
                    w[w_index]['uid'] = point.get('uid')
                    w[w_index]['x3'] = float(point.get('snappedX')) + float(room_x)
                    w[w_index]['y3'] = -float(point.get('snappedY')) - float(room_y)

                w_index = 0
                for wall in w: # get (x4, y4), the second point in each line segment - WARNING: relies on the assumption that the points are in order
                    w_index += 1
                    if w_index + 1 in list(w.keys()):
                        w[w_index]['x4'] = w[w_index + 1]['x3']
                        w[w_index]['y4'] = w[w_index + 1]['y3']
                    else:
                        w[w_index]['x4'] = w[1]['x3']
                        w[w_index]['y4'] = w[1]['y3']
                
                # print('ft', ':', ft)
                # print('rt', ':', '"' + rt + '"')
                # print('w', ':', w)
                
                for wall in w: # transfer values to nwa_dict (where wall key is "uid" instead of numbered index)
                    uid = w[wall]['uid']
                    nwa_dict[ft][rt][uid]['x3'] = w[wall]['x3']
                    nwa_dict[ft][rt][uid]['y3'] = w[wall]['y3']
                    nwa_dict[ft][rt][uid]['x4'] = w[wall]['x4']
                    nwa_dict[ft][rt][uid]['y4'] = w[wall]['y4']
                
                y = nwa_dict[ft][rt] # for brevity
                
                w_index = 0
                for wall in y:
                    w_index += 1
                    # print('wall', ':', wall)
                    
                    
                    
                    y[wall]['windows'] = []
                    y[wall]['net_a'] = y[wall]['a']
                    y[wall]['total_window_a'] = 0
                    for window in o:
                        if 'x1' not in list(o[window].keys()):
                            continue
                        if 'x3' not in list(o[window].keys()):
                            continue
                        # print(window)
                        if linear_subset(float(o[window]['x1']), float(o[window]['y1']), float(o[window]['x2']), float(o[window]['y2']), float(y[wall]['x3']), float(y[wall]['y3']), float(y[wall]['x4']), float(y[wall]['y4'])) == True:
                            y[wall]['windows'].append(window + ' (' + str(o[window]['a']) + ')')
                            y[wall]['net_a'] -= o[window]['a']
                            y[wall]['total_window_a'] += o[window]['a']
                            
                            # print('object ' + str(window) + ' (' + str(o[window]['x1']) + '\t' + str(o[window]['y1']) + ') -> (' + str(o[window]['x2']) + '\t' + str(o[window]['y2']) + ') is colinear with wall ' + str(wall) + ' (' + str(y[wall]['x3']) + '\t' + str(y[wall]['y3']) + ') -> (' + str(y[wall]['x4']) + '\t' + str(y[wall]['y4']) + ')')
                            # print('yes')
                    # print("w[wall]['windows']", ':', w[wall]['windows'])
                
                # print('y', ':', y)
                nwa_dict[ft][rt] = y
                
                
    except Exception as ex:
        output = str(ex) + "\n\n" + traceback.format_exc()
        # LOGGER.info('Exception : ' + str(traceback.format_exc()))
        print(output)
    
    finally:
        return xml_ref_dict, nwa_dict, xml_val_dict




def preBER(root):
    return
def inspection(root):
    return
def QA(root):
    return



def get_project_files(id, plan_name, headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"
            ,"key": "45170e50321733db78952dfa5901b0dfeeb8"
            , "customer": "63b5a4ae69c91"
            , "accept": "application/json"
            }):
            
    try:
        output = []
        generate_locally = True
        if (socket.gethostname()) == "PC1VXW6X":
            generate_locally = False
        # azure_upload(json_data)
        # account_url = os.environ['AZ_STR_URL']
        account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
        default_credential = DefaultAzureCredential()
        blob_service_client = BlobServiceClient(account_url, credential=default_credential)
        
        
        container_name = "project-files"
        container_client = blob_service_client.get_container_client(container_name)
        if not container_client.exists():
            container_client = blob_service_client.create_container(container_name)

        
        
        json_url = "https://cloud.magicplan.app/api/v2/plans/" + str(id) + "/files?include_photos=true"
        request = urllib.request.Request(json_url, headers=headers)
        JSON = urllib.request.urlopen(request).read()
        JSON = json.loads(JSON)
        
        # local_path = str(json_val_dict['Application ID'])
        # if not os.path.exists(local_path):
            # os.mkdir(local_path)
        
        
        
        
        for file in JSON["data"]["files"]:
            if file["file_type"] == "pdf":
                output.append(file["name"])
                # print('getting file: ' + file["name"])
                if generate_locally == True:
                    request = urllib.request.Request(file["url"], headers=headers)
                    file_content = urllib.request.urlopen(request).read()
                    local_file_name = file["name"]
                    blob_client = blob_service_client.get_blob_client(container=container_name, blob=os.path.join(plan_name, file["name"]))
                    blob_client.upload_blob(file_content, overwrite=True)

        
        for file in JSON["data"]["photos"]:
            local_file_name = file["name"]
            local_file_name = local_file_name.replace("10th Floor", "Ground Floor")
            local_file_name = local_file_name.replace("11th Floor", "1st Floor")
            local_file_name = local_file_name.replace("12th Floor", "2nd Floor")
            local_file_name = local_file_name.replace("13th Floor", "3rd Floor")
            local_file_name = local_file_name.replace("14th Floor", "4th Floor")
            output.append(local_file_name)
            # print('getting file: ' + file["name"], 'local: ', local_file_name)
            if generate_locally == True:
                request = urllib.request.Request(file["url"], headers=headers)
                file_content = urllib.request.urlopen(request).read()
                blob_client = blob_service_client.get_blob_client(container=container_name, blob=os.path.join(plan_name, local_file_name))
                blob_client.upload_blob(file_content, overwrite=True)
    
    except Exception as ex:
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # output = "Line " + str(exc_tb.tb_lineno) + ": " + exc_type 
        
        # output = str(ex)
        output = traceback.format_exc()
        print(output)
        
    finally:
        return output

def get_true_floor(ff):
    try:
        tf = ff
        tf = tf.replace("10th Floor", "Ground Floor")
        tf = tf.replace("11th Floor", "1st Floor")
        tf = tf.replace("12th Floor", "2nd Floor")
        tf = tf.replace("13th Floor", "3rd Floor")
        tf = tf.replace("14th Floor", "4th Floor")
    finally:
        return tf

def survey(root):
    try:
        output = ''
        xml_ref_dict, nwa_dict, xml_val_dict = XML_2_dict(root)
        # id = xml_val_dict['Application ID'] # take this out once all erroneous references have been updated
        id = xml_val_dict['id'] # take this out once all erroneous references have been updated
        plan_name = xml_val_dict['plan_name'] # take this out once all erroneous references have been updated
        
        
        # print('xml_val_dict', ':', xml_val_dict)
        json_val_dict = xml_val_dict # take this out once all erroneous references have been updated
        
        
        sfi = [] # a list to hold the numbers of roof (also wall?) types that are suitable for insulation
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"
            ,"key": "45170e50321733db78952dfa5901b0dfeeb8"
            , "customer": "63b5a4ae69c91"
            , "accept": "application/json"
            }


        ofl_filelist = []
        
        # if (socket.gethostname()) != "PC1VXW6X":
        # print('about to get project files for ' + plan_name + " (id: " + str(id) + ")")
        ofl_filelist = get_project_files(id, plan_name, headers)
        
        print('ofl_filelist', ':', ofl_filelist)
        print('finished getting project files')
        for filename in ofl_filelist:
            print(filename)
        
   
        # print('about to create (almost) empty attachment files for ' + plan_name + " (id: " + str(id) + ")")
        # populate_template(xml_val_dict) # adds an (almost) empty copy of the template to avoid potential Logic App error if file not found
        populate_template_new(xml_val_dict, 'template') # adds an (almost) empty copy of the template to avoid potential Logic App error if file not found
        # populate_template_new(xml_val_dict, 'template_mrc')
        print('finished creating empty attachment file')
            

        
        
        ofl_wos = ['ATTIC INSULATION'
                , 'Internal Wall Insulation: Sloped or flat (horizontal) surface'
                , 'Attic (Loft) Insulation 100 mm top-up'
                , 'Attic (Loft) Insulation 150 mm top-up'
                , 'Attic (Loft) Insulation 200 mm top-up'
                , 'Attic (Loft) Insulation 250 mm top up'
                , 'Attic (Loft) Insulation 300 mm'
                , 'Attic Storage (5m2)'
                , 'Installation of new attic hatch'
                , 'Additional Roof Ventilation (High Level)'
                , 'Additional Roof Ventilation (Low Level)'
                , 'WALLS'
                , 'Draught Proofing (<= 20m installed)'
                , 'Draught Proofing (> 20m installed)'
                , 'MEV 15l/s Bathroom'
                , 'MEV 30l/s Utility'
                , 'MEV 60l/s Kitchen'
                , 'Permanent ventilation wall vent (Certified Proprietary Integrated System)'
                , 'Background ventilation wall vent (Certified Proprietary Integrated System)'
                , 'Ducting existing cooker hood to exterior'
                , 'Cavity Wall Insulation Bonded Bead'
                , 'Loose Fibre Extraction'
                , 'External Wall Insulation: Less than 60m2'
                , 'External Wall Insulation: 60m2 to 85m2'
                , 'External Wall Insulation: Greater than 85m2'
                , 'ESB alteration'
                , 'GNI meter alteration'
                , 'GNI new connection'
                , 'RGI Meter_No Heating'
                , 'Internal Wall Insulation: Vertical Surface'
                , 'External wall insulation and CWI: less than 60m2'
                , 'External wall insulation and CWI: 60m2 to 85m2'
                , 'External wall insulation and CWI: greater than 85m2'
                , 'Window (same m2 rate will apply to windows with certified trickle vents)'
                , 'HEATING'
                , 'Basic gas heating system'
                , 'Basic oil heating system'
                , 'Full gas heating system installation'
                , 'Full oil heating system installation'
                , 'Gas boiler and controls (Basic & controls pack)'
                , 'Oil boiler and controls (Basic & controls pack)'
                , 'Hot Water Cylinder Jacket'
                , 'MECHANICAL VENTILATION SYSTEMS AND AIR TIGHTNESS TESTING & ENERGY'
                , 'Air-tightness test recommended?'
                , 'LED Bulbs: supply only (4 no.)'

                ]
        
        
        ofl_wos_2 = ['ATTIC INSULATION'
                , 'Internal Wall Insulation: Sloped or flat (horizontal) surface'
                , 'Attic (Loft) Insulation 100 mm top-up'
                , 'Attic (Loft) Insulation 150 mm top-up'
                , 'Attic (Loft) Insulation 200 mm top-up'
                , 'Attic (Loft) Insulation 250 mm top up'
                , 'Attic (Loft) Insulation 300 mm'
                , 'Attic Storage (5m2)'
                , 'Installation of new attic hatch'
                , 'Additional Roof Ventilation (High Level)'
                , 'Additional Roof Ventilation (Low Level)'
                , 'WALLS'
                , 'Draught Proofing (<= 20m installed)'
                , 'Draught Proofing (> 20m installed)'
                , 'MEV 15l/s Bathroom'
                , 'MEV 30l/s Utility'
                , 'MEV 60l/s Kitchen'
                , 'Permanent ventilation wall vent (Certified Proprietary Integrated System)'
                , 'Background ventilation wall vent (Certified Proprietary Integrated System)'
                , 'Ducting existing cooker hood to exterior'
                , 'Cavity Wall Insulation Bonded Bead'
                , 'Loose Fibre Extraction'
                , 'External Wall Insulation: Less than 60m2'
                , 'External Wall Insulation: 60m2 to 85m2'
                , 'External Wall Insulation: Greater than 85m2'
                , 'ESB alteration'
                , 'GNI meter alteration'
                , 'GNI new connection'
                , 'RGI Meter_No Heating'
                , 'Internal Wall Insulation: Vertical Surface'
                , 'External wall insulation and CWI: less than 60m2'
                , 'External wall insulation and CWI: 60m2 to 85m2'
                , 'External wall insulation and CWI: greater than 85m2'
                , 'Window (same m2 rate will apply to windows with certified trickle vents)'
                , 'HEATING'
                , 'Basic gas heating system'
                , 'Basic oil heating system'
                , 'Full gas heating system installation'
                , 'Full oil heating system installation'
                , 'Gas boiler and controls (Basic & controls pack)'
                , 'Oil boiler and controls (Basic & controls pack)'
                , 'Hot Water Cylinder Jacket'
                , 'MECHANICAL VENTILATION SYSTEMS AND AIR TIGHTNESS TESTING & ENERGY'
                , 'Air-tightness test recommended?'
                , 'LED Bulbs: supply only (4 no.)'

                ]
        
        ofl_pm = ['Internal Wall Insulation: Sloped or flat (horizontal) surface'
                , 'Attic (Loft) Insulation 100 mm top-up'
                , 'Attic (Loft) Insulation 150 mm top-up'
                , 'Attic (Loft) Insulation 200 mm top-up'
                , 'Attic (Loft) Insulation 250 mm top up'
                , 'Attic (Loft) Insulation 300 mm'
                , 'Cavity Wall Insulation Bonded Bead'
                , 'Loose Fibre Extraction'
                , 'External Wall Insulation: Less than 60m2'
                , 'External Wall Insulation: 60m2 to 85m2'
                , 'External Wall Insulation: Greater than 85m2'
                , 'Internal Wall Insulation: Vertical Surface'
                , 'External wall insulation and CWI: less than 60m2'
                , 'External wall insulation and CWI: 60m2 to 85m2'
                , 'External wall insulation and CWI: greater than 85m2'
                ]
        
        ofl_hpm = ['Basic gas heating system'
                , 'Basic oil heating system'
                , 'Full gas heating system installation'
                , 'Full oil heating system installation'
                , 'Gas boiler and controls (Basic & controls pack)'
                , 'Oil boiler and controls (Basic & controls pack)'
                , 'Hot Water Cylinder Jacket'
                ]
        

        
        ofl_s = ["Adequate Access*"
            , "No Access Details*"
            , "Cherry Picker Required*"
            , "Cherry Picker Required Details*"
            , "Mould/Mildew identified by surveyor; or reported by the applicant*"
            , "Mould/Mildew Details*"
            , "As confirmed by homeowner; property is a protected structure*"
            , "Protected Structure Details*"
            ]
        
        ofl_mr = ['Thermal Envelope - Heat loss walls, windows and doors' # all walls from floors 10 - 13 except Loadbearing/Party walls and internal walls)
                    , 'Thermal Envelope - Heat loss floor area' # surface area of floor 10 (excluding rooms as per attribute)
                    , 'Thermal Envelope - Heat loss roof area' # repeat value from above
                    , 'Heat loss Wall Area recommended for EWI and IWI' # 
                    , 'New Windows being recommended for replacement' # use value "replace_window_area" 
                    , 'Total Surface Area (m2)' # sum of first three above fields
                    , 'Total Surface Area receiving EWWR (m2)' # sum of 4th and 5th above fields
                    , 'Result %' # 7th divided by 6th * 100
                    , 'Is Major Renovation?' # yes if greater than or equal to 23%
                    ]
        
        ofl_mae = ["Number of habitable rooms in the property"
                    , "Number of wet rooms in the property"
                    , "No. of habitable/wet rooms w/ open flued appliance"
                    , "LED Bulbs: supply only (4 no.)"
                    , "Air-tightness test recommended?"
                    ]
        
        ofl_general = ['Dwelling Type*'
                        , 'Dwelling Age*'
                        , 'Age Extension 1'
                        , 'Age Extension 2'
                        , 'Asbestos Suspected'
                        , 'Asbestos Details' # only if suspected after 2000 
                        , 'Lot *' # blank for now
                        , 'Survey Date' # project creation date in MP
                        , 'Gross floor area (m2) *'
                        , 'Number of Storeys *' # 0-9
                        , 'Room in Roof' # yes if any "dormer / room in roof" else no
                        , 'No. Single Glazed Windows *'
                        , 'No. Double Glazed Windows *'
                        , 'Property Height (m)*'
                        , 'Internet Available'
                        ]
        
        ofl_roof = ['sloped_surface_area'
                , 'ins_100_area'
                , 'ins_150_area'
                , 'ins_200_area'
                , 'ins_250_area'
                , 'ins_300_area'
                , 'storage'
                , 'new_hatch_count'
                , 'high_roof_vent_area'
                , 'Roof 1 Type*'
                , 'Other Details Roof 1*'
                , 'Sloped Ceiling Roof 1*'
                , 'Roof 1 greater than 2/3 floor area*'
                , 'Roof 1 Pitch (degrees)*'
                , 'Roof Type 1 Insulation Exists*'
                , 'Can Roof Type 1 Insulation Thickness be Measured?*'
                , 'Roof 1 Thickness (mm)*'
                , 'Roof 1 Insulation Type*'
                , 'Required per standards (mm2) *'
                , 'Existing (mm2)*'
                , 'Area of Roof Type 1 with fixed flooring (m2)*'
                , 'Folding/stair ladder in Roof Type 1*'
                , 'Fixed light in Roof Type 1*'
                , 'Downlighters in Roof Type 1*'
                , 'High power cable in Roof Type 1 (6sq/10sq or higher)*'
                , 'Roof 2 Type'
                , 'Other Details Roof 2*'
                , 'Sloped Ceiling Roof 2*'
                , 'Roof 2 greater than 2/3 floor area*'
                , 'Roof 2 Pitch (degrees)*'
                , 'Roof 2 Insulation Exists*'
                , 'Can Roof Type 2 Insulation Thickness be Measured?'
                , 'Roof 2 Thickness (mm)'
                , 'Roof 2 Insulation Type'
                , 'Roof 2 Required per standards (mm2) *'
                , 'Roof 2 Existing (mm2) *'
                , 'Area of Roof Type 2 with fixed flooring (m2)'
                , 'Folding/stair ladder in Roof Type 2'
                , 'Fixed light in Roof Type 2'
                , 'Downlighters in Roof Type 2'
                , 'High power cable in Roof Type 2 (6sq/10sq or higher)'
                , 'Roof 3 Type'
                , 'Other Details Roof 3*'
                , 'Sloped Ceiling Roof 3*'
                , 'Roof 3 greater than 2/3 floor area*'
                , 'Roof 3 Pitch (degrees)*'
                , 'Roof 3 Insulation Exists*'
                , 'Can Roof Type 3 Insulation Thickness be Measured?*'
                , 'Roof 3 Thickness (mm)*'
                , 'Roof 3 Insulation Type*'
                , 'Roof 3 Required per standards (mm2) *'
                , 'Roof 3 Existing (mm2) *'
                , 'Area of Roof Type 3 with fixed flooring (m2)*'
                , 'Folding/stair ladder in Roof Type 3*'
                , 'Fixed light in Roof Type 3*'
                , 'Downlighters in Roof Type 3*'
                , 'High power cable in Roof Type 3 (6sq/10sq or higher)*'
                , 'Roof 4 Type'
                , 'Other Details Roof 4*'
                , 'Sloped Ceiling Roof 4*'
                , 'Roof 4 greater than 2/3 floor area'
                , 'Roof 4 Pitch (degrees)*'
                , 'Roof Type 4 Insulation Exists*'
                , 'Can Roof Type 4 Insulation Thickness be Measured?*'
                , 'Roof 4 Thickness (mm)'
                , 'Roof 4 Insulation Type*'
                , 'Roof 4 Required per standards (mm2) *'
                , 'Roof 4 Existing (mm2) *'
                , 'Area of Roof Type 4 with fixed flooring (m2)*'
                , 'Folding/stair ladder in Roof Type 4*'
                , 'Fixed light in Roof Type 4*'
                , 'Downlighters in Roof Type 4*'
                , 'High power cable in Roof Type 4 (6sq/10sq or higher)*'
                , 'Suitable for Insulation *'
                , 'Roof not suitable details*'
                , 'Notes (Roof)']
        
        ofl_walls = ['Wall Type 1*'
                    , 'Other Wall type 1 Details*'
                    , 'Wall 1 wall thickness (mm)*'
                    , 'Wall 1 Insulation Present?*'
                    , 'Wall 1 Insulation Type*'
                    , "Wall 1 Fill Type*"
                    , 'Wall 1 Residual Cavity Width (mm)*'
                    , 'Can Wall type 1 Insulation Thickness be Measured?*'
                    , "If 'Yes' enter Wall type 1 insulation thickness (mm)*"
                    , 'Wall Type 2'
                    , 'Other Wall type 2 Details*'
                    , 'Wall 2 wall thickness (mm)*'
                    , 'Wall 2 Insulation Present?*'
                    , 'Wall 2 Insulation Type*'
                    , "Wall 2 Fill Type*"
                    , 'Wall 2 Residual Cavity Width (mm)*'
                    , 'Can Wall type 2 Insulation Thickness be Measured?*'
                    , "If 'Yes' enter Wall type 2 insulation thickness (mm)*"
                    , 'Wall Type 3'
                    , 'Other Wall type 3 Details*'
                    , 'Wall 3 wall thickness (mm)*'
                    , 'Wall 3 Insulation Present?*'
                    , 'Wall 3 Insulation Type*'
                    , "Wall 3 Fill Type*"
                    , 'Wall 3 Residual Cavity Width (mm)*'
                    , 'Can Wall type 3 Insulation Thickness be Measured?*'
                    , "If 'Yes' enter Wall type 3 insulation thickness (mm)*"
                    , 'Wall Type 4'
                    , 'Other Wall type 4 Details*'
                    , 'Wall 4 wall thickness (mm)*'
                    , 'Wall 4 Insulation Present?*'
                    , 'Wall 4 Insulation Type*'
                    , "Wall 4 Fill Type*"
                    , 'Wall 4 Residual Cavity Width (mm)*'
                    , 'Can Wall type 4 Insulation Thickness be Measured?*'
                    , "If 'Yes' enter Wall type 4 insulation thickness (mm)*"
                    , "Is the property suitable for wall insulation? *"
                    , "No wall insulation details *"
                    , "EWI/IWI > 25% *"
                    , 'Suitable for Draught Proofing'
                    , 'Not suitable details Draughtproofing*'
                    , "Notes (Walls)"
                    
                    , "Draught Proofing (<= 20m installed)"
                    , "Draught Proofing (> 20m installed)"
                    , "MEV 15l/s Bathroom"
                    , "MEV 30l/s Utility"
                    , "MEV 60l/s Kitchen"
                    , "New Permanent Vent"
                    , "New Background Vent"
                    , "Duct Cooker Hood"
                    , "Cavity Wall Insulation Bonded Bead"
                    , "Loose Fibre Extraction"
                    , "External Wall Insulation: Less than 60m2"
                    , "External Wall Insulation: 60m2 to 85m2"
                    , "External Wall Insulation: Greater than 85m2"
                    , "ESB alteration"
                    , "GNI meter alteration"
                    # , "GNI new connection"
                    , "New Gas Connection"
                    , "RGI Meter_No Heating"
                    , 'Internal Wall Insulation: Vertical Surface'
                    , "External wall insulation and CWI: less than 60m2"
                    , "External wall insulation and CWI: 60m2 to 85m2"
                    , "External wall insulation and CWI: greater than 85m2"
                    , 'replace_window_area'
                    , 'Notes (Windows and Doors)'
                    ]
        
        ofl_heating = ['Heating System *'
                    , 'Qualifying Boiler'
                    , 'System Age *'
                    , 'Fully Working *'
                    , 'Requires Service *'
                    , "Other Primary Heating Details *"
                    , 'Not Working Details Primary Heating *'
                    , 'Requires Service (App?)*'
                    , 'Requires Service Details Primary Heating *'
                    
                    , 'Hot Water System Exists *'
                    , 'From Primary heating system'
                    , 'From Secondary heating system'
                    , 'Electric Immersion'
                    , 'Electric Instantaneous'
                    , 'Instantaneous Combi Boiler'
                    , 'Other'
                    , 'Other HW Details *'
                    , 'HWS'
                    
                    , 'Hot Water Cylinder*'
                    , 'Insulation *'
                    , 'Condition of Lagging Jacket *'
                    , 'HWC Controls *'
                    
                    , 'Heating Systems Controls *'
                    , 'Partial Details *'
                    , 'Programmer / Timeclock *'
                    , 'Room Thermostat Number *'
                    , 'Rads Number *'
                    , 'TRVs Number *'
                    
                    , 'Suitable for Heating Measures *'
                    , 'Not suitable details*'
                    , 'Notes (Heating)'
                    , 'Secondary Heating System'
                    , 'Secondary System Age *'
                    , 'Secondary System Fully Working *'
                    , 'Secondary System Requires Service *'
                    , 'Not Working Details Secondary Heating *'
                    , 'Secondary System Requires Service (App?)*'
                    , 'Requires Service Details Secondary Heating *'





                    ]

        
        
        
        
        
        
        
        
        
        habitable_room_types = ['Kitchen', 'Dining Room', 'Living Room', 'Bedroom', 'Primary Bedroom', "Children's Bedroom", 'Study', 'Music Room']
        wet_room_types = ['Kitchen', 'Bathroom', 'Half Bathroom', 'Laundry Room', 'Toilet', 'Primary Bathroom']
        
        
        
        json_val_dict['Hot Water Cylinder Jacket'] = ''
        
        json_val_dict["Electric Storage Heater age (years)*"] = ''
        json_val_dict["Warm Air System age (years)*"] = ''
        json_val_dict["Is there Mains Gas in the Area?"] = ''
        json_val_dict["Number of habitable rooms in the property"] = 0
        json_val_dict["Number of wet rooms in the property"] = 0
        json_val_dict["No. of habitable/wet rooms w/ open flued appliance"] = 0
        
        
        
        wt_dict = {}
        wt_dict['ext_wall_area_gross'], exploded_wall_dict = exterior_walls(root)
        # print("wt_dict['ext_wall_area_gross']", ':', wt_dict['ext_wall_area_gross'])
        wt_dict['gross'] = 0
        wt_dict['total'] = 0
        wt_dict['total_window_a'] = 0
        wt_dict['total_party_a'] = 0
        wt_dict['External.Wall.not.recieving.EWI.or.IWI'] = 0
        nwa_temp_dict = {}
        for floor in nwa_dict.keys():
            for room in nwa_dict[floor]:
                for wall in nwa_dict[floor][room]:
                    if 10 <= int(floor) <= 13:
                        # print(nwa_dict[floor][room][wall]['total_window_a'])
                        wt_dict['total_window_a'] += nwa_dict[floor][room][wall]['total_window_a']
                        if 'loadBearingWall' in list(nwa_dict[floor][room][wall].keys()):
                            if nwa_dict[floor][room][wall]['loadBearingWall'] == '1':
                                wt_dict['total_party_a'] += nwa_dict[floor][room][wall]['a']
                        # wt_dict['total_party_a'] += nwa_dict[floor][room][wall]['total_window_a']
                    if 10 <= int(floor) <= 13:
                        if 'type' in list(nwa_dict[floor][room][wall].keys()): # indicates wall recommended for insulation
                            # print("nwa_dict[floor][room][wall]['a']", ':', nwa_dict[floor][room][wall]['a'])
                            wt_dict['gross'] += float(nwa_dict[floor][room][wall]['a'])
                    for key in nwa_dict[floor][room][wall]:
                        name = 'floor ' + floor + '_' + room + '_wall ' + str(wall) + '_' + key
                        nwa_temp_dict[name] = nwa_dict[floor][room][wall][key]
                        if key == 'type':
                            if nwa_dict[floor][room][wall][key] == 'External.Wall.not.recieving.EWI.or.IWI':
                                req_area = nwa_dict[floor][room][wall]['a']
                            else:
                                req_area = nwa_dict[floor][room][wall]['net_a']
                            # print(name, 'req_area', ':', str(req_area))
                            if nwa_dict[floor][room][wall][key] in wt_dict.keys():
                                wt_dict[nwa_dict[floor][room][wall][key]] += req_area
                            else:
                                wt_dict[nwa_dict[floor][room][wall][key]] = req_area
                            wt_dict['total'] += req_area
        
        # print(nwa_temp_dict)
        
        
        
        
        wt_dict['ext_wall_area_net'] = wt_dict['ext_wall_area_gross'] - wt_dict['total_party_a']
        # print(wt_dict)
        
        # (if any value blank then 0)

        
        
        # print('nwa_dict[10]:')
        # print(nwa_dict['10'])
        # print('exploded_wall_dict:')
        # print(exploded_wall_dict)
        
        
        # json_data = json.dumps(
        # nwa_dict
        # )
        # j = r"d:\USERS\gshortall\Documents\Shortcut\investigate_A.json"
        # with open(j, "w") as investigate_file:
            # investigate_file.write(json_data)
        
        # json_data = json.dumps(
        # exploded_wall_dict
        # )
        # j = r"d:\USERS\gshortall\Documents\Shortcut\investigate_B.json"
        # with open(j, "w") as investigate_file:
            # investigate_file.write(json_data)
        
        # print(nwa_dict.keys())
        for floor in list(nwa_dict.keys()):
            if floor != "10":
                continue
            # print(floor)
            for room in nwa_dict[floor].keys():
                for wall in nwa_dict[floor][room].keys():
                    x1 = nwa_dict[floor][room][wall]["x1"]
                    y1 = nwa_dict[floor][room][wall]["y1"]
                    x2 = nwa_dict[floor][room][wall]["x2"]
                    y2 = nwa_dict[floor][room][wall]["y2"]
                    x3 = nwa_dict[floor][room][wall]["x3"]
                    y3 = nwa_dict[floor][room][wall]["y3"]
                    x4 = nwa_dict[floor][room][wall]["x4"]
                    y4 = nwa_dict[floor][room][wall]["y4"]
                    
                    # print(str(x1), str(y1))
                    # print(room, str(x1), str(y1), str(x3), str(y3))
                    # print(room, str(x2), str(y2), str(x4), str(y4))
                    # print(room, str(x3), str(y3))
                    # print(room, str(x4), str(y4))

        # else:
            # print('WARNING: No floor 10')
        
                    # for floor in exploded_wall_dict:
                    # print(floor)
                    d_min = 1
                    w_candidates = []
                    for wall in exploded_wall_dict[floor].keys():
                        # print(exploded_wall_dict[floor][wall])
                        x5 = exploded_wall_dict[floor][wall]["x1"]
                        y5 = exploded_wall_dict[floor][wall]["y1"]
                        x6 = exploded_wall_dict[floor][wall]["x2"]
                        y6 = exploded_wall_dict[floor][wall]["y2"]
                        w_type = exploded_wall_dict[floor][wall]["type"]
                        # print(wall, str(x5), str(y5))
                        # print(wall, str(x6), str(y6))
                        
                        d = cart_distance((x3, y3), (x5, y5))
                        if d == d_min:
                            w_candidates.append(wall)
                        if d < d_min:
                            d_min = d
                            w_candidates.append(wall)
                        d = cart_distance((x3, y3), (x6, y6))
                        if d == d_min:
                            w_candidates.append(wall)
                        if d < d_min:
                            d_min = d
                            w_candidates.append(wall)
                            
                    # print('d_min', ':', d_min)
                    print('w_candidates', ':', w_candidates)

        
        
        
        

        
        
        
        json_url = "https://cloud.magicplan.app/api/v2/plans/forms/" + str(xml_val_dict['id'])
        try:
            request = urllib.request.Request(json_url, headers=headers)
        except:
            output = traceback.format_exc()
            print(output)
        # finally:
            
        JSON = urllib.request.urlopen(request).read()
        JSON = json.loads(JSON)
        print('len(JSON["data"]): ', len(JSON["data"]))
        df = pd.DataFrame(JSON["data"])
        
        
        

        
        json_val_dict["Existing (mm2)*"] = int(0)
        json_val_dict['No. Single Glazed Windows *'] = 0
        # json_ref_dict = {}
        alt_double_glazed_count = 0
        



        replace_windows = []


        cylinder_stat = False
        json_val_dict['Heating System *'] = 'N/A'
        json_val_dict['Secondary Heating System'] = 'N/A'
        json_val_dict['HWS'] = ''
        json_val_dict['HWC Controls *'] = 'None'
        single_glazed_windows = []
        programmers = []
        room_thermostats = []
        
        
        json_val_dict["ESB alteration"] = ''
        json_val_dict["GNI meter alteration"] = ''
        json_val_dict["RGI Meter_No Heating"] = ''
        json_val_dict["New Gas Connection"] = ''
        
        # esb_alterations = []
        # gni_alterations = []
        # rgi_meter_no_heating = []
        # new_gas_connection = []
        
        
        
        condensing = ''
        linked_stove_bb = ''
        other_heating_notes = ''
        
        json_val_dict["Duct Cooker Hood"] = 0
        
        balanced_flues = []
        req_lagging_jackets = []
        slope_dict = {}
        roof_type_dict = {}
        h = {}
        h_index = 0
        for datum in JSON["data"]:
            # print(datum["symbol_name"])
            if 'Combi Boiler' in datum["symbol_name"]:
                json_val_dict['Hot Water System Exists *'] = True
                
            if 'Hot Water Cylinder' in datum["symbol_name"]:
                json_val_dict['Hot Water Cylinder*'] = True
                json_val_dict['Hot Water System Exists *'] = True
                if 'Bad' in datum["symbol_name"] or 'No Insulation' in datum["symbol_name"]:
                    json_val_dict['Condition of Lagging Jacket *'] = 'Bad'
                    req_lagging_jackets.append(datum["symbol_instance_id"])
                else:
                    json_val_dict['Condition of Lagging Jacket *'] = 'Good'
                
                if 'Lagging Jacket' in datum["symbol_name"]:
                    json_val_dict['Insulation *'] = 'Lagging Jacket'
                if 'Factory Fitted' in datum["symbol_name"]:
                    json_val_dict['Insulation *'] = 'Factory Fitted'
                if 'No Insulation' in datum["symbol_name"]:
                    json_val_dict['Insulation *'] = 'No Insulation'
                    
                    
                    
                
                
            if datum["symbol_name"] == "Programmer":
                programmers += datum["symbol_instance_id"]
            if datum["symbol_name"] == "Room Thermostat":
                room_thermostats += datum["symbol_instance_id"]


            if datum["symbol_name"] == "ESB alteration":
                json_val_dict["ESB alteration"] = 1
                # esb_alterations.append(datum["symbol_instance_id"])
            if datum["symbol_name"] == "GNI meter alteration":
                json_val_dict["GNI meter alteration"] = 1
                # gni_alterations.append(datum["symbol_instance_id"])
            
            if datum["symbol_name"] == "New Gas Connection":
                json_val_dict["New Gas Connection"] = 1
                # new_gas_connection.append(datum["symbol_instance_id"])
            if datum["symbol_name"] == "RGI Meter_No Heating":
                json_val_dict["RGI Meter_No Heating"] = 1
                # rgi_meter_no_heating.append(datum["symbol_instance_id"])
            
            for form in datum["forms"]:
                for section in form["sections"]:
                    for field in section["fields"]:
                        v = ''
                        if field["value"]["value"] == None:
                            vals = [val["value"] for val in field["value"]["values"]]
                            for val in vals:
                                v += val
                                v += '<BR>'
                        else:
                            v = field["value"]["value"]
                        # print(field["label"], ':', v)
                        json_val_dict[field["label"]] = v
                        
                        im = field["label"].replace(' *', '')
                        im = im.replace('*', '')
                        json_val_dict[im] = v # if the field is marked as mandatory this creates a non-marked copy with the same answer, note this doesn't solve all issues though
                        

                            
                        if field["label"] == "Is it a Balanced Flue?" and field["value"]["value"] == False:
                            balanced_flues.append(datum["symbol_instance_id"])
                        
                        if field["label"] == "Heating designation on Portal*" and field["value"]["value"] == "Primary":
                            json_val_dict['Heating System *'] = datum["symbol_name"]
                        if field["label"] == "Heating designation on Portal*" and field["value"]["value"] == "Secondary":
                            json_val_dict['Secondary Heating System'] = datum["symbol_name"]
                        if field["label"] == "Is there a timer?" and field["value"]["value"] == True:
                            json_val_dict['HWC Controls *'] = 'Independent Timer'
                        if field["label"] == "Is there a cylinder stat?" and field["value"]["value"] == True:
                            json_val_dict['HWC Controls *'] = 'Cylinder Thermostat'
                            cylinder_stat = True
                        if field["label"] == "Is the cylinder heated from the primary heating system?":
                            if field["value"]["value"] == True:
                                json_val_dict['From Primary heating system'] = True
                                if 'From Primary heating system' not in json_val_dict['HWS']:
                                    json_val_dict['HWS'] += ('From Primary heating system' + '<BR>')
                        if field["label"] == "Is the cylinder heated from the secondary heating system?":
                            if field["value"]["value"] == True:
                                json_val_dict['From Secondary heating system'] = True
                                if 'From Secondary heating system' not in json_val_dict['HWS']:
                                    json_val_dict['HWS'] += ('From Secondary heating system' + '<BR>')
                        if field["label"] == "Is there an electric immersion?" and field["value"]["value"] == True:
                            json_val_dict['Electric Immersion'] = True
                            if 'Electric Immersion' not in json_val_dict['HWS']:
                                json_val_dict['HWS'] += ('Electric Immersion' + '<BR>')
                        if field["label"] == "How is the cylinder heated? (Do not include immersion)" and field["value"]["has_value"] == True:
                            json_val_dict["Other HW Details *"] = field["value"]["value"]
                            if 'Other' not in json_val_dict['HWS']:
                                json_val_dict['HWS'] += ('Other' + '<BR>')
                            
                        if field['label'] == "Heating notes*":
                            if datum["symbol_name"] not in [json_val_dict['Heating System *'], json_val_dict['Secondary Heating System']]:
                                n = 'Other System: ' + field["value"]["value"] + '<BR>'
                                if n not in other_heating_notes:
                                    other_heating_notes += n



                        if field["label"] == "Existing Roof Ventilation (mm2)*":
                            if not field["value"]["value"].isdigit():
                                continue
                            json_val_dict["Existing (mm2)*"] += int(field["value"]["value"])
                        
                        if field["label"] == "Is the window Single glazed?" and field["value"]["value"] == True:
                            # json_val_dict['No. Single Glazed Windows *'] += 1
                            # single_glazed_windows.append(datum["symbol_instance_id"])
                            single_glazed_windows.append(datum["symbol_instance_id"])
                        if field["label"] == "Is it being recommended for replacement?" and field["value"]["value"] == True:
                            replace_windows.append(datum["symbol_instance_id"]) 
                        if datum["symbol_instance_id"] in xml_ref_dict.keys():
                            if field["label"] == "Roof Type*":
                                # print('before: ', xml_ref_dict[datum["symbol_instance_id"]])
                                xml_ref_dict[datum["symbol_instance_id"]] = field["value"]["value"]
                                # print('after: ', xml_ref_dict[datum["symbol_instance_id"]])
                                for n in range(1, 5):
                                    if field["value"]["value"] == f"Roof Type {n}":
                                        roof_type_dict[datum["symbol_instance_id"]] = n
                    g = {} # temporary dictionary containing the answers to all questions in a section (roof-related)
                    for field in section["fields"]:
                        g[field["label"]] = field["value"]["value"]
                    for n in range(1, 5):
                        if "Roof Type*" in g.keys() and f"Roof Type {n} Sloping Ceiling Suitable for Insulation*" in g.keys():
                            # print('g: ', g)
                            if g["Roof Type*"] == "Sloped Ceiling" and g[f"Roof Type {n} Sloping Ceiling Suitable for Insulation*"] == True:
                                if f"Roof {n} Pitch (degrees)*" in json_val_dict.keys():
                                    pitch = json_val_dict[f"Roof {n} Pitch (degrees)*"]
                                else:
                                    pitch = 30
                                slope_dict[datum["symbol_instance_id"]] = pitch
        # print('balanced_flues', ':', str(balanced_flues))

        print(1)


        json_val_dict['Notes (Heating)'] = ''
        # Go through Forms again to get values for Primary & Secondary Heating Systems
        for datum in JSON["data"]:
            if datum["symbol_name"] == json_val_dict['Heating System *']:
                print(datum["symbol_name"])
                for form in datum["forms"]:
                    for section in form["sections"]:
                        for field in section["fields"]:
                            print(field["label"])
                            if 'age (years)' in field['label']:
                                json_val_dict['System Age *'] = field["value"]["value"]
                            if field['label'] == 'Fully Working?':
                                json_val_dict['Fully Working *'] = field["value"]["value"]
                            if 'require service?' in field['label']:
                                json_val_dict['Requires Service *'] = field["value"]["value"]
                            if field['label'] == '':
                                json_val_dict["Other Primary Heating Details *"] = field["value"]["value"]
                            if field['label'] == "Not working details*":
                                # print(field["value"]["value"])
                                json_val_dict['Not Working Details Primary Heating *'] = field["value"]["value"]
                            # if field['label'] == 'Does the appliance require service?':
                                # json_val_dict['Requires Service (App?)*'] = field["value"]["value"]
                            if field['label'] == 'Service details':
                                json_val_dict['Requires Service Details Primary Heating *'] = field["value"]["value"]
                            if field['label'] == "Is the boiler Condensing?*" and field["value"]["value"] == False:
                                condensing = False
                            if field['label'] == "Interlinked with?" and field["value"]["value"] == "Stove + Back Boiler":
                                linked_stove_bb = True
                            if field['label'] == "Heating notes*" and field["value"]["value"] != None:
                                n = 'Primary System: ' + field["value"]["value"] + '<BR>'
                                print(n)
                                if n not in json_val_dict['Notes (Heating)']:
                                    json_val_dict['Notes (Heating)'] += n
                                    print('Notes (Heating)', ':', json_val_dict['Notes (Heating)'])
            # print(2)
            
            if datum["symbol_name"] == json_val_dict['Secondary Heating System']:
                for form in datum["forms"]:
                    for section in form["sections"]:
                        for field in section["fields"]:
                            if 'age (years)' in field['label']:
                                json_val_dict['Secondary System Age *'] = field["value"]["value"]
                            if field['label'] == 'Fully Working?':
                                json_val_dict['Secondary System Fully Working *'] = field["value"]["value"]
                            if 'require service?' in field['label']:
                                json_val_dict['Secondary System Requires Service *'] = field["value"]["value"]
                            if field['label'] == '':
                                json_val_dict["Other Primary Heating Details *"] = field["value"]["value"]
                            if field['label'] == 'Not working details*':
                                json_val_dict['Not Working Details Secondary Heating *'] = field["value"]["value"]
                            # if field['label'] == 'Does the appliance require service?':
                                # json_val_dict['Secondary System Requires Service (App?)*'] = field["value"]["value"]
                            if field['label'] == 'Service details':
                                json_val_dict['Requires Service Details Secondary Heating *'] = field["value"]["value"]
                            if field['label'] == "Heating notes*":
                                n = 'Secondary System: ' + field["value"]["value"] + '<BR>'
                                if n not in json_val_dict['Notes (Heating)']:
                                    json_val_dict['Notes (Heating)'] += n
                    
        print('Notes (Heating)', ':', json_val_dict['Notes (Heating)'])
        
        json_val_dict['Programmer / Timeclock *'] = 0
        json_val_dict['Room Thermostat Number *'] = 0
        json_val_dict['Rads Number *'] = 0
        json_val_dict['TRVs Number *'] = 0
        
        if other_heating_notes != '':
            json_val_dict['Notes (Heating)'] = json_val_dict['Notes (Heating)'] + other_heating_notes
        



        
        json_url = "https://cloud.magicplan.app/api/v2/plans/statistics/" + str(xml_val_dict['id'])
        request = urllib.request.Request(json_url, headers=headers)
        JSON = urllib.request.urlopen(request).read()
        JSON = json.loads(JSON)
        
        # df = pd.DataFrame(JSON["data"])
        
        df = pd.DataFrame(JSON["data"]["project_statistics"])
        # print()
        

        
        
        json_val_dict['No. Double Glazed Windows *'] = 0
        json_val_dict['Number of Storeys *'] = 0
        json_val_dict['Gross floor area (m2) *'] = 0.00
        new_draughtproofing = 0
        sum_low = 0
        sum_high = 0
        roof_area_sum = 0
        slope_roof_area_sum = 0
        new_hatch_count = 0
        
        json_val_dict["MEV 15l/s Bathroom"] = 0
        json_val_dict["MEV 30l/s Utility"] = 0
        json_val_dict["MEV 60l/s Kitchen"] = 0
        json_val_dict["New Permanent Vent"] = 0
        json_val_dict["New Background Vent"] = 0
        
        json_val_dict["Cavity Wall Insulation Bonded Bead"] = 0
        json_val_dict["Loose Fibre Extraction"] = 0
        external_wall_insulation = 0
        json_val_dict["External Wall Insulation: Less than 60m2"] = 0
        json_val_dict["External Wall Insulation: 60m2 to 85m2"] = 0
        json_val_dict["External Wall Insulation: Greater than 85m2"] = 0
        json_val_dict["Internal Wall Insulation: Vertical Surface"] = 0
        external_wall_insulation_and_cwi = 0
        json_val_dict["External wall insulation and CWI: less than 60m2"] = 0
        json_val_dict["External wall insulation and CWI: 60m2 to 85m2"] = 0
        json_val_dict["External wall insulation and CWI: greater than 85m2"] = 0

        # , 'Heating Systems Controls *'
        # , 'Partial Details *'
        
        
        rooms_with_balanced_flues = []
        req_lagging_jacket_count = 0
        
        
        json_val_dict['Thermal Envelope - Heat loss floor area'] = 0
        json_val_dict['replace_window_area'] = 0
        # print(xml_ref_dict)
        # print('replace_windows', ':', replace_windows)
        for floor in df.floors:
            if int(xml_ref_dict[floor["uid"]]) == 20:
                external_wall_insulation += floor["area_with_interior_walls_only"]
            if int(xml_ref_dict[floor["uid"]]) == 21:
                json_val_dict["Cavity Wall Insulation Bonded Bead"] += floor["area_with_interior_walls_only"]
            if int(xml_ref_dict[floor["uid"]]) == 22:
                json_val_dict["Internal Wall Insulation: Vertical Surface"] += floor["area_with_interior_walls_only"]
            if int(xml_ref_dict[floor["uid"]]) == 23:
                json_val_dict["Loose Fibre Extraction"] += floor["area_with_interior_walls_only"]
            if int(xml_ref_dict[floor["uid"]]) == 24:
                external_wall_insulation_and_cwi += floor["area_with_interior_walls_only"]
            if -1 <= int(xml_ref_dict[floor["uid"]]) <= 9:
                for room in floor["rooms"]:
                    
                    
                    for furniture in room["furnitures"]:
                        if furniture["uid"] in req_lagging_jackets:
                            # json_val_dict['Hot Water Cylinder Jacket'] += 1
                            req_lagging_jacket_count += 1
                        if furniture["name"] in ["Radiator", "Radiator with TRV", "Water Radiator"]:
                            json_val_dict['Rads Number *'] += 1
                        if furniture["name"] == "Radiator with TRV":
                            json_val_dict['TRVs Number *'] += 1
                        if furniture["name"] == "Electric Instantaneous":
                            json_val_dict['Electric Instantaneous'] = True
                        if furniture["name"] in ["Gas Combi Boiler", "Oil Combi Boiler"]:
                            json_val_dict['Instantaneous Combi Boiler'] = True
                    for wall_item in room["wall_items"]:
                        if wall_item["name"] == "Room Thermostat":
                            json_val_dict['Room Thermostat Number *'] += 1
                        if wall_item["name"] == "Programmer":
                            json_val_dict['Programmer / Timeclock *'] += 1
                            
                            
            
            
            
            
            if int(xml_ref_dict[floor["uid"]]) == 10:
                json_val_dict['Thermal Envelope - Heat loss floor area'] = floor["area_with_interior_walls_only"]
                for room in floor["rooms"]:
                    if room["uid"] in xml_ref_dict['exclude_rooms']:
                        json_val_dict['Thermal Envelope - Heat loss floor area'] -= room["area_with_interior_walls_only"]

            if -1 <= int(xml_ref_dict[floor["uid"]]) <= 9:
                json_val_dict['Number of Storeys *'] += 1
                json_val_dict['Gross floor area (m2) *'] += floor["area_with_interior_walls_only"]
                json_val_dict['No. Double Glazed Windows *'] += floor["window_count"]

                for room in floor["rooms"]:
                    # print(xml_ref_dict[room["uid"]])
                    if room["uid"] in xml_ref_dict['habitable_rooms']:
                        json_val_dict["Number of habitable rooms in the property"] += 1
                    if room["uid"] in xml_ref_dict['wet_rooms']:
                        json_val_dict["Number of wet rooms in the property"] += 1
                    if room["uid"] in xml_ref_dict['exclude_rooms']:
                        json_val_dict['Gross floor area (m2) *'] -= room["area_with_interior_walls_only"]
                        
                    for wall_item in room["wall_items"]:
                        if wall_item["uid"] in replace_windows:
                            json_val_dict['replace_window_area'] += (wall_item["width"] * wall_item["height"])
                        if wall_item["uid"] in single_glazed_windows:
                            # print(wall_item["uid"] + " found in " + str(single_glazed_windows))
                            json_val_dict['No. Single Glazed Windows *'] += 1
                            
                    for furniture in room["furnitures"]:
                        # print(furniture["name"])
                        if furniture["uid"] in balanced_flues:
                            rooms_with_balanced_flues.append(room["uid"])
                            
                        if furniture["name"] == "New Draughtproofing":
                            new_draughtproofing += 1
                        if furniture["name"] == "New Mechanical Vent":
                            # print(xml_ref_dict[room["uid"]])
                            if xml_ref_dict[room["uid"]] in ['Bathroom', 'Half Bathroom', 'Toilet']:
                                json_val_dict["MEV 15l/s Bathroom"] += 1
                            if xml_ref_dict[room["uid"]] in ['Laundry Room']:
                                json_val_dict["MEV 30l/s Utility"] += 1
                            if xml_ref_dict[room["uid"]] in ['Kitchen']:
                                json_val_dict["MEV 60l/s Kitchen"] += 1
                        if furniture["name"] == "New Permanent Vent":
                            json_val_dict["New Permanent Vent"] += 1
                        if furniture["name"] == "New Background Vent":
                            json_val_dict["New Background Vent"] += 1
                        if furniture["name"] == "Duct Cooker Hood":
                            json_val_dict["Duct Cooker Hood"] += 1
                        if furniture["name"] == "Duct Mechanical Extract Vent":
                            json_val_dict["Duct Cooker Hood"] += 1
                        if furniture["name"] == "New Hatch": # only found in -1 to 9 and Roof?
                            new_hatch_count += 1
                
                
            for room in floor["rooms"]:
                for furniture in room["furnitures"]:
                    if furniture["name"] == "New Low Level Roof Ventilation": # only found in Roof?
                        sum_low += float(furniture["width"])
                    if furniture["name"] == "New High Level Roof Ventilation": # only found in Roof?
                        sum_high += float(furniture["width"])
        
        
        json_val_dict['No. Double Glazed Windows *'] -= json_val_dict['No. Single Glazed Windows *']

        for room in rooms_with_balanced_flues:
            if room in (xml_ref_dict['habitable_rooms'] + xml_ref_dict['wet_rooms']):
                json_val_dict["No. of habitable/wet rooms w/ open flued appliance"] += 1
        


        
        if new_draughtproofing == 0:
            json_val_dict["Draught Proofing (<= 20m installed)"] = 'N/A'
            json_val_dict["Draught Proofing (> 20m installed)"] = 'N/A'
        if 1 <= new_draughtproofing <= 3:
            json_val_dict["Draught Proofing (<= 20m installed)"] = 1
            json_val_dict["Draught Proofing (> 20m installed)"] = 'N/A'
        if new_draughtproofing >= 4:
            json_val_dict["Draught Proofing (<= 20m installed)"] = 'N/A'
            json_val_dict["Draught Proofing (> 20m installed)"] = 1
        
        
        json_val_dict["External Wall Insulation: Less than 60m2"] = round(external_wall_insulation) if external_wall_insulation <= 60 else 'N/A'
        
        json_val_dict["External Wall Insulation: 60m2 to 85m2"] = round(external_wall_insulation) if 60 < external_wall_insulation <= 85  else 'N/A'
        
        json_val_dict["External Wall Insulation: Greater than 85m2"] = round(external_wall_insulation) if external_wall_insulation > 85  else 'N/A'


        json_val_dict["External wall insulation and CWI: less than 60m2"] = round(external_wall_insulation_and_cwi) if external_wall_insulation_and_cwi <= 60 else 'N/A'
        
        json_val_dict["External wall insulation and CWI: 60m2 to 85m2"] = round(external_wall_insulation_and_cwi) if 60 < external_wall_insulation_and_cwi <= 85 else 'N/A'
        
        json_val_dict["External wall insulation and CWI: greater than 85m2"] = round(external_wall_insulation_and_cwi) if external_wall_insulation_and_cwi > 85 else 'N/A'
            
        
        
        
        for floor in df.floors:
            if int(xml_ref_dict[floor["uid"]]) == 1000: # i.e. type "Roof"
                # roof_area_total = floor["area_with_interior_walls_only"]
                # print('roof_area_total (unused variable): ', roof_area_total)
                
                slope_roof_area_sum = 0
                for n in range(1, 5):
                    json_val_dict[f"roof_{n}_area"] = 0

                for room in floor["rooms"]:
                    # print('room["uid"]', ':', room["uid"])
                    # print(xml_ref_dict[room["uid"]])
                    # print('area:', room["area_with_interior_walls_only"])
                    
                    if room["uid"] in roof_type_dict.keys():
                        n = roof_type_dict[room["uid"]]
                        json_val_dict[f"roof_{n}_area"] += room["area_with_interior_walls_only"]
                    if room["uid"] in slope_dict.keys():
                        this_slope_area = room["area_with_interior_walls_only"] / cos(slope_dict[room["uid"]]/57.2958)
                        slope_roof_area_sum += this_slope_area
                    for furniture in room["furnitures"]:
                        if furniture["name"] == "New Hatch": # only found in -1 to 9 and Roof?
                            new_hatch_count += 1
        # print(json_val_dict)
        
        
        roof_general(json_val_dict) # adds a number of fields contingent on the above
        
        walls_general(json_val_dict)
        
        
        json_val_dict['Gross floor area (m2) *'] = round(json_val_dict['Gross floor area (m2) *'], 2)
        json_val_dict['Required per standards (mm2) *'] = round(sum_low * 10000)
        

        
        
        HSC_count = 0
        # Yes to Cylinder stat in form for hot water cylinder
        if cylinder_stat == True:
            HSC_count += 1
        # Object count of "Programmer" >0
        if json_val_dict['Programmer / Timeclock *'] > 0:
            HSC_count += 1
        # Object count of "Room Thermostat" >0
        if json_val_dict['Room Thermostat Number *'] > 0:
            HSC_count += 1
        # % of Radiators/Rads with TRVs >=50%
        if json_val_dict['TRVs Number *'] > 0:
            if json_val_dict['Rads Number *'] > 0: # should be redundant due to preceeding condition... (apparently not!)
                r = json_val_dict['TRVs Number *'] / json_val_dict['Rads Number *']
                if r >= 0.5:
                    HSC_count += 1
        else:
            r = 0
        # print('json_val_dict["TRVs Number *"]', ':', json_val_dict['TRVs Number *'])
        # print('json_val_dict["Rads Number *"]', ':', json_val_dict['Rads Number *'])
        # print('r', ':', str(r))
        # print('HSC_count', ':', str(HSC_count))
        
        if HSC_count == 0:
            json_val_dict['Heating Systems Controls *'] = 'No Controls'
        if 1 <= HSC_count <= 3:
            cylinder_stat_yn = "Yes" if cylinder_stat == True else "No"
            percentage = str(round(r * 100)) + '%'
            json_val_dict['Heating Systems Controls *'] = 'Partial Controls'
            json_val_dict["Partial Details *"] = 'No of Programmers: ' + str(json_val_dict['Programmer / Timeclock *']) + "<BR>" + 'No of Room Stats: ' + str(json_val_dict['Room Thermostat Number *']) + "<BR>" + '% of Radiators  with TRVs: ' + percentage + "<BR>" + 'Cylinder Stat?: ' + cylinder_stat_yn
        if HSC_count == 4:
            json_val_dict['Heating Systems Controls *'] = 'Full zone control to spec'
            
        
        
        print(3)
        
        
        
        # Work Order Recommendation (Roof):
        json_val_dict['sloped_surface_area'] = round(slope_roof_area_sum) if round(slope_roof_area_sum) != 0 else 'N/A'
        
        print('sfi_dict', ':', json_val_dict["sfi_dict"])
        json_val_dict['Attic Storage (5m2)'] = 0
        for t in [100, 150, 200, 250, 300]:
            if str(t) in json_val_dict["sfi_dict"].keys():
                json_val_dict[f'ins_{t}_area'] = round(json_val_dict["sfi_dict"][str(t)])
                json_val_dict['Attic Storage (5m2)'] = 1
        
        
        json_val_dict['new_hatch_count'] = new_hatch_count
        json_val_dict['high_roof_vent_area'] = round(sum_high * 5000)
        # json_val_dict['low_roof_vent_area'] = json_val_dict['Required per standards (mm2) *']
        
        for n in range(1, 5):
            if f"Wall Type {n}" in json_val_dict.keys():
                json_val_dict[f"Wall Type {n}"] = json_val_dict[f"Wall Type {n}"] if json_val_dict[f"Wall Type {n}"] != '' else 'N/A'
            if f"Wall Type {n} Residual Cavity Width (mm)" in json_val_dict.keys():
                json_val_dict[f"Wall Type {n} Residual Cavity Width (mm)"] = json_val_dict[f"Wall Type {n} Residual Cavity Width (mm)"] if json_val_dict[f"Wall Type {n} Residual Cavity Width (mm)"] != 0 else 'N/A'
            if f"Wall Type {n} Fill Type" in json_val_dict.keys():
                json_val_dict[f"Wall Type {n} Fill Type"] = json_val_dict[f"Wall Type {n} Fill Type"] if json_val_dict[f"Wall Type {n} Fill Type"] != 0 else 'N/A'
        
        
        
        
        # Fixed Values: (should these only be added to output_dict as they are not JSON values?)
        json_val_dict['Roof 2 Required per standards (mm2) *'] = 0
        json_val_dict['Roof 2 Existing (mm2) *'] = 0
        json_val_dict['Roof 3 Required per standards (mm2) *'] = 0
        json_val_dict['Roof 3 Existing (mm2) *'] = 0
        json_val_dict['Roof 4 Required per standards (mm2) *'] = 0
        json_val_dict['Roof 4 Existing (mm2) *'] = 0
        
        
        wt_dict['EWI/IWI'] = round(wt_dict['total'] - wt_dict['External.Wall.not.recieving.EWI.or.IWI'])
        
        # (if any value blank then 0)
        json_val_dict['Thermal Envelope - Heat loss walls, windows and doors'] = round(wt_dict['gross'], 2) # round(wt_dict['ext_wall_area_net'], 2)
        json_val_dict['Thermal Envelope - Heat loss floor area'] = round(json_val_dict['Thermal Envelope - Heat loss floor area'], 2)
        json_val_dict['Thermal Envelope - Heat loss roof area'] = json_val_dict['Thermal Envelope - Heat loss floor area']
        json_val_dict['Heat loss Wall Area recommended for EWI and IWI'] = round(wt_dict['EWI/IWI'], 2) # round(wt_dict['total'], 2)
        json_val_dict['New Windows being recommended for replacement'] = round(json_val_dict['replace_window_area'], 2) if json_val_dict['replace_window_area'] > 0.5 else 0
        json_val_dict['Total Surface Area (m2)'] = round(json_val_dict['Thermal Envelope - Heat loss walls, windows and doors'] + (2 * json_val_dict['Thermal Envelope - Heat loss floor area']), 2)
        json_val_dict['Total Surface Area receiving EWWR (m2)'] = round(float(wt_dict['EWI/IWI']) + float(json_val_dict['replace_window_area']), 2)
        json_val_dict['Result %'] = round(100 * (json_val_dict['Total Surface Area receiving EWWR (m2)'] / json_val_dict['Total Surface Area (m2)']), 2) if json_val_dict['Total Surface Area (m2)'] > 0 else 0
        json_val_dict['Is Major Renovation?'] = 'Yes' if json_val_dict['Result %'] >= 23 else 'No'
        
        json_val_dict['THERMAL ENVELOPE OF BUILDING AREA'] = json_val_dict['Total Surface Area (m2)']
        json_val_dict['TOTAL SURFACE AREA FOR MAJOR RENOVATION WORKS'] = json_val_dict['Total Surface Area receiving EWWR (m2)']
        json_val_dict['Total surface area for MR works / Thermal Envelope'] = json_val_dict['Result %']
        json_val_dict['WARMER HOMES MAJOR RENOVATION RESULT'] = json_val_dict['Is Major Renovation?']
        
        
        
        json_val_dict['EWI/IWI > 25% *'] = json_val_dict['Is Major Renovation?']
        
        

        
        json_val_dict["ESB alteration"] = json_val_dict["ESB alteration"] if json_val_dict["ESB alteration"] != 0 else ''
        json_val_dict["GNI meter alteration"] = json_val_dict["GNI meter alteration"] if json_val_dict["GNI meter alteration"] != 0 else ''
        
        
        

        # print(1)
        
        
        
        
        for pm in ofl_pm:
            # print(pm)
            if pm not in json_val_dict.keys():
                json_val_dict[pm] = ''
            # print('json_val_dict[pm]', ':', json_val_dict[pm])
        
        
        # print('ins_200_area', ':', ins_200_area)
        # print("json_val_dict['ins_200_area']", ':', json_val_dict['ins_200_area'])
        
        json_val_dict['Internal Wall Insulation: Sloped or flat (horizontal) surface'] = json_val_dict['sloped_surface_area']
        if 'ins_100_area' in json_val_dict.keys():
            json_val_dict['Attic (Loft) Insulation 100 mm top-up'] = json_val_dict['ins_100_area']
        if 'ins_150_area' in json_val_dict.keys():
            json_val_dict['Attic (Loft) Insulation 150 mm top-up'] = json_val_dict['ins_150_area']
        if 'ins_200_area' in json_val_dict.keys():
            json_val_dict['Attic (Loft) Insulation 200 mm top-up'] = json_val_dict['ins_200_area']
        if 'ins_250_area' in json_val_dict.keys():
            json_val_dict['Attic (Loft) Insulation 250 mm top up'] = json_val_dict['ins_250_area']
        if 'ins_300_area' in json_val_dict.keys():
            json_val_dict['Attic (Loft) Insulation 300 mm'] = json_val_dict['ins_300_area']
        
        
        # json_val_dict['Cavity Wall Insulation Bonded Bead']
        # json_val_dict['Loose Fibre Extraction']
        # json_val_dict['External Wall Insulation: Less than 60m2']
        # json_val_dict['External Wall Insulation: 60m2 to 85m2']
        # json_val_dict['External Wall Insulation: Greater than 85m2']
        # json_val_dict['Internal Wall Insulation: Vertical Surface']
        # json_val_dict['External wall insulation and CWI: less than 60m2']
        # json_val_dict['External wall insulation and CWI: 60m2 to 85m2']
        # json_val_dict['External wall insulation and CWI: greater than 85m2']
        json_val_dict['Basic gas heating system'] = ''
        json_val_dict['Basic oil heating system'] = ''
        json_val_dict['Full gas heating system installation'] = ''
        json_val_dict['Full oil heating system installation'] = ''
        json_val_dict['Gas boiler and controls (Basic & controls pack)'] = ''
        json_val_dict['Oil boiler and controls (Basic & controls pack)'] = ''
        
        
        
        # print('req_lagging_jackets', ':', req_lagging_jackets)
        # print('req_lagging_jacket_count', ':', req_lagging_jacket_count)
        
        for pm in ofl_pm:
            # print(json_val_dict[pm])
            if str(json_val_dict[pm]) not in ['', '0', 'N/A']: # if any primary measure has any valid value
                json_val_dict["LED Bulbs: supply only (4 no.)"] = 1
                json_val_dict["Hot Water Cylinder Jacket"] = req_lagging_jacket_count
        
        # print("Hot Water Cylinder Jacket", ':', json_val_dict["Hot Water Cylinder Jacket"])
        # print('req_lagging_jacket_count', ':', req_lagging_jacket_count)
        
        
        # print(external_wall_insulation)

        # print(json_val_dict["Internal Wall Insulation: Vertical Surface"])
        
        # print('sum of Ex/In: ', float(external_wall_insulation) + float(json_val_dict["Internal Wall Insulation: Vertical Surface"]))
        json_val_dict["Air-tightness test recommended?"] = 1 if float(external_wall_insulation) + float(json_val_dict["Internal Wall Insulation: Vertical Surface"]) > 0 else ''
        

        json_val_dict["Cavity Wall Insulation Bonded Bead"] = round(json_val_dict["Cavity Wall Insulation Bonded Bead"]) if json_val_dict["Cavity Wall Insulation Bonded Bead"] != 0 else 'N/A'
        json_val_dict["Loose Fibre Extraction"] = round(json_val_dict["Loose Fibre Extraction"]) if json_val_dict["Loose Fibre Extraction"] != 0 else 'N/A'
        json_val_dict["Internal Wall Insulation: Vertical Surface"] = round(json_val_dict["Internal Wall Insulation: Vertical Surface"]) if json_val_dict["Internal Wall Insulation: Vertical Surface"] != 0 else 'N/A'
        json_val_dict['replace_window_area'] = round(json_val_dict['replace_window_area']) if json_val_dict['replace_window_area'] != 0 else 'N/A'
        json_val_dict['replace_window_area'] = 1 if json_val_dict['replace_window_area'] == 0 else json_val_dict['replace_window_area']
        # json_val_dict['Notes (Windows and Doors)'] = json_val_dict['Notes (Windows and Doors)'] if json_val_dict['Notes (Windows and Doors)'] != '' else 'N/A'
        # json_val_dict['No. Double Glazed Windows *'] = json_val_dict['No. Double Glazed Windows *'] - json_val_dict['No. Single Glazed Windows *']
        
        
        
        
        
        warnings = 'Major Renovation Error:'
        # print("Is a Major Renovation calculation necessary?*", ':', json_val_dict["Is a Major Renovation calculation necessary?*"])
        if "Is a Major Renovation calculation necessary?*" in json_val_dict.keys():
            if json_val_dict["Is a Major Renovation calculation necessary?*"] in [True, "Yes"]:
                # print(json_val_dict['Thermal Envelope - Heat loss walls, windows and doors'])
                # print(json_val_dict['Thermal Envelope - Heat loss floor area'])
                # print(json_val_dict['Heat loss Wall Area recommended for EWI and IWI'])
                if 0 in [json_val_dict['Thermal Envelope - Heat loss walls, windows and doors'], json_val_dict['Thermal Envelope - Heat loss floor area'], json_val_dict['Heat loss Wall Area recommended for EWI and IWI']]:
                    warnings = warnings + "<BR>" + 'Major Renovation calculation has been confirmed as necessary but not provided. The Values for "EWI/IWI >25%", "Qualifying Boiler", Heating Recommendations and the Lot Type may be incorrect as the required information has not been provided. Please resubmit to include Major Renovation calculations on Heat Loss 10th floor, 11th floor etc. and include values for Walls receiving EWI or IWI and External Walls not receiving EWI or IWI and confirm the response to the form question "Is there Mains Gas in the Area?"'
                    json_val_dict["EWI/IWI > 25% *"] = "No"

                # for q in ['Thermal Envelope - Heat loss walls, windows and doors', 'Thermal Envelope - Heat loss floor area', 'Heat loss Wall Area recommended for EWI and IWI']:
                    # if json_val_dict[q] == 0:
                        # warnings = warnings + "<BR>" + "Setting 'EWI/IWI > 25% *' to 'No', since Major Renovation calculation has been confirmed as necessary but zero value provided for '" + q + "'"
                        # json_val_dict["EWI/IWI > 25% *"] = "No"
            
            
            
        print(warnings)
            
            
        if "Is a Major Renovation calculation necessary?*" in json_val_dict.keys():
            if json_val_dict["Is a Major Renovation calculation necessary?*"] == False:
                # print(json_val_dict["Reason Major Renovation calculation is not necessary?*"])
                if json_val_dict["Reason Major Renovation calculation is not necessary?*"] == "The proportions of EWI/IWI are significantly greater than 25%":
                    json_val_dict["EWI/IWI > 25% *"] = "Yes"
                    # if json_val_dict['Qualifying Boiler'] == 'N/A':
                        # warnings = warnings + "<BR>" + "Qualifying Boiler question must be answered Yes/No"
                    
        
        if warnings == 'Major Renovation Error:':
            warnings = ''
        
        
        json_val_dict['Qualifying Boiler'] = False
        
        if json_val_dict['EWI/IWI > 25% *'] == 'No':
            json_val_dict['Qualifying Boiler'] = 'N/A'
        else:
            if condensing == False:
                json_val_dict['Qualifying Boiler'] = True
                if linked_stove_bb == True:
                    json_val_dict['Qualifying Boiler'] = False
        
        
        
        print(4)
        
        json_val_dict['Suitable for Heating Measures *'] = False
        
        if json_val_dict['Qualifying Boiler'] == True:
            json_val_dict['Suitable for Heating Measures *'] = True
        
        # print(json_val_dict['Suitable for Heating Measures *'])
        # print("Is there Mains Gas in the Area?", ':', json_val_dict["Is there Mains Gas in the Area?"])
        
        
        if json_val_dict['Suitable for Heating Measures *'] == False:
            json_val_dict['Not suitable details*'] = json_val_dict['Notes (Heating)']
        
        # print(json_val_dict["Is there Mains Gas in the Area?"])
        if json_val_dict["Is there Mains Gas in the Area?"] != "N/A - No heating measured being recommended":
            if json_val_dict['Qualifying Boiler'] == True:
                if json_val_dict['Heating Systems Controls *'] == 'Full zone control to spec':
                    if json_val_dict["Is there Mains Gas in the Area?"] == "Yes":
                        json_val_dict['Basic gas heating system'] = 1
                    else:
                        json_val_dict['Basic oil heating system'] = 1
            
            # print("Electric Storage Heater age (years)*", ':', json_val_dict["Electric Storage Heater age (years)*"])
            # print("Warm Air System age (years)*", ':', json_val_dict["Warm Air System age (years)*"])
            # print('Heating System *', ':', json_val_dict['Heating System *'])
            
            if (json_val_dict['Heating System *'] in ["Electric Storage Heater", "Warm Air System"] and json_val_dict['System Age *'] == "25+") or json_val_dict['Heating System *'] in ["Open Fire with Back Boiler", "Open Fire with Back Boiler With Enclosure Door", "Solid Fuel Range", "Solid Fuel Range with Back Boiler", "Fixed Electric Inset Fire", "Gas Fire Room Heater", "Oil Stove Room Heater", "Open Fire", "Open Fire With Enclosure Door"]:
                if json_val_dict["Is there Mains Gas in the Area?"] == "Yes":
                    json_val_dict['Full gas heating system installation'] = 1
                else:
                    json_val_dict['Full oil heating system installation'] = 1
            
            if json_val_dict['Qualifying Boiler'] == True:
                if json_val_dict['Heating Systems Controls *'] != 'Full zone control to spec':
                    if json_val_dict["Is there Mains Gas in the Area?"] == "Yes":
                        json_val_dict['Gas boiler and controls (Basic & controls pack)'] = 1
                    else:
                        json_val_dict['Oil boiler and controls (Basic & controls pack)'] = 1
            
        # print(2)
        # print("Hot Water Cylinder Jacket", ':', json_val_dict["Hot Water Cylinder Jacket"])
        for field in ofl_hpm:
            if json_val_dict[field] != '' and field != 'Hot Water Cylinder Jacket':
                print(field, ':', json_val_dict[field])
                json_val_dict['Hot Water Cylinder Jacket'] = ''
        # print("Hot Water Cylinder Jacket", ':', json_val_dict["Hot Water Cylinder Jacket"])
        # print(3)
        json_val_dict['Permanent ventilation wall vent (Certified Proprietary Integrated System)'] = json_val_dict["New Permanent Vent"]
        json_val_dict['Background ventilation wall vent (Certified Proprietary Integrated System)'] = json_val_dict["New Background Vent"]
        json_val_dict['Ducting existing cooker hood to exterior'] = json_val_dict['Duct Cooker Hood']
        json_val_dict['Window (same m2 rate will apply to windows with certified trickle vents)'] = int(round(json_val_dict['New Windows being recommended for replacement'], 0))
        json_val_dict['GNI new connection'] = json_val_dict['New Gas Connection']
        json_val_dict['Additional Roof Ventilation (Low Level)'] = json_val_dict['Required per standards (mm2) *'] - json_val_dict["Existing (mm2)*"]
        json_val_dict['Additional Roof Ventilation (High Level)'] = json_val_dict['high_roof_vent_area']
        
        # print('Window (same m2 rate will apply to windows with certified trickle vents)', ':', json_val_dict['Window (same m2 rate will apply to windows with certified trickle vents)'])

        
        # xl_2_pdf(xl_path)
        # print(json_val_dict)
        output_dict = json_val_dict
        
        for field in output_dict:
            value = output_dict[field]
            if (type(value) == bool and value == True):
                output_dict[field] = "Yes"
            if (type(value) == bool and value == False):
                output_dict[field] = "No"
        
        # xl_path = 
        # populate_template(output_dict)
        
        
        # Below all N/A unless type = other        
        
        missing = ['Thermal Envelope - Heat loss walls, windows and doors'
                , 'Other Details Roof 1 *'
                , 'Other Details Roof 2 *'
                , 'Other Details Roof 3 *'
                , 'Other Details Roof 4 *'
                , 'Other W1 Details *'
                , 'Other W2 Details *'
                , 'Other W3 Details *'
                , 'Other W4 Details *'

                , 'Internal Wall Insulation: Sloped or flat (horizontal) surface'
                , 'Attic (Loft) Insulation 100 mm top-up'
                , 'Attic (Loft) Insulation 150 mm top-up'
                , 'Attic (Loft) Insulation 200 mm top-up'
                , 'Attic (Loft) Insulation 250 mm top up'
                , 'Attic (Loft) Insulation 300 mm'
                
                , 'Attic Storage (5m2)' # "1" if any of the 5 above (not sloped) is non-zero
                
                , 'Additional Roof Ventilation (High Level)' # high_roof_vent_area
                , 'Additional Roof Ventilation (Low Level)' # 'Required per standards (mm2) *' - 'Existing (mm2)*'
                
                # , 'Roof 2 Insulation Type'
                , 'Wall 1 Residual Cavity Width (mm)*'
                , 'Wall 2 Residual Cavity Width (mm)*'
                , 'Wall 3 Residual Cavity Width (mm)*'
                , 'Wall 4 Residual Cavity Width (mm)*'

                # , 'Insulation *'
                # , 'Condition of Lagging Jacket *'
                , 'Partial Details *'
                
                ]
        
        print('missing:')
        for m in missing:
            if m not in json_val_dict.keys():
                print(m, 'not in json_val_dict.keys')
            else:
                print(m, ':', json_val_dict[m])
        
        
        
        
        # print(output_dict)
        output_dict['Lot *'] = 'Lot ' + lot(output_dict)
        print(output_dict['plan_name'], 'Lot *', ':', output_dict['Lot *'])
        
        # print('Not Working Details Primary Heating *', ':', json_val_dict['Not Working Details Primary Heating *'])
        
        populate_template_new(output_dict, 'template')
        
                
        # print(ofl_wos)
        
        # ofl_wos_2 = ofl_wos
        
        for f in ofl_wos_2:
            # print('f', ':', f)
            if f.isupper():
                continue
            else:
                if f in json_val_dict.keys():
                    # print(f, 'found in json_val_dict:', "'" + str(json_val_dict[f]) + "'")
                    if json_val_dict[f] in ['', 'N/A', None, 0]:
                        ofl_wos.remove(f)
                        # print('removed ' + f)
                else:
                    # print('removing ', f)
                    ofl_wos.remove(f)
        output = ''
        # json_val_dict["Is a Major Renovation calculation necessary?*"] = True
        # print('json_val_dict["Is a Major Renovation calculation necessary?*"]', ':', json_val_dict["Is a Major Renovation calculation necessary?*"])
        if "Is a Major Renovation calculation necessary?*" in json_val_dict.keys():
            if json_val_dict["Is a Major Renovation calculation necessary?*"] == "Yes":
                print('generating template_mrc')
                output, filename = populate_template_new(output_dict, 'template_mrc')
                
                ofl_filelist.append(filename)
                # ofl_filelist.append(filename)
                # print(ofl_filelist)
        

        
        
        # print(output)
        # print('Attic (Loft) Insulation 200 mm top-up', ':', json_val_dict['Attic (Loft) Insulation 200 mm top-up'])
        
        # print("Hot Water Cylinder Jacket", ':', json_val_dict["Hot Water Cylinder Jacket"])
        if output == '':
            styling = "border=\"1\""
            output = f"""\
                <h1>Work Order Summary</h1> \
                {create_table_text(output_dict, headers = ['Measure Item Name', 'Total Quantity'], styling=styling, do_not_sum=['All'], order_list = ofl_wos)} \
                <h1>General</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_general)} \
                <h1>Major Renovation</h1> \
                {create_table_text(output_dict, headers = ['Building Thermal Envelope', 'm2'], styling=styling, do_not_sum=['All'], order_list = ofl_mr)} \
                <h1>Primary Measures</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_pm)} \
                <h1>Roof</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_roof)} \
                <h1>Walls</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_walls)} \
                <h1>Heating</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_heating)} \
                <h1>Mechanical Ventilation Systems, Air Tightness Testing & Energy</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_mae)} \
                <h1>Supplementary</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_s)} \
                <h1>Heating Primary Measures</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_hpm)} \
                <h1>File List</h1> \
                {create_table_text(output_dict, headers = ['name', 'value'], styling=styling, do_not_sum=['All'], order_list = ofl_filelist)} \

                </div>"""


            if warnings != '':
                output = f"""\
                    <h1>Warnings</h1> \
                    {warnings} \
                    </div>""" + output

        # print(output)

    except Exception as ex:
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # output = "Line " + str(exc_tb.tb_lineno) + ": " + exc_type 
        
        # populate_template(xml_val_dict) # adds an (almost) empty copy of the template to avoid potential Logic App error if file not found
        
        populate_template_new(xml_val_dict, 'template')
        # populate_template_new(output_dict, 'template_mrc')
        
        # output = str(ex)
        output = traceback.format_exc()
        # LOGGER.info('Exception : ' + str(traceback.format_exc()))
        
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print(exc_type, fname, exc_tb.tb_lineno)
    finally:
        return output
    return output


def no_2_alph(no):
    alph = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    return alph[no - 1]


def Azure_2_Local(file_name
                , account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
                , default_credential = DefaultAzureCredential()
                , container_from = 'attachment'
                , local_dir = "/tmp"
                ):
    '''
    Read the contents of a file from Azure Blob Storage to the local environment and return the local file path
    '''
    
    try:
        instance_fp = os.path.join(os.getcwd(), local_dir, file_name)
        print(instance_fp)
        
        # Create the BlobServiceClient object
        blob_service_client = BlobServiceClient(account_url, credential=default_credential)
        container_client = blob_service_client.get_container_client(container= container_from) 
        
        with open(file=instance_fp, mode="wb") as download_file:
            download_file.write(container_client.download_blob(file_name).readall())
            
        output = instance_fp
    
    except:
        output = traceback.format_exc()
        print(output)
        
    return output



def XL_2_dict_new(xl_file_path):
    '''
    Read the contents of an Excel Workbook and return as a dictionary
    
    Three types of worksheet expected format: [trivial, table, lookup_table]
    
    trivial:
        2 columns, B & C, field_name and field_req
        field_req is one of the following four : [exact text, variable name, count, logic_dict]
        
    table: 
        need to update populate_template_new() to handle multicol outputs
        need to establish how many columns?
        done properly this will also cover trivial above
    
    lookup_table:
        2 cols
        create lookup_table_dict
        lookup is: field_value = lookup_table_dict[col_1]?
    
    '''
    
    try:
        
        wb = openpyxl.load_workbook(xl_file_path, data_only = True)
        
        output_tables = [
                        "1. Survey Details P1"
                        , "7. Thermal Mass P1"
                        , "8. Ventilation P1"
                        , "9. Space Heating P4"
                        , "9.4 Heating System Controls"
                        , "9.4 Heating System Controls (2)"
                        , "9.4 Pumps and Fans"
                        , "11. Lighting P1"
                        ]
        
        
        lookup_tables = ['1.1 Assessor Details Table'
                            , '2.2 Referance Table S8'
                            , '3.1 Referance Table S5'
                            , '3.2 Referance Table S4'
                            , '3.3 Referance Table Non Default'
                            , '7.1 Referance Table'
                            , 'lookup Age Band'
                            , 'lookup Assumed Roof U-Value'
                            ]
       
        
        multicol_tables = [ # use different name for these?
                            '2 Building Average Storey'
                            , '2 Building Average Storey (Floors)'
                            # , '2 Building Average Storey (Rooms)'
                            , '2.3 Floor Schedule Table'
                            , '3.4 Roof Type Schedule Table'
                            , '3.5 Roof Type Summary Table'
                            # , '4.1 Wall Schedule Table'
                            , '4.3 Wall Summary Table'
                            , '5.1 Windows Summary Table'
                            , '5.2 Window Schedule Table'
                            # , '5.3 Building | Doors P1'
                            , '5.4 Door Summary Table'
                            , '5.5 Door Schedule Table'
                            , '6. Colour Area Table P1'
                            , '8.1 Attic Hatches'
                            , '8.2 Ventilation Items'
                            

                            #, '10. Water Heating P3'
                            #, '10.3 Showers and Baths Table'
                            #, '10.4 Solar Thermal Table'
                            #, '12. Renewables P5'
                            #, '12.1 Renewables Table'
                            #, '9. Space Heating'
                            #, '9.1 Space Heating Schedule'
                            #, '9.2 Space Heating Category'
                            #, '9.5 Pumps and Fans'

                            
                            
                            
                            , '9.3 Space Heating Category'
                            , '11.1 Lighting Schedule'
                            ]
        
        output = {}
        lookup = {}
        print('XL_2_dict_new()', ':', 'sheets present in input file', ':')
        for sheet in wb.worksheets:
            print(sheet.title)
            
            
            if sheet.title in output_tables:
                output[sheet.title] = {}
                # print('output_table', ':', sheet.title)
                for i, row in enumerate(list(sheet.values)):
                    field_name = row[1]
                    # print(field_name)
                    if field_name == None:
                        continue
                    field_name = field_name.strip()
                    # print('field_name', ':', field_name)
                    
                    field_req = row[2]
                    if field_req == None:
                        continue
                    field_req = field_req.strip()
                    field_loc = no_2_alph(2) + str(i)
                    default_val = ''
                    if len(row) >= 5 and row[4] != None:
                        # print('row: ', str(i), 'col 4:', row[4])
                        default_val = eval(row[4])
                    # print('field_loc', ':', field_loc)
                    output[sheet.title][field_name] = {"field_req": field_req, "field_loc": field_loc, "default_val": default_val}
                
            elif sheet.title in multicol_tables:
                headers = []
                output[sheet.title] = {}
                # print('output_table', ':', sheet.title)
                # Do we need to take the headers?
                # if sheet.title in ['5.1 Windows Summary Table', '5.4 Door Summary Table']:
                    # headers = ['key']
                

                # ************ INCLUDE HEADERS FROM INPUT FILE ************
                
                headers = headers + [cell.value for cell in sheet[3]]
                
                # ************************         ************************
                
                
                if sheet.title in ['6. Colour Area Table P1']:
                    headers = ['colour', 'floor: area']
                
                if sheet.title in ['8.1 Attic Hatches']:
                    headers = [
                            'Room'
                            , 'Type'
                            , 'Count'
                            , 'Description'
                            ]
                
                if sheet.title in ['8.2 Ventilation Items']:
                    headers = [
                            'Room'
                            , 'Type'
                            , 'Count'
                            , 'Description'
                            ]
                
                
                
                if sheet.title in ['11.1 Lighting Schedule']:
                    headers = [
                            # 'uid'
                            'key'
                            , 'name'
                            , 'room_name'
                            # , 'floor_name'
                            , 'Count'
                            # , 'Type'
                            , 'Description'
                            # , 'Bulb Type'
                            # , 'Efficiency [lm/W]'
                            ]
                
                if sheet.title in ['5.2 Window Schedule Table']:
                    headers = [
                            'uid'
                            , 'floor_name'
                            , 'Room'
                            , 'Type'
                            , 'Construction'
                            , 'Description'
                            , 'No. of opes'
                            , 'No. of opes draught- stripped'
                            , 'In roof'
                            , 'Over shading'
                            , 'Orientation'
                            , 'cardinal_direction'
                            , 'Window height (m)'
                            , 'Window width (m)'
                            , 'Area [m2]'
                            , 'U-Value [W/m2K]'
                            # , 'Is the Window Shading Estimated to be Average or unknown (20% - 60% of sky blocked by obstacles)?'
                            # , 'Are the Number of Window Openings required to be calculated?'
                            # , 'Number of Window Openings'
                            # , 'Are all of the Window Openings Draught-stripped?'
                            ]
                
                if sheet.title in ['5.5 Door Schedule Table']:
                    headers = [
                            'uid'
                            , 'Room'
                            , 'Type'
                            , 'Description'
                            , 'Draught Stripped'
                            , 'Door Area [m2]'
                            , 'U-Value [W/m2K]'
                            , 'floor_name'
                            , 'Glazing Area (m²)'
                            , 'Glazing Type'
                            , 'Number of openings'
                            , 'Number of openings draughtstripped'
                            , 'Door height (m)'
                            , 'Door width (m)'
                            ]
                
                
                
                
                if sheet.title in ['2.3 Floor Schedule Table']:
                    headers = [
                            'uid'
                            , 'floor_name'
                            , 'True Floor'
                            , 'Floor Type'
                            , 'description'
                            , 'underfloor heating?'
                            , 'age band'
                            , 'perimeter'
                            , 'area (m2)'
                            , 'U-value calculation required?'
                            , 'calculated U-value'
                            , 'U-Value'
                            
                            # , 'dwelling age band?'
                            ]
                
                if sheet.title in ['3.4 Roof Type Schedule Table']:
                    headers = [
                                'uid'
                                ,  'roof type'
                                ,  'description'
                                ,  'insulation thickness (mm)'
                                ,  'age band'
                                ,  'area (m2)'
                                ,  'calculated U-value'
                                ,  'thermal conductivity'
                                ,  'roof pitch (degrees)'
                                ,  'insulation thickness known?'
                                
                                # ,  'name'
                                # ,  'dwelling age band?'
                                # ,  'U-value calculation required?'
                                # ,  'U-value'
                                ]
                
                
                if sheet.title in ['3.5 Roof Type Summary Table']:
                    headers = [
                                'uid'
                                ,  'roof type'
                                ,  'description'
                                ,  'insulation thickness (mm)'
                                ,  'age band'
                                # ,  'area (m2)'
                                ,  'U-value'
                                ,  'Total roof type Area (m²)'
                                ]
                
                
                if sheet.title in ['4.3 Wall Summary Table']:
                    headers = [
                                'sku'
                                , 'total_surface'
                                , 'wall type'
                                , 'description'
                                , 'Is semi exposed'
                                , 'age band'
                                , 'other U-value'
                                , 'U-value'
                                , 'insulation thickness'
                                , 'thermal conductivity (λ)'
                                
                                # , 'other wall type'
                                # , 'dwelling age band?'
                                # , 'insulation'
                                # , 'explanation'
                                # , 'default U-value?'
                                # , 'can u-value be substantiated?'
                                # , 'can thermal conductivity be substantiated?'
                                ]
                
                
                if sheet.title in ['5.1 Windows Summary Table']:
                    headers = [
                                'key'
                                , 'Count of the common windows'
                                , 'Type'
                                , 'Description'
                                , 'Construction'
                                , 'No. of opes'
                                , 'No. of opes draught- stripped'
                                , 'In roof'
                                , 'Over shading'
                                , 'cardinal_direction'
                                , 'Area [m2]'
                                , 'U-Value [W/m2K]'
                                , 'Orientation'
                                ]
                
                if sheet.title in ['5.4 Door Summary Table']:
                    headers = [
                                'key'
                                , 'Count'
                                , 'Type'
                                , 'Door Description'
                                , 'Draught Stripped'
                                
                                # , 'Number of openings'
                                # , 'Number of openings draught stripped'
                                # , 'Glazing Area (m²)'
                                # , 'Glazing Area (%)'
                                # , 'Glazing Type'
                                # , 'Door U-Value (Wm2K)'
                                # , 'Total Door Area (m²)'
                                # , 'Door Type'
                                ]
                
                #, '10. Water Heating P3'
                #, '10.3 Showers and Baths Table'
                #, '10.4 Solar Thermal Table'
                #, '12. Renewables P5'
                #, '12.1 Renewables Table'
                #, '9. Space Heating'
                #, '9.1 Space Heating Schedule'
                #, '9.2 Space Heating Category'
                #, '9.5 Pumps and Fans'
                            
                
                # if sheet.title in ['2 Building Average Storey' , '2 Building Average Storey (Floors)', '2 Building Average Storey (Rooms)']:
                
                # if sheet.title in ['2 Building Average Storey (Floors)']: # does this sheet title even exist yet? Maybe none of this should be here
                    # headers = ['uid'
                            # , 'number'
                            # ,  'name'
                            # ,  'use_floor_level_height'
                            # ,  'ceiling_height'
                            # ,  'thermal_envelope'
                            # , 'Living Area (m2)'
                            # ]
                    
                    
                    
                # if 'Building Average Storey (Rooms' in sheet.title:
                    # headers = ['uid'
                            # ,  'ceiling_height'
                            # ,  'name'
                            # ,  'floor_type'
                            # ,  'use_floor_level_height'
                            # ,  'volume'
                            # ,  'room_type'
                            # ,  'thermal_envelope']
                
                
                if sheet.title in ['9.3 Space Heating Category']:
                    headers = ["uid"
                                , "Object Name"
                                # ,"Primary/Secondary/Neither/Fuel Cost Comparson Required"
                                ,"Heat Source Type on DEAP"
                                ,"Manufacturer"
                                ,"Model"
                                ,"Ventilation Type"
                                # ,"Heats Water?"
                                ,"Does the heat source heat water?"
                                ,"Fuel Type"
                                , "Is there one central heating pump for this heat source?"
                                , "Is the central heating pump inside dwelling?"
                                # ,"Number of Central Heating Pumps"
                                ,"Number of central heating pumps?"
                                # ,"Number of Central Heating Pumps Inside"
                                ,"Number of central heating pumps inside?"
                                ,"Number of Oil Boiler Pumps"
                                ,"Number of Oil Boiler Pumps Inside"
                                ,"Is there a gas boiler flue fan?"
                                ,"Number of Gas Boiler Flue Fans"
                                # ,"Heating System category (Primary Heating System only)"
                                , "Heating System Category"
                                # ,"Sub Category (Primary Heating System only)"
                                , "Sub Category"
                                ,"Sub Category2 (Primary Heating System only)"
                                # ,"Heating System (Primary Heating System only)"
                                , "Heating System"
                                ,"Delayed start thermostat present?"
                                ,"Passive flue gas heat recovery device present?"
                                ,"Load or weather compensation?"
                                ,"Integrated thermal store present?"
                                ]
                
                headers = list(filter((None).__ne__, headers))
                
                
                
                # print(sheet.title, ':', output[sheet.title])
                
                
                
                
                output[sheet.title]["headers"] = headers
                
            elif sheet.title in lookup_tables:
                lookup[sheet.title] = {}
                for i, row in enumerate(list(sheet.values)):
                    field_name = row[1]
                    if field_name == None:
                        continue
                    # print('field_name', ':', field_name)
                    
                    field_req = row[2]
                    field_loc = no_2_alph(2) + str(i)
                    # print('field_loc', ':', field_loc)
                    lookup[sheet.title][field_name] = {"field_req": field_req, "field_loc": field_loc}
                
            elif sheet.title == 'Floor reference': # one-to-many
                lookup[sheet.title] = {}
                for i, row in enumerate(list(sheet.values)):
                    KSN_ref = row[0].strip()
                    MP_ref = row[1]
                    if MP_ref == '-2 to 8':
                        MP_ref = list(x for x in range(-2, 9))
                    if MP_ref == '9 to 14':
                        MP_ref = list(x for x in range(9, 15))
                    lookup[sheet.title][KSN_ref] = MP_ref
            
            elif sheet.title == 'Object Reference': # one-to-many
                lookup[sheet.title] = {}
                lookup[sheet.title + ' alt'] = {}
                for i, row in enumerate(list(sheet.values)):
                    field_category = row[0]
                    field_name = row[1]
                    field_id = row[4]
                    if field_category == None or field_id == None:
                        continue
                    if row[0] not in lookup[sheet.title].keys():
                        lookup[sheet.title][field_category] = []
                    lookup[sheet.title][field_category].append(field_id)
                    if field_category not in lookup[sheet.title + ' alt'].keys():
                        lookup[sheet.title + ' alt'][field_category] = {}
                    lookup[sheet.title + ' alt'][field_category][field_id] = field_name
                    lookup[sheet.title + ' alt'][field_category][field_name] = field_id

        
        lookup['Object Reference']['Floor Reference'] = lookup['Floor reference']
        lookup['Object Reference']['alt'] = lookup['Object Reference alt']
        
        
        # print("output['8. Ventilation P1']", ':')
        # pprint.pprint(output['8. Ventilation P1'])
        # print("lookup['Object Reference']", ':')
        # pprint.pprint(lookup['Object Reference'])
        # print("lookup['Floor reference']", ':')
        # pprint.pprint(lookup['Floor reference'])
                
    except:
        output = traceback.format_exc()
        print(output)
        
        
    return output, lookup



def XL_2_dict(
                file_name
                , account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
                , default_credential = DefaultAzureCredential()
                , container_from = 'attachment'
                , local_path_from = "/tmp"
                # , container_to = 'project-files'
                # , local_path_to = plan_name
                ):
    '''
    Read the contents of an Excel Workbook from Azure Blob Storage and return as a dictionary
    
    Three types of worksheet expected format: [trivial, table, lookup_table]
    
    trivial:
        2 columns, B & C, field_name and field_req
        field_req is one of the following four : [exact text, variable name, count, logic_dict]
        
    table: 
        need to update populate_template_new() to handle multicol outputs
        need to establish how many columns?
    
    lookup_table:
        2 cols
        create lookup_table_dict
        lookup is: field_value = lookup_table_dict[col_1]?
    
    '''
    
    try:
        azure_source_fp = file_name
        instance_fp = os.path.join(os.getcwd(), local_path_from, file_name)
        print(instance_fp)
        
        # file_name = plan_name + ' Major Renovation calculation' + '.xlsx'


        # Create the BlobServiceClient object
        blob_service_client = BlobServiceClient(account_url, credential=default_credential)
        container_client = blob_service_client.get_container_client(container= container_from) 
        
        with open(file=instance_fp, mode="wb") as download_file:
            download_file.write(container_client.download_blob(azure_source_fp).readall())
        
        # file_content = container_client.download_blob(azure_source_fp).readall()
        
        wb = openpyxl.load_workbook(instance_fp, data_only = True)
        
        output_tables = ["1. Survey Details P1"
                        , "7. Thermal Mass P1"
                        , "8. Ventilation P1"
                        , "11. Lighting P1"
                        ]
        
        
        lookup_tables = ['1.1 Assessor Details Table'
                            , '2.2 Referance Table S8'
                            , '3.1 Referance Table S5'
                            , '3.2 Referance Table S4'
                            , '3.3 Referance Table Non Default'
                            , '7.1 Referance Table'
                            ]
       
        multicol_tables = ['5.1 Windows Summary Table'
                            , '5.2 Window Schedule Table'
                            , '5.3 Building | Doors P1'
                            , '6. Colour Area Table P1'
                            , '8.1 Ventilation Items'
                            , '11.1 Lighting Schedule'

                            ]
        output = {}
        lookup = {}
        for sheet in wb.worksheets:
            # print(sheet.title)
            
            
            if sheet.title in multicol_tables:
                print('multicol_table', ':', sheet.title)
            
            elif sheet.title in output_tables:
                    output[sheet.title] = {}
                    for i, row in enumerate(list(sheet.values)):
                        field_name = row[1]
                        if field_name == None:
                            continue
                        # print('field_name', ':', field_name)
                        
                        field_req = row[2]
                        field_loc = no_2_alph(2) + str(i)
                        # print('field_loc', ':', field_loc)
                        output[sheet.title][field_name] = {"field_req": field_req, "field_loc": field_loc}
                
            elif sheet.title in lookup_tables:
                    lookup[sheet.title] = {}
                    for i, row in enumerate(list(sheet.values)):
                        field_name = row[1]
                        if field_name == None:
                            continue
                        # print('field_name', ':', field_name)
                        
                        field_req = row[2]
                        field_loc = no_2_alph(2) + str(i)
                        # print('field_loc', ':', field_loc)
                        lookup[sheet.title][field_name] = {"field_req": field_req, "field_loc": field_loc}
                

            
    except:
        output = traceback.format_exc()
        print(output)
        
        
    return output, lookup

def get_stats_data(project_id, headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"
            ,"key": "45170e50321733db78952dfa5901b0dfeeb8"
            , "customer": "63b5a4ae69c91"
            , "accept": "application/json"
            }
            , xml_ref_dict = {}
            , xl_ref_dict = {}
            , floor_type_dict = {}
            ):
    try:
        output = {}
        
        
        json_url = "https://cloud.magicplan.app/api/v2/plans/statistics/" + str(project_id)
        request = urllib.request.Request(json_url, headers=headers)
        print(type(request))
        if not type(request) is urllib.request.Request:
            raise Exception(request)
        
        
        JSON = urllib.request.urlopen(request).read()
        print('type(JSON)', ':', type(JSON))
        if not type(JSON) is bytes:
            raise Exception(JSON)
        
        JSON = json.loads(JSON)
        print('type(JSON)', ':', type(JSON))
        if not isinstance(JSON, dict):
            raise Exception(JSON)
        
        
        floor_dict = {} # need to know floor uid/names to combine with colour/uid from XML?
        roof_dict = {}
        
        count_dict = {}
        door_dict = {}
        window_dict = {}
        bulb_dict = {}
        vent_dict = {}
        
        storey_height_dict = {}
        
        
        wall_dict = {} # this comes from XML? But then requires additional info from Forms
        
        # colour_dict = {} 
        
        # below are lists of ids but statistics file uses names:
        doors = xl_ref_dict['Door']
        windows = xl_ref_dict['Windows']
        bulbs = xl_ref_dict['Lighting']
        vents = xl_ref_dict['Wall Ventilation'] + xl_ref_dict['Mechanical Ventilation Systems']
        roof_objects = xl_ref_dict['Roofs']
        
        
        composite_count_objects = ['Number of Intermittent Fans', 'Number of openings', 'Number of openings Draughtproofed'] # add column(s) to Object Reference to convey this information *KSN need to be able to control this so it shouldn't be hard coded
        
        
        
        intermittent_fans = ["Broken Cooker Hood", "Broken Mechanical Vent", "New Mechanical Vent", "Ducted Cooker Hood", "Existing Mechanical Vent"] # as above
        
        # Hatches: all in "Roofs" category, add column to indicate Not/Draughtproofed subcategory)
        # thermal envelope only (specifiy in input tables - add column that refers to "Floor Reference" tab
        # for roof_object in roof_objects:
            # print('roof_object', ':', roof_object)
        
        
        attic_hatches_draught_stripped = ["Fixed Ladder Hatch Draughtproofed", "Attic Hatch Draughtproofed", "Wall Hatch Draughtproofed"] 
        attic_hatches_not_draught_stripped = ["Fixed Ladder Hatch Not Draughtproofed", "Attic Hatch Not Draughtproofed", "Wall Hatch Not Draughtproofed"]
        openings = windows + doors + attic_hatches_draught_stripped + attic_hatches_not_draught_stripped # add column (should this object be considered an opening?)
        
        # need to add Chimneys and Flues to count_dict
        # these only exist as "Ventilation Types" associated with certain Heating Objects
        # only visible in forms
        cond_count_objects = ["Chimney", "Flue"]
               
        # print("xl_ref_dict", ':')
        # pprint.pprint(xl_ref_dict)
        
        count_objects = composite_count_objects + bulbs + intermittent_fans + openings + cond_count_objects # add column
        print('count_objects', ':', count_objects)
        for co_id in count_objects:
            for category in xl_ref_dict['alt'].keys():
                # print(category.keys())
                if co_id in xl_ref_dict['alt'][category].keys():
                    co_name = xl_ref_dict['alt'][category][co_id]
                    count_dict[co_name] = 0
                
            # count_dict[co_id] = 0
        
        
        # print('floor_type_dict', ';')
        # pprint.pprint(floor_type_dict)
        
        
        
        # some of our objects appear as wall items, some as furnitures, check both?
        
        # print('type(JSON["data"]["project_statistics"]["floors"])', ':', type(JSON["data"]["project_statistics"]["floors"]))
        if not isinstance(JSON["data"]["project_statistics"]["floors"], list):
            raise Exception(JSON["data"]["project_statistics"]["floors"])
        
        for floor in JSON["data"]["project_statistics"]["floors"]:
            floor_uid = floor["uid"]
            floor_name = floor["name"]
            floor_dict[floor_uid] = floor_name
            
            storey_height_dict[floor["uid"]] = {}
            storey_height_dict[floor["uid"]]['value'] = {}
            storey_height_dict[floor["uid"]]['value']['area'] = floor["area"]
            storey_height_dict[floor["uid"]]['value']['height'] = floor["height"]
            storey_height_dict[floor["uid"]]['value']['name'] = floor["name"]
            for room in floor["rooms"]:
                storey_height_dict[room["uid"]] = {}
                storey_height_dict[room["uid"]]['value'] = {}
                storey_height_dict[room["uid"]]['value']['area'] = room["area"]
                storey_height_dict[room["uid"]]['value']['height'] = room["height"]
                storey_height_dict[room["uid"]]['value']['name'] = room["name"]
                storey_height_dict[room["uid"]]['value']['volume'] = round(float(room["area"]) * float(room["height"]), 2)
                if room["uid"] in floor_type_dict.keys():
                    if 'value' not in floor_type_dict[room["uid"]].keys():
                        floor_type_dict[room["uid"]]['value'] = {}
                    floor_type_dict[room["uid"]]['value']['area'] = room["area"]
                    floor_type_dict[room["uid"]]['value']['name'] = room["name"]
                    floor_type_dict[room["uid"]]['value']['perimeter'] = room["perimeter"]
                    floor_type_dict[room["uid"]]['value']['floor_uid'] = floor_uid
                    floor_type_dict[room["uid"]]['value']['floor_name'] = floor_name
            
            
            if floor["uid"] in xml_ref_dict.keys(): # when would it not be? why is this condition necessary here?
                # if xml_ref_dict[floor["uid"]] == '1000':
                    # del floor_dict[floor["uid"]]
                
                for room in floor["rooms"]:
                    room_uid = room["uid"]
                    
                    if xml_ref_dict[floor["uid"]] == '1000':
                        roof_dict[room["uid"]] = {}
                        roof_dict[room["uid"]]['value'] = {}
                        for room_stat in room:
                            if room_stat in ["name", "area", "area_without_walls", "area_with_interior_walls_only", "area_with_walls"]:
                                roof_dict[room["uid"]]['value'][room_stat] = room[room_stat]
                        
                    
                    # Need special one for roof/skylights - furnitures on a non-true-floor (1000? could this loop be tabbed right under the above umbrella?) that need to appear in window table
                    for furniture in room["furnitures"]:
                        if furniture["id"] in windows: # can this only happen on floor 1000?
                            window_dict[furniture["uid"]] = {}
                            window_dict[furniture["uid"]]['value'] = {}
                            window_dict[furniture["uid"]]['value']["name"] = furniture["name"]
                            window_dict[furniture["uid"]]['value']["height"] = furniture["depth"] # note difference for roof lights
                            window_dict[furniture["uid"]]['value']["width"] = furniture["width"]
                            window_dict[furniture["uid"]]['value']["room_uid"] = room["uid"]
                            window_dict[furniture["uid"]]['value']["room_name"] = room["name"]
                            window_dict[furniture["uid"]]['value']["floor_name"] = floor["name"]
                            window_dict[furniture["uid"]]['value']["floor_uid"] = floor["uid"]
                    
                    
                    if floor_uid in xml_ref_dict['true_floors'] and room_uid in xml_ref_dict['thermal_envelope_uids']:
                        if room_uid not in count_dict.keys():
                            count_dict[room_uid] = {}
                    
                        for furniture in room["furnitures"]:
                            # print('furniture["name"]', ':', furniture["name"], ' ', 'furniture["id"]', ':', furniture["id"])
                            if furniture["id"] in count_objects:
                                furniture["name"] = furniture["name"].strip()
                                count_dict[furniture["name"]] += 1
                                if furniture["name"] not in count_dict[room_uid].keys():
                                    count_dict[room_uid][furniture["name"]] = 0
                                count_dict[room_uid][furniture["name"]] += 1
                            
                            if furniture["id"] in bulbs:
                                bulb_dict[furniture["uid"]] = {}
                                bulb_dict[furniture["uid"]]['value'] = {}
                                bulb_dict[furniture["uid"]]['value']["name"] = furniture["name"]
                                bulb_dict[furniture["uid"]]['value']["room_uid"] = room["uid"]
                                bulb_dict[furniture["uid"]]['value']["room_name"] = room["name"]
                                bulb_dict[furniture["uid"]]['value']["floor_name"] = floor["name"]
                                bulb_dict[furniture["uid"]]['value']["floor_uid"] = floor["uid"]

                            if furniture["id"] in vents:
                                vent_dict[furniture["uid"]] = {}
                                vent_dict[furniture["uid"]]['value'] = {}
                                vent_dict[furniture["uid"]]['value']["name"] = furniture["name"]
                                vent_dict[furniture["uid"]]['value']["room_uid"] = room["uid"]
                                vent_dict[furniture["uid"]]['value']["room_name"] = room["name"]
                                vent_dict[furniture["uid"]]['value']["floor_name"] = floor["name"]
                                vent_dict[furniture["uid"]]['value']["floor_uid"] = floor["uid"]
                        
                        for wall_item in room["wall_items"]:
                            # print(wall_item["name"])
                            # print(wall_item["uid"])
                            if wall_item["id"] in count_objects:
                                count_dict[wall_item["name"]] += 1
                                if wall_item["name"] not in count_dict[room_uid].keys():
                                    count_dict[room_uid][wall_item["name"]] = 0
                                count_dict[room_uid][wall_item["name"]] += 1
                                
                            if wall_item["id"] in vents:
                                vent_dict[wall_item["uid"]] = {}
                                vent_dict[wall_item["uid"]]['value'] = {}
                                vent_dict[wall_item["uid"]]['value']["name"] = wall_item["name"]
                                vent_dict[wall_item["uid"]]['value']["room_uid"] = room["uid"]
                                vent_dict[wall_item["uid"]]['value']["room_name"] = room["name"]
                                vent_dict[wall_item["uid"]]['value']["floor_name"] = floor["name"]
                                vent_dict[wall_item["uid"]]['value']["floor_uid"] = floor["uid"]
                                
                            if wall_item["id"] in doors:
                                door_dict[wall_item["uid"]] = {}
                                door_dict[wall_item["uid"]]['value'] = {}
                                door_dict[wall_item["uid"]]['value']["name"] = wall_item["name"]
                                door_dict[wall_item["uid"]]['value']["height"] = wall_item["height"]
                                door_dict[wall_item["uid"]]['value']["width"] = wall_item["width"]
                                door_dict[wall_item["uid"]]['value']["room_uid"] = room["uid"]
                                door_dict[wall_item["uid"]]['value']["room_name"] = room["name"]
                                door_dict[wall_item["uid"]]['value']["floor_name"] = floor["name"]
                                door_dict[wall_item["uid"]]['value']["floor_uid"] = floor["uid"]
                            if wall_item["id"] in windows:
                                window_dict[wall_item["uid"]] = {}
                                window_dict[wall_item["uid"]]['value'] = {}
                                window_dict[wall_item["uid"]]['value']["name"] = wall_item["name"]
                                window_dict[wall_item["uid"]]['value']["height"] = wall_item["height"]
                                window_dict[wall_item["uid"]]['value']["width"] = wall_item["width"]
                                window_dict[wall_item["uid"]]['value']["room_uid"] = room["uid"]
                                window_dict[wall_item["uid"]]['value']["room_name"] = room["name"]
                                window_dict[wall_item["uid"]]['value']["floor_name"] = floor["name"]
                                window_dict[wall_item["uid"]]['value']["floor_uid"] = floor["uid"]
                
                # Get "rooflight_area" (to be subtracted later after slope calculations)
                for w in window_dict:
                    room_uid = window_dict[w]['value']["room_uid"]
                    if room_uid in roof_dict.keys():
                        a = window_dict[w]['value']["height"] * window_dict[w]['value']["width"] # note this calculation is performed again later in window_forms_append()
                        if 'rooflight_area' not in roof_dict[room_uid]['value'].keys():
                            roof_dict[room_uid]['value']['rooflight_area'] = 0
                        roof_dict[room_uid]['value']['rooflight_area'] += a
                
                
        # print('roof_dict', ';')
        # pprint.pprint(roof_dict)
        # print('floor_type_dict (post)', ';')
        # pprint.pprint(floor_type_dict)
        
        
        for intermittent_fan in intermittent_fans:
            if 'Number of Intermittent Fans' not in count_dict.keys():
                count_dict['Number of Intermittent Fans'] = 0
            if intermittent_fan in count_dict.keys():
                count_dict['Number of Intermittent Fans'] += count_dict[intermittent_fan]
        
        for opening in openings:
            if 'Number of openings' not in count_dict.keys():
                count_dict['Number of openings'] = 0
            if opening in count_dict.keys():
                count_dict['Number of openings'] += count_dict[opening]
        
        # for opening in openings_draughtproofed:
            # count_dict['Number of openings Draughtproofed'] += count_dict[opening]
        
        # print('door_dict', ':')
        # pprint.pprint(door_dict)
        # print('floor_dict', ':')
        # pprint.pprint(floor_dict)
        
        
        
        output['count_dict'] = count_dict
        output['door_dict'] = door_dict
        output['window_dict'] = window_dict
        output['bulb_dict'] = bulb_dict
        output['vent_dict'] = vent_dict
        output['floor_dict'] = floor_dict
        output['floor_type_dict'] = floor_type_dict
        output['roof_dict'] = roof_dict
        output['storey_height_dict'] = storey_height_dict

    except:
        output = traceback.format_exc()
        print(output)
    
    return output

def JSON_2_dict(project_id, headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"
            ,"key": "45170e50321733db78952dfa5901b0dfeeb8"
            , "customer": "63b5a4ae69c91"
            , "accept": "application/json"
            }
            , forms_data = {} # more efficient to pass it in like this than to extract it again?
            , xml_ref_dict = {}
            , xl_ref_dict = {}
            , wall_dict={}
            , wo={}
            , nwa_dict={}
            , colours_dict={}
            ):
    '''
    This function retrieves Forms & Statistics JSON files from MagicPlan
    Combines and filters the data to return json_dict:
        floor; room; object; details
        label_val_dict
    '''
    try:
        json_dict = {}
        
        # json_dict: need to combine "forms_full_dict" and statistics (statistics_full_dict?)
        
        # **************** GET FORMS DATA **************** 
        if forms_data == {}:
            forms_data = get_forms_data(project_id)
        
        
        # print("forms_data['wall_type_dict']", ':')
        # pprint.pprint(forms_data['wall_type_dict'])
        
        
        wall_type_dict = {}
        for wt in forms_data['wall_type_dict']:
            for key in forms_data['wall_type_dict'][wt].keys():
                if 'Is there a' in key or wt == 'Wall Type 1':
                    if forms_data['wall_type_dict'][wt][key] == True or wt == 'Wall Type 1':
                        wall_type_dict[wt] = forms_data['wall_type_dict'][wt]
                        
                        wall_type_dict[wt + ' - Semi-Exposed'] = {}
                        wall_type_dict[wt + ' - Semi-Exposed']['value'] = {}
                        for key in forms_data['wall_type_dict'][wt].keys():
                            if key == 'value':
                                for v in forms_data['wall_type_dict'][wt][key]:
                                    wall_type_dict[wt + ' - Semi-Exposed'][key][v] = forms_data['wall_type_dict'][wt][key][v]
                            else:
                                wall_type_dict[wt + ' - Semi-Exposed'][key] = forms_data['wall_type_dict'][wt][key]
                            
                        # wall_type_dict[wt + ' - Semi-Exposed'] = forms_data['wall_type_dict'][wt + ' - Semi-Exposed']
                        wall_type_dict[wt + ' - Semi-Exposed']['value']['Is semi exposed'] = 'Yes'
                        # print("wall_type_dict[wt + ' - Semi-Exposed']['Is semi exposed']", ':', wall_type_dict[wt + ' - Semi-Exposed']['Is semi exposed'])
        
        
        
        
        
        
        form_val_dict = forms_data['form_val_dict']
        forms_full_dict = forms_data['forms_full_dict']
        forms_uid_dict = forms_data['forms_uid_dict']
        
        json_dict = form_val_dict
        json_uid_dict = forms_uid_dict
        
        window_detail_dict = forms_data['window_detail_dict']
        floor_type_dict = forms_data['floor_type_dict']
        
        # print('wo', ':')
        # pprint.pprint(wo)
        # print('forms_data["floor_type_dict"]', ':')
        # pprint.pprint(forms_data["floor_type_dict"])
        
        # **************** GET STATS DATA **************** 
        
        stats_data = get_stats_data(project_id
                                , xml_ref_dict=xml_ref_dict # real floors/excluded rooms
                                , xl_ref_dict=xl_ref_dict
                                , floor_type_dict = floor_type_dict
                                )
        if not isinstance(stats_data, dict):
            raise Exception(stats_data)
        
        
        
        if 'count_dict' in stats_data.keys(): # why wouldn't it be? Don't like this reliance on this condition
            count_dict = stats_data['count_dict']
            # print('count_dict', ':')
            # pprint.pprint(count_dict)
            # add counts to json_dict as strings (top-level key:value pairs - should this be happening here or is it more suited to return the table as-is and then do the output prep all together? what about multi-col tables? any counts in those? if so, will this approach work?):
            for key in count_dict.keys():
                json_dict[key] = str(count_dict[key])
        
            # where should we exclude floors/ex-thermal envelope?
            for object_uid in json_uid_dict:
                if 'Ventilation Type' in json_uid_dict[object_uid].keys():
                    # print('Ventilation Type', ':', json_uid_dict[object_uid]['Ventilation Type'])
                    if json_uid_dict[object_uid]['Ventilation Type'] in count_dict.keys():
                        count_dict[json_uid_dict[object_uid]['Ventilation Type']] += 1
        
        
        # adding these like this for now, might change depending on what is most convenient later:
        json_dict["bulb_dict"] = stats_data["bulb_dict"]
        json_dict["vent_dict"] = stats_data["vent_dict"]
        json_dict["storey_height_dict"] = stats_data["storey_height_dict"]
        
        # Now need to go through Forms adding fields to our object dicts by uid
        # print('stats_data["door_dict"]', ':')
        # pprint.pprint(stats_data["door_dict"])
        # print('wall_type_dict', ':')
        # pprint.pprint(wall_type_dict)
        # json_dict["wall_dict"] = stats_append(wall_dict, forms_uid_dict)
        
        json_dict["door_dict"] = stats_append(stats_data["door_dict"], forms_uid_dict) 
        json_dict["window_dict"] = stats_append(stats_data["window_dict"], forms_uid_dict) # forms_append?
        json_dict["roof_dict"] =  stats_append(stats_data["roof_dict"], forms_uid_dict) 
        
        # print('json_dict["door_dict"]', ':')
        # pprint.pprint(json_dict["door_dict"])
        
        
        # 66d588c8.6ad3e7ff
        
        json_dict["floor_name_dict"] = stats_data["floor_dict"]

        
        # ************** WINDOW ORIENTATION ***************
        
        for window in json_dict['window_dict']:
            if window in wo.keys():
                json_dict['window_dict'][window]['value']['cardinal_direction'] = wo[window]
        
        
        
        json_dict["heating_dict"] = {}
        for ft in forms_data['heating_dict']:
            if ft not in json_dict["heating_dict"].keys():
                json_dict["heating_dict"][ft] = {}
            json_dict["heating_dict"][ft]['value'] = forms_data['heating_dict'][ft]
        
        # print('json_dict["heating_dict"]', ':')
        # pprint.pprint(json_dict["heating_dict"])
        
        
        
        json_dict["floor_type_dict"] = {}
        for ft in forms_data['floor_type_dict']:
            if ft not in json_dict["floor_type_dict"].keys():
                json_dict["floor_type_dict"][ft] = {}
            json_dict["floor_type_dict"][ft]['value'] = forms_data['floor_type_dict'][ft]
        
        
        
        for ft in stats_data['floor_type_dict']:
            if 'value' in stats_data['floor_type_dict'][ft].keys():
                if ft not in json_dict["floor_type_dict"].keys():
                    json_dict["floor_type_dict"][ft] = {}
                for v in stats_data['floor_type_dict'][ft]['value']:
                    json_dict["floor_type_dict"][ft]['value'][v] = stats_data['floor_type_dict'][ft]['value'][v]
                    
        
        
        
        
        json_dict["door_dict"] = door_forms_append(json_dict["door_dict"], forms_uid_dict)
        json_dict["window_dict"] = window_forms_append(json_dict["window_dict"], forms_uid_dict, window_detail_dict)
        
        # print('json_dict["window_dict"]', ':')
        # pprint.pprint(json_dict["window_dict"])
        
        
        json_dict['window_summary_dict'] = window_summary(json_dict["window_dict"])
        json_dict['door_summary_dict'] = door_summary(json_dict["door_dict"])
        json_dict['bulb_summary_dict'] = bulb_summary(json_dict["bulb_dict"])
        
        print('json_dict["bulb_summary_dict"]', ':')
        pprint.pprint(json_dict["bulb_summary_dict"])
        
        
        
        
        # print('about to condense floor_dict')
        # json_dict["floor_dict"] = condense(json_dict["floor_dict"], json_dict)
        
        print('about to condense floor_type_dict')
        json_dict["floor_dict"] = condense(json_dict["floor_type_dict"], json_dict)
        
        print('about to condense roof_dict')
        json_dict["roof_dict"] = condense(json_dict["roof_dict"], json_dict)
        
        # print('json_dict["roof_dict"]', ':')
        # pprint.pprint(json_dict["roof_dict"])
        
        for rt in json_dict["roof_dict"]:
            if 'roof pitch (degrees)' in json_dict["roof_dict"][rt]['value'].keys():
                ra = json_dict["roof_dict"][rt]['value']['area (m2)']
                rp = json_dict["roof_dict"][rt]['value']['roof pitch (degrees)']
                if rp != '':
                    sa = ra / cos(rp / 57.2958)
                    # print('sa', ':', sa)
                    json_dict["roof_dict"][rt]['value']['area (m2)'] = round(sa, 3)
            
            if 'rooflight_area' in json_dict["roof_dict"][rt]['value'].keys():
                json_dict["roof_dict"][rt]['value']['area (m2)'] -= json_dict["roof_dict"][rt]['value']['rooflight_area']
            
        # print('post calc', ':')
        # pprint.pprint(json_dict["roof_dict"])
        
        json_dict['roof_summary_dict'] = roof_summary(json_dict["roof_dict"])
        
        
        
        
        
        
        
        # below is the Wall Type condensation operation, still separate as it is slightly different to the first two but should be incorporated
        for w in wall_type_dict:
            wall_type_dict[w]['value'] = {}
            for x in wall_type_dict[w]:
                # print('x', ':', x)
                if x == w:
                    wall_type_dict[w]['value']['wall type'] = wall_type_dict[w][x]
                
                if x == 'Other ' + w:
                    wall_type_dict[w]['value']['other wall type'] = wall_type_dict[w][x]
                    
                if 'Description' in x: 
                    wall_type_dict[w]['value']['description'] = wall_type_dict[w][x]
                    
                if 'age band match' in x: 
                    wall_type_dict[w]['value']['dwelling age band?'] = wall_type_dict[w][x]
                    if wall_type_dict[w][x] == True:
                        wall_type_dict[w]['value']['age band'] = json_dict['Age: Dwelling']['value']

                if x == 'Wall Age Band': 
                    wall_type_dict[w]['value']['age band'] = json_dict[wall_type_dict[w][x]]['value']
                
                if 'Wall Insulation' in x: 
                    wall_type_dict[w]['value']['insulation'] = wall_type_dict[w][x]
                
                if x == 'Explain':
                    wall_type_dict[w]['value']['explanation'] = wall_type_dict[w][x]
                
                if 'default wall U-value?' in x: 
                    wall_type_dict[w]['value']['default U-value?'] = wall_type_dict[w][x]
                
                if 'U-value (W/m' in x: 
                    wall_type_dict[w]['value']['U-value'] = wall_type_dict[w][x]
                
                if 'U-value be substantiated?' in x: 
                    wall_type_dict[w]['value']['can u-value be substantiated?'] = wall_type_dict[w][x]
                
                if 'insulation thickness' in x: 
                    wall_type_dict[w]['value']['insulation thickness'] = wall_type_dict[w][x]
                
                if 'thermal conductivity be substantiated' in x: 
                    wall_type_dict[w]['value']['can thermal conductivity be substantiated?'] = wall_type_dict[w][x]
                
                if 'thermal conductivity (λ)' in x: 
                    wall_type_dict[w]['value']['thermal conductivity (λ)'] = wall_type_dict[w][x]
        
        
        
        
        
        
        for wt in wall_type_dict:
            if 'Semi' not in wt:
                for k in wall_type_dict[wt]['value'].keys():
                    v = wall_type_dict[wt]['value'][k]
                    wall_type_dict[wt + ' - Semi-Exposed']['value'][k] = v
                wall_type_dict[wt + ' - Semi-Exposed']['value']['Is semi exposed'] = 'Yes'
        
        
        
        json_dict['wall_type_dict'] = wall_type_dict
        json_dict['count_dict'] = count_dict
        
        
        # print('json_dict["count_dict"]', ':')
        # pprint.pprint(json_dict['count_dict'])
        
        
        
        output = json_dict
    
    except:
        output = traceback.format_exc()
        print(output)
   
    return output


def condense(old_dict, json_dict):
    try:
        # print('json_dict', ':')
        # pprint.pprint(json_dict)
        # print('old_dict', ':')
        # pprint.pprint(old_dict)
        
        new_dict = old_dict.copy() # is this still necessary?
        for e in old_dict:
            # new_dict[e]['value'] = {} # why did we have to do this with wall_type_dict below?
            new_dict[e]['value_condensed'] = {}
            
            
            
            
            for x in new_dict[e]['value']:
                # print('x', ':', x)
                
                if x == 'area':
                    new_dict[e]['value_condensed']['area (m2)'] = new_dict[e]['value'][x]
                if x == 'rooflight_area':
                    new_dict[e]['value_condensed']['rooflight_area'] = new_dict[e]['value'][x]
                if x == 'name':
                    new_dict[e]['value_condensed']['name'] = new_dict[e]['value'][x]
                if x == 'perimeter':
                    new_dict[e]['value_condensed']['perimeter'] = new_dict[e]['value'][x]
                if x == 'Floor Type':
                    new_dict[e]['value_condensed']['Floor Type'] = new_dict[e]['value'][x]
                if x == 'floor_name':
                    new_dict[e]['value_condensed']['floor_name'] = new_dict[e]['value'][x]
                
                # is this floor being used ... DEAP? - probably not the right place to apply this filter
                
                if x == 'Floor Age Band': # year range?
                    new_dict[e]['value_condensed']['age band'] = json_dict[new_dict[e]['value'][x]]['value']
                
                if x == 'Is there underfloor heating?':
                    new_dict[e]['value_condensed']['underfloor heating?'] = new_dict[e]['value'][x]
                
                
                
                if 's this floor being used' in x:
                    new_dict[e]['value_condensed']['is this floor being used?'] = new_dict[e]['value'][x]
                if 'description' in x:
                    new_dict[e]['value_condensed']['description'] = new_dict[e]['value'][x]
                
                
                if 'U-value calculation required?' in x:
                    new_dict[e]['value_condensed']['U-value calculation required?'] = new_dict[e]['value'][x]
                if 'espoke calculation U-Value [W' in x:
                    new_dict[e]['value_condensed']['calculated U-value'] = new_dict[e]['value'][x]
                
                # need to identify "U-Value" field? How to differentiate from above?
                
                
                if 'BER Roof Type' in x:
                    new_dict[e]['value_condensed']['roof type'] = new_dict[e]['value'][x]
                if '(degrees)' in x:
                    new_dict[e]['value_condensed']['roof pitch (degrees)'] = new_dict[e]['value'][x]
                if 'age band match' in x:
                    new_dict[e]['value_condensed']['dwelling age band?'] = new_dict[e]['value'][x] # True or False
                    if new_dict[e]['value'][x] == True:
                        new_dict[e]['value_condensed']['age band'] = json_dict['Age: Dwelling']['value']
                    
                
                if x == 'Roof Age Band': # e.g. "Age: Extension 1" (only ever appears if above is False)
                    if new_dict[e]['value'][x] in json_dict.keys(): # 
                        new_dict[e]['value_condensed']['age band'] = json_dict[new_dict[e]['value'][x]]['value']
                    else:
                        new_dict[e]['value_condensed']['age band'] = new_dict[e]['value'][x]
                    


                if 'thickness known?' in x:
                    new_dict[e]['value_condensed']['insulation thickness known?'] = new_dict[e]['value'][x]
                if 'insulation thickness (mm)' in x:
                    new_dict[e]['value_condensed']['insulation thickness (mm)'] = new_dict[e]['value'][x]
                if 'hermal conductivity (W' in x:
                    new_dict[e]['value_condensed']['thermal conductivity'] = new_dict[e]['value'][x]
                if 'roof U-value (W/' in x:
                    new_dict[e]['value_condensed']['U-value (W/m2K)'] = new_dict[e]['value'][x]

            
            
            
            
            new_dict[e]['value'] = new_dict[e]['value_condensed']
        
        
        # print('new_dict', ':')
        # pprint.pprint(new_dict)
        
        old_dict = new_dict
        output = new_dict
        
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output



def floor_stats_append(stats_dict, forms_uid_dict):
    try:
        # print('stats_dict', ':')
        # pprint.pprint(stats_dict)
        # print('forms_uid_dict', ':')
        # pprint.pprint(forms_uid_dict)
        
        
        for item in stats_dict:
            for room in stats_dict[item]['rooms']:
                if room in forms_uid_dict:
                    for field in forms_uid_dict[room]:
                        stats_dict[item]['value'][field] = forms_uid_dict[room][field]
        
        
        
        output = stats_dict
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output

def stats_append(stats_dict, forms_dict):
    try:
        # print('stats_dict', ':')
        # pprint.pprint(stats_dict)
        # print('forms_dict', ':')
        # pprint.pprint(forms_dict)
        
        
        for item in stats_dict:
            if item in forms_dict.keys():
                for field in forms_dict[item]:
                    stats_dict[item]['value'][field] = forms_dict[item][field]
        
        
        output = stats_dict
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output



def roof_summary(roof_dict):
    try:
        roof_summary_dict = {}
        for roof in roof_dict:
            # print("roof", ':', roof)
            # print("json_dict['roof_dict'][roof]['value']", ':')
            # pprint.pprint(roof_dict[roof]['value'])
            key = ''
            for keypart in ['roof type', 'insulation thickness (mm)', 'age band']:
                if keypart in roof_dict[roof]['value'].keys():
                    key += (str(roof_dict[roof]['value'][keypart]) + '_')
            key = str(hash(key))
            if key not in roof_summary_dict.keys():
                roof_summary_dict[key] = {}
                roof_summary_dict[key]['value'] = {}
                roof_summary_dict[key]['value']['Description'] = ''
                roof_summary_dict[key]['value']['Total roof type Area (m²)'] = 0
                # roof_summary_dict[key]['value']['Count'] = 0
                # roof_summary_dict[key]['value']['Number of openings'] = 0
                # roof_summary_dict[key]['value']['Number of openings draughtstripped'] = 0
            
            for keypart in ['roof type', 'description', 'insulation thickness (mm)', 'age band', 'U-value']:
                if keypart in roof_dict[roof]['value'].keys():
                    roof_summary_dict[key]['value'][keypart] = roof_dict[roof]['value'][keypart] 
            
            roof_summary_dict[key]['value']['Total roof type Area (m²)'] += roof_dict[roof]['value']['area (m2)']
            roof_summary_dict[key]['value']['Total roof type Area (m²)'] = round(roof_summary_dict[key]['value']['Total roof type Area (m²)'], 2)
            
            # roof_summary_dict[key]['value']['Count'] += 1
            # roof_summary_dict[key]['value']['Number of openings'] += roof_dict[roof]['value']['Number of openings']
            # roof_summary_dict[key]['value']['Number of openings draughtstripped'] += roof_dict[roof]['value']['Number of openings draughtstripped']
        
        
        
        
        output = roof_summary_dict
        
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output


def door_summary(door_dict):
    try:
        door_summary_dict = {}
        for door in door_dict:
            # print("door", ':', door)
            # print("json_dict['door_dict'][door]['value']", ':')
            # pprint.pprint(door_dict[door]['value'])
            key = ''
            for keypart in ['Type', 'Draught Stripped']:
                if keypart in door_dict[door]['value'].keys():
                    key += (str(door_dict[door]['value'][keypart]) + '_')
            key = str(hash(key))
            if key not in door_summary_dict.keys():
                door_summary_dict[key] = {}
                door_summary_dict[key]['value'] = {}
                door_summary_dict[key]['value']['Total Door Area (m²)'] = 0
                door_summary_dict[key]['value']['Count'] = 0
                door_summary_dict[key]['value']['Number of openings'] = 0
                door_summary_dict[key]['value']['Number of openings draughtstripped'] = 0
            
            for keypart in ['Type', 'Door Type', 'Glazing Area (m\u00b2)', 'Glazing Type', 'Door U-Value (Wm2K)']:
                if keypart in door_dict[door]['value'].keys():
                    door_summary_dict[key]['value'][keypart] = door_dict[door]['value'][keypart] 
            
            door_summary_dict[key]['value']['Total Door Area (m²)'] += door_dict[door]['value']['Door Area [m2]']
            door_summary_dict[key]['value']['Total Door Area (m²)'] = round(door_summary_dict[key]['value']['Total Door Area (m²)'], 2)
            door_summary_dict[key]['value']['Count'] += 1
            door_summary_dict[key]['value']['Number of openings'] += door_dict[door]['value']['Number of openings']
            door_summary_dict[key]['value']['Number of openings draughtstripped'] += door_dict[door]['value']['Number of openings draughtstripped']
        
        
        
        
        output = door_summary_dict
        
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output

def window_summary(window_dict):
    try:
        
        # print('window_dict', ':')
        # pprint.pprint(window_dict)
        
        window_summary_dict = {}
        for window in window_dict:
            
            key = ''
            for keypart in ['Type', 'Description', 'In roof', 'Over shading', 'cardinal_direction']: # , 'No. of opes', 'No. of opes draught- stripped']:
                if keypart in window_dict[window]['value'].keys():
                    key += (str(window_dict[window]['value'][keypart]) + '_')
            key = str(hash(key))
            if key not in window_summary_dict.keys():
                window_summary_dict[key] = {}
                window_summary_dict[key]['value'] = {}
                window_summary_dict[key]['value']['Area [m2]'] = 0
                window_summary_dict[key]['value']['Count of the common windows'] = 0
                window_summary_dict[key]['value']['No. of opes'] = 0
                window_summary_dict[key]['value']['No. of opes draught- stripped'] = 0
            
            for keypart in ['Type', 'Description', 'Construction', 'In roof', 'Over shading', 'cardinal_direction', 'U-Value [W/m2K]']: # Openings? U-value?
                if keypart in window_dict[window]['value'].keys():
                    window_summary_dict[key]['value'][keypart] = window_dict[window]['value'][keypart] 
            
            # if 'cardinal_direction' not in window_summary_dict[key]['value'].keys():
                # print(window_summary_dict[key]['value'])
            # window_summary_dict[key]['value']['Orientation'] = window_summary_dict[key]['value']['cardinal_direction']
            
            window_summary_dict[key]['value']['Area [m2]'] += window_dict[window]['value']['Area [m2]']
            window_summary_dict[key]['value']['Area [m2]'] = round(window_summary_dict[key]['value']['Area [m2]'], 2)
            window_summary_dict[key]['value']['Count of the common windows'] += 1
            window_summary_dict[key]['value']['No. of opes'] += int(window_dict[window]['value']['No. of opes'])
            window_summary_dict[key]['value']['No. of opes draught- stripped'] += int(window_dict[window]['value']['No. of opes draught- stripped'])
        
        output = window_summary_dict
        
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output


def bulb_summary(bulb_dict):
    try:
        
        print('bulb_dict', ':')
        pprint.pprint(bulb_dict)
        
        bulb_summary_dict = {}
        for bulb in bulb_dict:
            
            key = ''
            for keypart in ['room_uid', 'name']: 
                if keypart in bulb_dict[bulb]['value'].keys():
                    key += (str(bulb_dict[bulb]['value'][keypart]) + '_')
            key = str(hash(key))
            if key not in bulb_summary_dict.keys():
                bulb_summary_dict[key] = {}
                bulb_summary_dict[key]['value'] = {}
                bulb_summary_dict[key]['value']['Count'] = 0
            
            # for keypart in ['Count', 'Room', 'Type', 'Description']:
            for keypart in ['Count', 'room_name', 'name', 'Description']:
                if keypart in bulb_dict[bulb]['value'].keys():
                    bulb_summary_dict[key]['value'][keypart] = bulb_dict[bulb]['value'][keypart] 
            
            bulb_summary_dict[key]['value']['Count'] += 1
            
        output = bulb_summary_dict
        
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output





def window_forms_append(object_dict, forms_uid_dict, window_detail_dict): # forms_uid_dict is not used in this function
    
    try:
        
        # print('object_dict', ':')
        # pprint.pprint(object_dict)
        
        for window in object_dict:
            object_dict[window]['value']['Room'] = object_dict[window]['value']['room_name']
            object_dict[window]['value']['Window height (m)'] = object_dict[window]['value']['height']
            object_dict[window]['value']['Window width (m)'] = object_dict[window]['value']['width']
            object_dict[window]['value']['Area [m2]'] = round(object_dict[window]['value']['width'] * object_dict[window]['value']['height'],2)
            
            # print("object_dict[window].keys()", ':', object_dict[window].keys())
            
            if 'Is the Rooflight Orientation Horizontal?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Is the Rooflight Orientation Horizontal?'] == True:
                    object_dict[window]['value']['Orientation'] = 'Horizontal'
                
            
            if 'Window Orientation' in object_dict[window]['value'].keys():
                object_dict[window]['value']['Orientation'] = object_dict[window]['value']['Window Orientation']
            
            
            object_dict[window]['value']['Type'] = 'Window Type 1'
            # print(window, ' is not Type 1')
            if 'Is the Rooflight Window Type 1?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Is the Rooflight Window Type 1?'] == False:
                    # print(window, ' is not Type 1')
                    object_dict[window]['value']['Type'] = object_dict[window]['value']['Other Rooflight Window Type']
            if 'Is the Window Type 1?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Is the Window Type 1?'] == False:
                    # print(window, ' is not Type 1')
                    object_dict[window]['value']['Type'] = object_dict[window]['value']['Other Window Type']
            
            
            object_dict[window]['value']['Construction'] = window_detail_dict[object_dict[window]['value']['Type']]
            
            if (object_dict[window]['value']['Type'] + ' Description') in window_detail_dict.keys():
                object_dict[window]['value']['Description'] = window_detail_dict[object_dict[window]['value']['Type'] + ' Description']
                
            
            
            if (object_dict[window]['value']['Type'] + ' U-Value (W/m2K)') in window_detail_dict.keys():
                object_dict[window]['value']['U-Value [W/m2K]'] = window_detail_dict[object_dict[window]['value']['Type'] + ' U-Value (W/m2K)']
            
            object_dict[window]['value']['In roof'] = True if object_dict[window]['value']['name'] == 'Skylight' else False
            
            object_dict[window]['value']['No. of opes'] = 1
            if 'Are the Number of Rooflight Openings required to be calculated?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Are the Number of Rooflight Openings required to be calculated?'] == True:
                    object_dict[window]['value']['No. of opes'] = object_dict[window]['value']['Number of Rooflight Openings']
            if 'Are the Number of Window Openings required to be calculated?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Are the Number of Window Openings required to be calculated?'] == True:
                    object_dict[window]['value']['No. of opes'] = object_dict[window]['value']['Number of Window Openings']
            
            object_dict[window]['value']['No. of opes draught- stripped'] = object_dict[window]['value']['No. of opes']
            if 'Are all of the Window Openings Draught-stripped?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Are all of the Window Openings Draught-stripped?'] == False:
                    object_dict[window]['value']['No. of opes draught- stripped'] = 'need exact text' # object_dict[window]['Other Window Type']
            
            
            
            object_dict[window]['value']['Over shading'] = 'Average or unknown (20% - 60% of sky blocked by obstacles)'
            if object_dict[window]['value']['symbol_name'] == 'Skylight':
                object_dict[window]['value']['Over shading'] = 'Very Little (<20% of sky blocked by obstacles)'
            
            if 'Is the Rooflight Shading Estimated to be Very Little (<20% of sky blocked by obstacles)?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Is the Rooflight Shading Estimated to be Very Little (<20% of sky blocked by obstacles)?'] == False:
                    object_dict[window]['value']['Over shading'] = object_dict[window]['value']['Other Rooflight Shading Estimate']
            
            if 'Is the Window Shading Estimated to be Average or unknown (20% - 60% of sky blocked by obstacles)?' in object_dict[window]['value'].keys():
                if object_dict[window]['value']['Is the Window Shading Estimated to be Average or unknown (20% - 60% of sky blocked by obstacles)?'] == False:
                    object_dict[window]['value']['Over shading'] = object_dict[window]['value']['Other Window Shading Estimate']
                    
            
        
        
        output = object_dict
    
    
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output

def door_forms_append(object_dict, forms_uid_dict): # shouldn't any part of the below that involves uid_dict be already covered by stats_append()?
    
    try:
        object_dict_2 = object_dict.copy()
        for door in object_dict:
            if object_dict[door]['value']["Is this door considered a heat loss door as per BER methodology?"] == False:
                del object_dict_2[door]
        object_dict = object_dict_2
            
        for door in object_dict:
            object_dict[door]['value']['Room'] = object_dict[door]['value']['room_name']
            object_dict[door]['value']['Door height (m)'] = object_dict[door]['value']['height']
            object_dict[door]['value']['Door width (m)'] = object_dict[door]['value']['width']
            object_dict[door]['value']['Door Area [m2]'] = round(object_dict[door]['value']['width'] * object_dict[door]['value']['height'], 2)
            
            
            # for f in ['Type', 'Description', 'Draught Stripped', 'U-Value [W/m2K]', 'Glazing Area (m²)', 'Glazing Type']:
            for f in ['Type', 'Draught Stripped', 'U-Value [W/m2K]', 'Glazing Area (m²)', 'Glazing Type']:
                object_dict[door]['value'][f] = ''
            
            
            # print("object_dict[door]['value']", ':')
            # pprint.pprint(object_dict[door]['value'])
            
            
            for key in object_dict[door]['value']:
                if 'Door Type' in key:
                    object_dict[door]['value']['Type'] = object_dict[door]['value'][key]
                # if 'Description' in key:
                    # object_dict[door]['value']['Description'] = object_dict[door]['value'][key]
                if 'Is the Door Opening Draught Stripped?' in key:
                    object_dict[door]['value']['Draught Stripped'] = object_dict[door]['value'][key]
                if 'U-Value' in key:
                    object_dict[door]['value']['U-Value [W/m2K]'] = object_dict[door]['value'][key]
                if 'Area of glazing' in key:
                    object_dict[door]['value']['Glazing Area (m²)'] = object_dict[door]['value'][key]
                if 'Glazing Type' in key:
                    object_dict[door]['value']['Glazing Type'] = object_dict[door]['value'][key]
            
            object_dict[door]['value']['Number of openings'] = 1
            if 'Draught Stripped' in object_dict[door]['value'].keys():
                if object_dict[door]['value']['Draught Stripped'] == True:
                    object_dict[door]['value']['Number of openings draughtstripped'] = 1
                else:
                    object_dict[door]['value']['Number of openings draughtstripped'] = 0
        
        
        output = object_dict
    
    
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output








def XML_2_dict_new(root, t = "floor"):
    '''
    Returns [xml_ref_dict, # links floors and rooms with their uids
            , nwa_dict # net wall area
            , obj_dict
            , xml_val_dict # includes any variables required as output
            , 
            ]
        
    '''
    try:
        # d = {}
        xml_ref_dict = {}
        nwa_dict = {}
        obj_dict = {}
        xml_val_dict = {}
        est_dict = {}
        storey_height_dict = {}
        
        
        xml_val_dict['project_id'] = root.get('id')
        
        project_name = root.get('name')
        if project_name[-1] == ' ':
            project_name = project_name[:-1]
        xml_val_dict['project_name'] = project_name
        
        project_address = ''
        address_fields = ['street', 'city', 'province', 'country', 'postalCode']
        for af in address_fields:
            f = root.get(af)
            if f is not None:
                project_address = (project_address + ', ' + str(f)) if project_address != '' else str(f)
        xml_val_dict['project_address'] = project_address
        
        
        xml_val_dict['eircode'] = root.get('postalCode')
        
        
        
        
        MagicPlan_2_SEAI_dict = {
            "author": "Surveyor"
            , "date": "survey_date"
            , "notes": "project_notes"
            , "qf.34d66ce4q3": "rating_type"
            , "qf.34d66ce4q4": "rating_purpose"
            # , "qf.34d66ce4q6": "age_dwelling"
            # , "qf.34d66ce4q7": "age_extension_1"
            # , "qf.57124e1fq101": "age_dwelling_ALT"
            # , "qf.57124e1fq102": "age_extension_1_ALT"
            }
        # note MagicPlan also has a separate "Surveyor" field ("qf.34d66ce4q1") but "author" is the one used for SEAI survey purposes
        # ToDo: confirm if also the case for BER
        
        values = root.findall('values/value')
        for mpk in MagicPlan_2_SEAI_dict:
            xml_val_dict[MagicPlan_2_SEAI_dict[mpk]] = ''
            for value in values:
                k = value.attrib["key"]
                if k == mpk:
                    xml_val_dict[MagicPlan_2_SEAI_dict[mpk]] = value.text
            print(MagicPlan_2_SEAI_dict[mpk], ':', xml_val_dict[MagicPlan_2_SEAI_dict[mpk]])
        
        # w = {}
        wd_list = ['634004d284d12@edit:0063fa41-fa2d-4493-9f86-dcd0263e8108'
                    , '634004d284d12@edit:0ecdca7d-a4c3-4692-893a-89e6eaa76e74'
                    , '634004d284d12@edit:28960da1-84f6-4f3b-a446-7c72b9febe9f'
                    , '634004d284d12@edit:28b0fb8c-47a4-4d9e-8ce5-2b35a1a0404e'
                    , '634004d284d12@edit:2b72a58f-7380-4b6c-9d74-667f937a9b57'
                    , '634004d284d12@edit:32b043c7-432a-409f-972d-a75b386b1789'
                    , '634004d284d12@edit:60194a47-84ce-414b-8368-69ec53167111'
                    , '634004d284d12@edit:6976cc78-3a2e-4935-99c6-6aff8011be8a'
                    , '634004d284d12@edit:735122f1-ab8b-47e8-b5ca-d4ec4d492f1c'
                    , '634004d284d12@edit:7d851726-6ff6-48f7-8371-9ea09bd5179f'
                    , '634004d284d12@edit:7f6101da-4b6d-4c31-9293-d59552aeff3a'
                    , '634004d284d12@edit:a9a0a953-0fd3-4733-b161-de4f08fe5d49'
                    , '634004d284d12@edit:e6026a1e-3089-4fe7-9ec4-8504b001eb2e'
                    , '634004d284d12@edit:fc02c0c5-d9d8-4679-8a77-dc75edf7f592'
                    , 'arcdoor'
                    , 'doorbypass'
                    , 'doorbypassglass'
                    , 'doordoublefolding'
                    , 'doordoublehinged'
                    , 'doordoublesliding'
                    , 'doorfolding'
                    , 'doorfrench'
                    , 'doorgarage'
                    , 'doorglass'
                    , 'doorhinged'
                    , 'doorpocket'
                    , 'doorsliding'
                    , 'doorslidingglass'
                    , 'doorswing'
                    , 'doorwithwindow'
                    , 'windowarched'
                    , 'windowawning'
                    , 'windowbay'
                    , 'windowbow'
                    , 'windowcasement'
                    , 'windowfixed'
                    , 'windowfrench'
                    , 'windowhopper'
                    , 'windowhung'
                    , 'windowsliding'
                    , 'windowtrapezoid'
                    , 'windowtriangle'
                    , 'windowtskylight1'
                    , 'windowtskylight2'
                    , 'windowtskylight3'
                    , 'compass'
                    , '634004d284d12@edit:775a214c-7656-4548-ab5d-2c490583a32b' # cloned from compass
                    ]
        
        xml_ref_dict['habitable_rooms'] = []
        xml_ref_dict['wet_rooms'] = []
        xml_ref_dict['exclude_rooms'] = []
        
        xml_ref_dict['exclude_room_types'] = ['Attic'
                                            , 'Balcony'
                                            , 'Storage'
                                            , 'Patio'
                                            , 'Deck'
                                            , 'Porch'
                                            , 'Cellar'
                                            , 'Garage'
                                            , 'Furnace Room'
                                            , 'Outbuilding'
                                            , 'Unfinished Basement'
                                            , 'Workshop'
                                            ]
        
        xml_ref_dict['habitable_room_types'] = ['Kitchen'
                                                , 'Dining Room'
                                                , 'Living Room'
                                                , 'Bedroom'
                                                , 'Primary Bedroom'
                                                , "Children's Bedroom"
                                                , 'Study'
                                                , 'Music Room'
                                                ]
        xml_ref_dict['wet_room_types'] = ['Kitchen'
                                        , 'Bathroom'
                                        , 'Half Bathroom'
                                        , 'Laundry Room'
                                        , 'Toilet'
                                        , 'Primary Bathroom'
                                        ]

        
        
        
        
        colours_dict : dict[str, list[float]] = {}
        colour_hex_2_name_dict = {
                                    # "#eecaffff" : "Blue Chalk"
                                    # , "#ffcc01ff" : "Yellow"
                                    # , "#adadadff" : "Gray68"
                                    # , "#ffecc5ff" : "Bisque"
                                    # , "#ff3a2fff" : "Red Orange"
                                    # , "#ff624fff" : "Tomato"
                                    # , "#32c659ff" : "Green"
                                    # , "#f5ec00ff" : "Golden Yellow"
                                    }

        
        
        floors = root.findall('interiorRoomPoints/floor')
        print('len(floors)', ':', len(floors))
        for floor in floors:
            ft = floor.get('floorType')
            floor_uid = floor.get('uid')
            xml_ref_dict[floor.get('floorType')] = floor.get('uid')
            xml_ref_dict[floor.get('uid')] = floor.get('floorType')
            nwa_dict[ft] = {}
            
            
            
            # room_points = floor.findall('floorRoom/point/values/value[@key="qcustomfield.e8660a0cq0.lo6b23iucno"]../../..')
            
            for room in floor.findall('floorRoom/values/value[@key="ground.color"]../..'):
                colour = room.find('values/value[@key="ground.color"]').text
                area = float(room.get('area'))
                if colour in colour_hex_2_name_dict.keys():
                    colour = colour_hex_2_name_dict[colour]
                # print('colour', ':', colour)
                if colour not in colours_dict:
                    colours_dict[colour] = {}
                    colours_dict[colour]['value'] = {}
                    colours_dict[colour]['value'][floor_uid] = 0.00
                if floor_uid not in colours_dict[colour]['value']:
                    colours_dict[colour]['value'][floor_uid] = 0.00
                
                colours_dict[colour]['value'][floor_uid] += area
                colours_dict[colour]['value'][floor_uid] = round(colours_dict[colour]['value'][floor_uid], 2)
            
            for room in floor.findall('floorRoom'):
                if room.get('type') not in xml_ref_dict.keys():
                    xml_ref_dict[room.get('type')] = []
                xml_ref_dict[room.get('type')].append(room.get('uid'))
                xml_ref_dict[room.get('uid')] = room.get('type')
                # print(room.get('type'))
                if room.get('type') in xml_ref_dict['habitable_room_types']:
                    xml_ref_dict['habitable_rooms'].append(room.get('uid'))
                    xml_ref_dict['habitable_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid'))
                
                if room.get('type') in xml_ref_dict['wet_room_types']:
                    xml_ref_dict['wet_rooms'].append(room.get('uid'))
                    xml_ref_dict['wet_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid'))
                
                if room.get('type') in xml_ref_dict['exclude_room_types']:
                    xml_ref_dict['exclude_rooms'].append(room.get('uid'))
                    xml_ref_dict['exclude_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid') + " (" + room.get('area') + ")")
                # else:
                    # xml_ref_dict['include_rooms'].append(room.get('uid'))
                    # xml_ref_dict['include_rooms'].append('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid') + " (" + room.get('area') + ")")
                
                # print('exclude_rooms', ':', xml_ref_dict['exclude_rooms'])
                
                
                
                for value in room.findall('values/value'):
                    key = value.get('key')
                    # print(key)
                    if key == "qcustomfield.2979903aq1": # Include?
                        # print(room.get('type'))
                        floor_area_include = value.text
                        # print('floor_area_include', ':', floor_area_include)
                        # if floor_area_include == '0':
                            # xml_ref_dict['exclude_rooms'].append(room.get('uid'))
                        if floor_area_include == '1':
                            if room.get('uid') in xml_ref_dict['exclude_rooms']:
                                xml_ref_dict['exclude_rooms'].remove(room.get('uid'))
                                xml_ref_dict['exclude_rooms'].remove('floor ' + ft + " - " + room.get('type') + " - " + room.get('uid') + " (" + room.get('area') + ")")
                                # print(xml_ref_dict['exclude_rooms'])
                                # print(room.get('type'))
                
                # print('exclude_rooms', ':', xml_ref_dict['exclude_rooms'])
                
                rt = room.get('type') + ' (' + room.get('uid') + ')'
                x = {}
                room_x = room.get('x')
                room_y = room.get('y')
                w_index = 0
                for point in room.findall('point'):
                    w_index += 1
                    # uid = point.get('uid')
                    x[w_index] = {}
                    for value in point.findall('values/value'):
                        if value.get('key') in ['qf.c52807ebq1', 'qf.bdbaf056q1', 'qf.c52807ebq1']:
                            x[w_index]['type'] = value.text
                    # if 'type' not in list(x[w_index].keys()):
                        # x.pop(w_index)
                        # continue
                    x[w_index]['uid'] = point.get('uid')
                    x[w_index]['x1'] = float(point.get('snappedX')) + float(room_x)
                    x[w_index]['y1'] = -float(point.get('snappedY')) - float(room_y)
                    x[w_index]['h'] = point.get('height')
                    for value in point.findall('values/value'):
                        if value.get('key') == "loadBearingWall":
                            # print("loadBearingWall", ':', value.text)
                            x[w_index]['loadBearingWall'] = value.text
                # print('ft', ':', ft)
                # print('rt', ':', rt)
                # print('x', ':', x)
                # print('len(x)', ':', len(x))
                
                        
                w_index = 0
                for wall in x:
                    w_index += 1
                    # print(list(x[1].keys()))
                    if w_index + 1 in list(x.keys()):
                        x[w_index]['x2'] = x[w_index + 1]['x1']
                        x[w_index]['y2'] = x[w_index + 1]['y1']
                    else:
                        x[w_index]['x2'] = x[1]['x1']
                        x[w_index]['y2'] = x[1]['y1']
                    x[w_index]['l'] = cart_distance((x[w_index]['x1'], x[w_index]['y1']), (x[w_index]['x2'], x[w_index]['y2']))
                    x[w_index]['a'] = float(x[w_index]['l']) * float(x[w_index]['h'])
                
                y = {}
                for wall in x:
                    uid = x[wall]['uid']
                    y[uid] = {}
                    if 'type' in list(x[wall].keys()):
                        y[uid]['type'] = x[wall]['type']
                    y[uid]['x1'] = x[wall]['x1']
                    y[uid]['y1'] = x[wall]['y1']
                    y[uid]['x2'] = x[wall]['x2']
                    y[uid]['y2'] = x[wall]['y2']
                    y[uid]['h'] = x[wall]['h']
                    y[uid]['l'] = x[wall]['l']
                    y[uid]['a'] = x[wall]['a']
                    # print(list(x[wall].keys()))
                    if 'loadBearingWall' in list(x[wall].keys()):
                        y[uid]['loadBearingWall'] = x[wall]['loadBearingWall']
                        
                
                # print('len(y)', ':', len(y))
                # print('y', ':', y)
                # print('adding wall dict y for room ' + rt + ' to nwa_dict')
                nwa_dict[ft][rt] = y
                
        # print('nwa_dict', ':')
        # pprint.pprint(nwa_dict)
        
        # print("xml_ref_dict['exclude_rooms']", ':', str(xml_ref_dict['exclude_rooms']))
        # print("xml_ref_dict['include_rooms']", ':', str(xml_ref_dict['include_rooms']))
        
        
        pi = 3.14159
        offset = 0
        for floor in floors:
            ft = floor.get("floorType")
            if ft == "0":
                for si in floor.findall("symbolInstance"):
                    s = si.get("symbol")
                    if s in ['compass', '634004d284d12@edit:775a214c-7656-4548-ab5d-2c490583a32b']:
                        siid = si.get("id")
                        for f in floor.findall("furniture"):
                            si2 = f.get("symbolInstance")
                            if si2 == siid:
                                offset = f.get("angle")
        
        # print('offset', ':', offset)
        
        
        
        # Create Object Dictionary 
            # - first get list of all objects on each floor
            # - then add any additional details available from "exploded" section (linked via "id" e.g. "W-1-5")
            # - is compass included? it is now but a better more rigorous solution should be found
            # - also need to loop through each window/door in each floorRoom to associate room_x, room_y 
        wo = {}
        floors = root.findall('floor')
        for floor in floors:
            ft = floor.get('floorType')
            uid = floor.get('uid')
            storey_height_dict[ft] = {}
            storey_height_dict[ft]['value'] = {}
            storey_height_dict[ft]['value']['uid'] = uid
            # storey_height_dict[ft]['value']['storey_height'] = xml_val_dict['storey_height']
            
            storey_height_dict[uid] = {}
            storey_height_dict[uid]['value'] = {}
            storey_height_dict[uid]['value']['floor_type'] = ft
            # storey_height_dict[uid]['value']['storey_height'] = xml_val_dict['storey_height']
            
            for p in floor.findall('symbolInstance'):
                if p.get('symbol') == 'floor':
                    for value in p.findall('values/value'):
                        if value.get('key') == 'qcustomfield.5d0165e3q1':
                            storey_height_dict[uid]['value']['use_floor_level_height'] = value.text
                            # if value.text == "0"
                                # del storey_height_dict[uid]
            
            
            o = {}
            total_surface = 0
            for p in floor.findall('symbolInstance'):
                # print("p.get('uid')", ':', p.get('uid'))
                if p.get('symbol') in wd_list:
                    # print("p.get('symbol')", ':', p.get('symbol'))
                    id = p.get('id')
                    o[id] = {} # We do need id to get the other info but does it need to be the top level? See est_dict below
                    o[id]['uid'] = p.get('uid')
                    o[id]['symbol'] = p.get('symbol')
                
                if p.get('isEstimated') == '1':
                    for value in p.findall('values/value'):
                        if value.get('key') == 'sku':
                            sku = value.text
                            sku = sku.replace('Wall type', 'Wall Type')
                            sku = sku.replace('Semi exposed', 'Semi-Exposed')
                            if sku not in est_dict.keys():
                                est_dict[sku] = {}
                        if value.get('key') == 'totalsurface':
                            total_surface = value.text
                    # est_dict[sku]['floor_type'] = ft
                    est_dict[sku][ft] = total_surface
                    if 'total_surface' not in est_dict[sku].keys():
                        est_dict[sku]['total_surface'] = 0 
                    est_dict[sku]['total_surface'] += float(total_surface)
                    
                    for lt in p.findall('linkedTo'):
                        uid = lt.get('uid')
                        if uid not in est_dict.keys():
                            est_dict[uid] = {}
                        est_dict[uid]['floor_type'] = ft
                        
                        for value in p.findall('values/value'):
                            if value.get('key') == 'sku':
                                est_dict[uid]['sku'] = value.text
                            if value.get('key') == 'totalsurface':
                                est_dict[uid]['total_surface'] = value.text
                            
                    
            
            
            for p in floor.findall('furniture'):
                si = p.get('symbolInstance')
                if si in list(o.keys()): # currently "compass" is the only furniture included in o
                    o[si]['angle'] = p.get('angle')
                
            for p in floor.findall('exploded/door'):
                si = p.get('symbolInstance')
                # print('si', ':', si)
                if si in list(o.keys()):
                # o[si] = {}
                # o[si]['symbolInstance'] = window.get('symbolInstance')
                    o[si]['x1'] = p.get('x1')
                    o[si]['y1'] = -float(p.get('y1'))
                    o[si]['x2'] = p.get('x2')
                    o[si]['y2'] = -float(p.get('y2'))
                    o[si]['w'] = p.get('width')
                    o[si]['d'] = p.get('depth')
                    o[si]['h'] = p.get('height')
                    o[si]['a'] = float(o[si]['w']) * float(o[si]['h'])
            
            for p in floor.findall('exploded/window'):
                # o_index += 1
                si = p.get('symbolInstance')
                # print('si', ':', si)
                if si in list(o.keys()):
                # o[si] = {}
                # o[si]['symbolInstance'] = window.get('symbolInstance')
                    o[si]['x1'] = p.get('x1')
                    o[si]['y1'] = -float(p.get('y1'))
                    o[si]['x2'] = p.get('x2')
                    o[si]['y2'] = -float(p.get('y2'))
                    o[si]['w'] = p.get('width')
                    o[si]['d'] = p.get('depth')
                    o[si]['h'] = p.get('height')
                    o[si]['a'] = float(o[si]['w']) * float(o[si]['h'])
            
            
            
            
            # print('o', ':')
            # pprint.pprint(o)
            
            # obj_dict[ft] = o
            
            for room in floor.findall('floorRoom'):
                uid = room.get('uid')
                rt = room.get('type')
                storey_height_dict[uid] = {}
                storey_height_dict[uid]['value'] = {}
                storey_height_dict[uid]['value']['room_type'] = rt
                storey_height_dict[uid]['value']['floor_type'] = ft
                for value in room.findall('values/value'):
                    if value.get('key') == "ceilingHeight":
                        storey_height_dict[uid]['value']['ceiling_height'] = value.text
                    if value.get('key') == "qcustomfield.2979903aq1":
                        storey_height_dict[uid]['value']['thermal_envelope'] = value.text
                    if value.get('key') == "qcustomfield.347f643dq1":
                        storey_height_dict[uid]['value']['thermal_envelope'] = value.text
                
                
                rt = room.get('type') + ' (' + room.get('uid') + ')'
                
                w = {}
                room_x = room.get('x')
                room_y = room.get('y')
                
                for window in room.findall('window'):
                    si = window.get('symbolInstance')
                    if si in o.keys():
                        o[si]['room_x'] = room_x
                        o[si]['room_y'] = room_y
                    else:
                        print(si, 'not found in o.keys()')
                        # print('o', ':')
                        # pprint.pprint(o)
                
                w_index = 0
                for point in room.findall('point'): # get (x3, y3)
                    w_index += 1
                    w[w_index] = {}
                    w[w_index]['uid'] = point.get('uid')
                    w[w_index]['x3'] = float(point.get('snappedX')) + float(room_x)
                    w[w_index]['y3'] = -float(point.get('snappedY')) - float(room_y)

                w_index = 0
                for wall in w: # get (x4, y4), the second point in each line segment - WARNING: relies on the assumption that the points are in order
                    w_index += 1
                    if w_index + 1 in list(w.keys()):
                        w[w_index]['x4'] = w[w_index + 1]['x3']
                        w[w_index]['y4'] = w[w_index + 1]['y3']
                    else:
                        w[w_index]['x4'] = w[1]['x3']
                        w[w_index]['y4'] = w[1]['y3']
                
                # print('ft', ':', ft)
                # print('rt', ':', '"' + rt + '"')
                # print('w', ':')
                # pprint.pprint(w)
                
                for wall in w: # transfer values to nwa_dict (where wall key is "uid" instead of numbered index)
                    uid = w[wall]['uid']
                    nwa_dict[ft][rt][uid]['x3'] = w[wall]['x3']
                    nwa_dict[ft][rt][uid]['y3'] = w[wall]['y3']
                    nwa_dict[ft][rt][uid]['x4'] = w[wall]['x4']
                    nwa_dict[ft][rt][uid]['y4'] = w[wall]['y4']
                
                y = nwa_dict[ft][rt] # for brevity
                
                w_index = 0
                for wall in y:
                    w_index += 1
                    # print('wall', ':', wall)
                    
                    y[wall]['windows'] = []
                    y[wall]['net_a'] = y[wall]['a']
                    y[wall]['total_window_a'] = 0
                    for window in o:
                        if 'x1' not in list(o[window].keys()):
                            continue
                        if 'x3' not in list(y[wall].keys()):
                            continue
                        # print('window', ':')
                        # pprint.pprint(o[window])
                        if linear_subset(float(o[window]['x1']), float(o[window]['y1']), float(o[window]['x2']), float(o[window]['y2']), float(y[wall]['x3']), float(y[wall]['y3']), float(y[wall]['x4']), float(y[wall]['y4'])) == True:
                            y[wall]['windows'].append(window + ' (' + str(o[window]['a']) + ')')
                            y[wall]['net_a'] -= o[window]['a']
                            y[wall]['total_window_a'] += o[window]['a']
                            
                            # print('object ' + str(window) + ' (' + str(o[window]['x1']) + '\t' + str(o[window]['y1']) + ') -> (' + str(o[window]['x2']) + '\t' + str(o[window]['y2']) + ') is colinear with wall ' + str(wall) + ' (' + str(y[wall]['x3']) + '\t' + str(y[wall]['y3']) + ') -> (' + str(y[wall]['x4']) + '\t' + str(y[wall]['y4']) + ')')
                            # print('yes')
                    # print("w[wall]['windows']", ':', w[wall]['windows'])
                
                # print('y', ':', y)
                nwa_dict[ft][rt] = y
                
            # print('o', ':')
            # pprint.pprint(o)
            
            # pi = 3.14159
            # offset = 0
            # for window in o:
                # if o[window]['symbol'] in ['compass', '634004d284d12@edit:775a214c-7656-4548-ab5d-2c490583a32b']:
                    # print("o[window]", ':')
                    # pprint.pprint(o[window])
                    # offset = float(o[window]['angle'])
                    # print('offset', ':', offset)
            
            for window in o:
                if o[window]['symbol'] == 'compass':
                    continue
                if 'x2' not in o[window].keys():
                    continue
                try:
                    if round(float(o[window]['x2']), 2) == round(float(o[window]['x1']), 2):
                        o[window]['angle'] = pi / 2
                        if float(o[window]['y2']) < float(o[window]['y1']):
                            o[window]['slope'] = 'infinite'
                        else:
                            o[window]['slope'] = 'infinite_neg'
                            o[window]['angle'] += pi
                        
                    else:
                        o[window]['slope'] = (round(float(o[window]['y2']), 2) - round(float(o[window]['y1']), 2))/(round(float(o[window]['x2']), 2) - round(float(o[window]['x1']), 2))
                except:
                    o[window]['slope'] = ''
                    print(traceback.format_exc())
                
                # if 'slope' not in o[window].keys():
                    # print('o[' + window + ']', ':')
                    # pprint.pprint(o[window])
                
                if o[window]['slope'] not in ['', 'infinite', 'infinite_neg']:
                    o[window]['angle'] = round(float(math.atan(o[window]['slope'])), 2)
                
                    if float(o[window]['x2']) > float(o[window]['x1']):
                        o[window]['angle'] += pi
                
                
                
                
                # o[window]['angle_real'] = round((round(float(o[window]['angle']), 2) + round(float(offset), 2)) % round(pi * 2, 2), 2)
                o[window]['angle_real'] = round((round(float(o[window]['angle']), 2) + round(float(offset), 2)), 2)
                
                cardinal_directions = ['S', 'SE', 'E', 'NE', 'N', 'NW', 'W', 'SW']
                o[window]['angle_real_degrees'] = o[window]['angle_real'] * (180 / pi)
                o[window]['cardinal_index'] = int(round(o[window]['angle_real_degrees'] / 45, 0) % 8)
                
                o[window]['cardinal_direction'] = cardinal_directions[o[window]['cardinal_index']]
                
                wo[o[window]['uid']] = o[window]['cardinal_direction']
        
        
        
                
    except Exception as ex:
        output = str(ex) + "\n\n" + traceback.format_exc()
        # LOGGER.info('Exception : ' + str(traceback.format_exc()))
        print(output)
    
    finally:
        return xml_ref_dict, nwa_dict, xml_val_dict, colours_dict, wo, est_dict, storey_height_dict


def get_forms_data(id, headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"
            ,"key": "45170e50321733db78952dfa5901b0dfeeb8"
            , "customer": "63b5a4ae69c91"
            , "accept": "application/json"
            }):
    
    try:
        output = {}
        form_val_dict = {}
        forms_full_dict = {}
        forms_uid_dict = {}
        missing_vals = {}
        window_detail_dict = {}
        wall_type_dict = {}
        floor_type_dict = {} # pick up floor details from every room - then group based on [age band, ...]
        heating_dict = {}
        
        # wall_type_vals_dict = {}

        json_url = "https://cloud.magicplan.app/api/v2/plans/forms/" + id
        request = urllib.request.Request(json_url, headers=headers)
        
        JSON = urllib.request.urlopen(request).read()
        print('type(JSON)', ':', type(JSON))
        if not type(JSON) is bytes:
            raise Exception(JSON)
        
        JSON = json.loads(JSON)
        print('type(JSON)', ':', type(JSON))
        if not isinstance(JSON, dict):
            raise Exception(JSON)
        

        for datum in JSON["data"]:
            if datum["symbol_instance_id"] not in forms_uid_dict.keys():
                forms_uid_dict[datum["symbol_instance_id"]] = {}
                forms_uid_dict[datum["symbol_instance_id"]]["symbol_name"] = datum["symbol_name"]
            if datum["symbol_type"] not in forms_full_dict.keys():
                forms_full_dict[datum["symbol_type"]] = {}
            if datum["symbol_name"] not in forms_full_dict[datum["symbol_type"]].keys():
                forms_full_dict[datum["symbol_type"]][datum["symbol_name"]] = {}
            forms_full_dict[datum["symbol_type"]][datum["symbol_name"]]['uid'] = datum["symbol_instance_id"]
                
            for form in datum["forms"]:
                # print(form["title"])
                for section in form["sections"]:
                    if form["title"] == "b. Building | Walls": # BER only
                        if section["name"] == "":
                            continue
                        wall_type_dict[section["name"]] = {}
                        wall_type_dict[section["name"] + ' - Semi-Exposed'] = {}
                    
                    if form["title"] == "BER Floor Details": # BER only
                        # if section["name"] == "":
                            # continue
                        floor_type_dict[datum["symbol_instance_id"]] = {}
                    
                    if form["title"] == "BER Space Heating": # BER only
                        if datum["symbol_instance_id"] not in heating_dict:
                            heating_dict[datum["symbol_instance_id"]] = {}
                            heating_dict[datum["symbol_instance_id"]]['Object Name'] = datum["symbol_name"]
                        # print('Object Name', ':', datum["symbol_name"])
                    
                    for field in section["fields"]:
                        im = field["label"].replace(' *', '')
                        im = im.replace('*', '')
                        v = ''
                        if field["value"]["value"] == None:
                            vals = []
                            vals = [val["value"] for val in field["value"]["values"]]
                            for val in vals:
                                v += val
                                v += '<BR>'
                        else:
                            v = field["value"]["value"]
                        
                        if im not in form_val_dict.keys():
                            form_val_dict[im] = {}
                        key = 'value'
                        if key not in form_val_dict[im].keys():
                            form_val_dict[im][key] = v
                        
                        forms_full_dict[datum["symbol_type"]][datum["symbol_name"]][im] = v
                        forms_uid_dict[datum["symbol_instance_id"]][im] = v
                        
                        # if form["title"] == "BER Space Heating":
                            # print(im, ':', v)
                        
                        
                        
                        if field["is_required"] == True and field["value"]["has_value"] == False:
                            missing_vals[datum["symbol_name"]] = im
                            
                        if form["title"] == "c. Building | Windows":
                            if im not in window_detail_dict.keys():
                                window_detail_dict[im] = {}
                            window_detail_dict[im] = v
                        if form["title"] == "b. Building | Walls":
                            wall_type_dict[section["name"]][im] = v
                        if form["title"] == "BER Floor Details":
                            floor_type_dict[datum["symbol_instance_id"]][im] = v
                        if form["title"] == "BER Space Heating":
                            heating_dict[datum["symbol_instance_id"]][im] = v
                
                        # if datum["symbol_instance_id"] == '65e0b636.65c3ebff':
                            # print('heating_dict[datum["symbol_instance_id"]]', ':')
                            # pprint.pprint(heating_dict[datum["symbol_instance_id"]])
            
            
                            
        # print('wall_type_dict', ':')
        # pprint.pprint(wall_type_dict)
        # print('heating_dict', ':')
        # pprint.pprint(heating_dict)
        
        
        
        output = {}
        output['form_val_dict'] = form_val_dict
        output['forms_full_dict'] = forms_full_dict
        output['forms_uid_dict'] = forms_uid_dict
        output['window_detail_dict'] = window_detail_dict
        output['missing_vals'] = missing_vals
        # output['wall_type_vals_dict'] = wall_type_vals_dict
        output['wall_type_dict'] = wall_type_dict
        output['floor_type_dict'] = floor_type_dict
        output['heating_dict'] = heating_dict
    
    except:
        output = traceback.format_exc()
        print('exception', ':', output)
    
    return output



def distributor_function(form, root = ''):
    '''
    # Processing steps common to all projects are performed first
        # Establish fundamental parameters
            # plan_name (name given to the project by the user)
            # email (of the user who exported the project)
            # xml (a url where the project XML data can be retrieved)
        # Retrieve XML data
            # Load as dict
            # Get the MagicPlan Project ID (ToDo: check is this available directly from the form?)
        
        # Retrieve API data (forms, statistics, files) # Need forms here to get "This project is a" but others could probably wait... don't forget "get_project_files" returns a list that is currently included in the email body HTML tables so probably leave that where it is in the project-specific function
            # Load as dict x 3
            # ToDo: avoid any potential Key Errors by giving these all a default value of ''
        # Identify what type of project it is
    
    # Project-specific processing is then carried out by dedicated functions:
        # survey()
        # BER()
    # These return a HTML formatted table as output which will appear as the email body
    
    # Return "json_data" to be uploaded to Azure blob storage where it will be processed by the Logic App
    '''
    
    try:

        output = ''
        if root == '':
            email = form['email']
            xml = form['xml']
            root : ET.Element # why is this line necessary? Does this initialise the variable? Does it enforce the type?
            with urllib.request.urlopen(xml) as f:
                s = f.read().decode('utf-8')
            root = dET.fromstring(s)
        else:
            email = 'gtsupport@ie.gt.com'
            xml = ''

        project_name = root.get('name')
        if project_name[-1] == ' ':
            project_name = project_name[:-1]
        
        
        
        val_dict = {}
        val_dict["plan_name"] = project_name
        
        
        
        
        project_id = root.get('id')
        print('project_id', ':', project_id)
        
        forms_data = get_forms_data(project_id)
        if isinstance(forms_data, dict):
            if 'form_val_dict' in forms_data.keys():
                form_val_dict = forms_data['form_val_dict']
            # forms_full_dict = forms_data['forms_full_dict']
            if 'missing_vals' in forms_data.keys():
                missing_vals = forms_data['missing_vals']
                print('missing_vals', ':', missing_vals)
        
        
        
            if "This project is a" in form_val_dict.keys():
                for pt in form_val_dict["This project is a"]:
                    project_type = form_val_dict["This project is a"][pt]
                print("This project is a", ':', project_type)
                if project_type == "BER":
                    template = "template_ber"
                    populate_template_new(val_dict, template)
                    output = BER(root, email=email, forms_data=forms_data)
                if project_type == "Survey":
                    output = survey(root)
        
        # if output == '':
            # output = survey(root)
        
        output = output + '<h2>' + xml + '</h2></div>'
        
        # populate_template_new(xml_val_dict, 'template')

        json_data = json.dumps({
            'email' : email
            , 'name'  : project_name
            , 'table' : output
        })
        output = json_data
    except:
        output = traceback.format_exc()
        print(output)

    return output


def BER(root, output = '', email = '', forms_data = {}):
    # read template xlsx (xlst?)
    # create dicts of required table/tab contents
    # populate the dicts
    # use the dicts to populate the template and save it to be sent as an email attachment
    # use the dicts to generate this function's "output" HTML to serve as the body of the return email
    
    try:
        # lists of MagicPlan room names to be included/excluded by default in the thermal envelope:
        thermal_envelope = [ # do we need this? We can just assume anything not on the other list is on this one (unless explicitly told otherwise, since the default is "included")
                            'Archives'
                            , 'Bathroom'
                            , 'Bedroom'
                            , 'Cafeteria'
                            , 'Children Bedroom'
                            , 'Closet'
                            , 'Conference Room'
                            , 'Den'
                            , 'Dining Room'
                            , 'Elevators'
                            , 'Half Bathroom'
                            , 'Hall'
                            , 'Hallway'
                            , 'Hatched Room'
                            , 'Kitchen'
                            , 'Kitchenette'
                            , 'Lab'
                            , 'Laundry Room'
                            , 'Living Room'
                            , 'Lounge'
                            , 'Maintenance Room'
                            , 'Meeting Room'
                            , 'Music Room'
                            , 'Open Space'
                            , 'Other'
                            , 'Photocopy Room'
                            , 'Playroom'
                            , 'Porch'
                            , 'Primary Bathroom'
                            , 'Primary Bedroom'
                            , 'Private Office'
                            , 'Reception'
                            , 'Restrooms'
                            , 'Server Room'
                            , 'Shared Office'
                            , 'Stairway'
                            , 'Storage'
                            , 'Study'
                            , 'Toilet'
                            , 'Training Room'
                            , 'Vestibule'
                            , 'Waiting Room'
                            ]
        
        ex_thermal_envelope = [
                                'Attic/Loft'
                                , 'Balcony'
                                , 'Cellar'
                                , 'Deck'
                                , 'Furnace Room'
                                , 'Garage'
                                , 'Outbuilding'
                                , 'Patio'
                                , 'Unfinished Basement'
                                , 'Workshop'
                                ]
        
        # true_floors = range(-2, 8)
        # print('true_floors', ';', true_floors)
        
        
        
        project_id = root.get("id")
        project_name = root.get("name") # ToDo: are we going to pass in here xml_dict or do we need to produce a project-specific one?
        xml_dict = XML_2_dict_new(root)
        
        # if not isinstance(output_dict, dict):
            # print('not a dict:', output_dict)
            # raise Exception(output_dict)
        
        
        # *****************************
        # Update the below - let XML_dict = XML_2_dict_new() be a single dictionary instead of the current tuple of dictionaries
        # then add Error Handling
        
        xml_ref_dict = xml_dict[0]
        nwa_dict = xml_dict[1]
        xml_val_dict = xml_dict[2]
        colours_dict = xml_dict[3]
        wo = xml_dict[4]
        est_dict = xml_dict[5]
        storey_height_dict = xml_dict[6]
        
        # print('nwa_dict["11"]', ':')
        # pprint.pprint(nwa_dict["11"])
        
        # *****************************
        
        # All four of the below should probably be incorporated into XML_2_dict_new() above
        
        # 1. thermal_envelope_uids
        # 2. ext_perim - matching of walls to exploded walls
        # 3. walls - total_area = net_a (not sure this is required?)
        # 4. est_dict - assigning of values to wall_dict
        
        
        # *****************************
        
        
        # print('storey_height_dict', ':')
        # pprint.pprint(storey_height_dict)
        
        
        thermal_envelope_uids = []
        for e in storey_height_dict:
            if 'room_type' in storey_height_dict[e]['value'].keys():
                # print(storey_height_dict[e]['value'])
                if storey_height_dict[e]['value']['room_type'] not in ex_thermal_envelope:
                    thermal_envelope_uids.append(e)
        # print(thermal_envelope_uids)
        xml_ref_dict['thermal_envelope_uids'] = thermal_envelope_uids
        
        # *****************************
        
        
        # print('about to get ex wa')
        wt_dict_ewag, exploded_wall_dict = exterior_walls(root)
        
        r_to = 2
        
        
        
        point_list = []
        for floor in nwa_dict:
            # if floor != '11':
                # continue
            
            for room in nwa_dict[floor]:
                for wall in nwa_dict[floor][room]:
                    x1 = round(nwa_dict[floor][room][wall]['x1'], r_to)
                    y1 = round(nwa_dict[floor][room][wall]['y1'], r_to)
                    a = [x1, y1]
                    point_list.append(a)
                    x2 = round(nwa_dict[floor][room][wall]['x2'], r_to)
                    y2 = round(nwa_dict[floor][room][wall]['y2'], r_to)
                    b = [x2, y2]
                    point_list.append(b)
                
                
                # print('point_list', ':', point_list)
                # l = np.array(point_list)
                # point_list = []
                # datapoints = l.T
                # plt.plot(datapoints[0], datapoints[1])
            
            
            
            for room in nwa_dict[floor]:
                for wall in nwa_dict[floor][room]:
                    nwa_dict[floor][room][wall]['x3'] = round(nwa_dict[floor][room][wall]['x3'], r_to)
                    nwa_dict[floor][room][wall]['y3'] = round(nwa_dict[floor][room][wall]['y3'], r_to)
                    x3 = round(nwa_dict[floor][room][wall]['x3'], r_to)
                    y3 = round(nwa_dict[floor][room][wall]['y3'], r_to)
                    a = [x3, y3]
                    point_list.append(a)
                    nwa_dict[floor][room][wall]['x4'] = round(nwa_dict[floor][room][wall]['x4'], r_to)
                    nwa_dict[floor][room][wall]['y4'] = round(nwa_dict[floor][room][wall]['y4'], r_to)
                    x4 = round(nwa_dict[floor][room][wall]['x4'], r_to)
                    y4 = round(nwa_dict[floor][room][wall]['y4'], r_to)
                    b = [x4, y4]
                    point_list.append(b)
                
                
                # print('point_list', ':', point_list)
                # l = np.array(point_list)
                # point_list = []
                # datapoints = l.T
                # plt.plot(datapoints[0], datapoints[1])
        
        
        
        # obs_room = 'Hall (66bc8575.a5205bff)'
        obs_room = 'Kitchen (66c611bf.a8f4b3ff)'
        
        point_list = []
        for floor in exploded_wall_dict:
            # if floor != '11':
                # continue
            for ex_wall in exploded_wall_dict[floor]:
                type_ex = exploded_wall_dict[floor][ex_wall]['type']
                if exploded_wall_dict[floor][ex_wall]['type'] != 'exterior':
                    continue
                exploded_wall_dict[floor][ex_wall]['x1'] = round(exploded_wall_dict[floor][ex_wall]['x1'], r_to)
                exploded_wall_dict[floor][ex_wall]['y1'] = round(exploded_wall_dict[floor][ex_wall]['y1'], r_to)
                x1 = round(exploded_wall_dict[floor][ex_wall]['x1'], r_to)
                y1 = round(exploded_wall_dict[floor][ex_wall]['y1'], r_to)
                a = [x1, y1]
                point_list.append(a)
                exploded_wall_dict[floor][ex_wall]['x2'] = round(exploded_wall_dict[floor][ex_wall]['x2'], r_to)
                exploded_wall_dict[floor][ex_wall]['y2'] = round(exploded_wall_dict[floor][ex_wall]['y2'], r_to)
                x2 = round(exploded_wall_dict[floor][ex_wall]['x2'], r_to)
                y2 = round(exploded_wall_dict[floor][ex_wall]['y2'], r_to)
                b = [x2, y2]
                point_list.append(b)
                # print(a, '\t', b)
                
                for room in nwa_dict[floor]:
                    if 'ext_perim' not in nwa_dict[floor][room].keys():
                        nwa_dict[floor][room]['ext_perim'] = 0
                    # print('room', ':', room)
                    for wall in nwa_dict[floor][room]:
                        if not isinstance(nwa_dict[floor][room][wall], dict):
                            continue
                        # print('wall', ':', wall)
                        # pprint.pprint(nwa_dict[floor][room][wall])
                        if 'loadBearingWall' in nwa_dict[floor][room][wall].keys():
                            if nwa_dict[floor][room][wall]['loadBearingWall'] == '1':
                                continue
                        
                        x3 = nwa_dict[floor][room][wall]['x3']
                        y3 = nwa_dict[floor][room][wall]['y3']
                        c = [x3, -y3]
                        # print('c', ':', c)
                        x4 = nwa_dict[floor][room][wall]['x4']
                        y4 = nwa_dict[floor][room][wall]['y4']
                        d = [x4, -y4]
                        # print('d', ':', d)
                
                        # print(a, '\t', b, '\t', c, '\t', d)
                        # print('room', ':', room, '\t', 'type_ex', ':', type_ex,  '\t', 'wall', ':', wall,  '\t', x1, '\t', y1, '\t', x2, '\t', y2, '\t', x3, '\t', y3, '\t', x4, '\t', y4)
                        # print('room', ':', room, '\t', 'wall', ':', wall, '\t', x1, '\t', y1, '\t', x3, '\t', y3)
                        # print('room', ':', room, '\t', 'wall', ':', wall, '\t', x2, '\t', y2, '\t', x4, '\t', y4)
                        
                        
                        string = ('segment ' + str(ex_wall) + '\t' 
                                            + str(x1) + '\t' 
                                            + str(y1) + '\t' 
                                            + str(x2) + '\t' 
                                            + str(y2) + '\t' 
                                            + ' is colinear with wall ' + '\t' 
                                            + str(wall) + '\t' 
                                            + str(x3) + '\t' 
                                            + str(y3) + '\t' 
                                            + str(x4) + '\t' 
                                            + str(y4))
                        if linear_subset(x1, y1, x2, y2, x3, y3, x4, y4, epsilon=0.05, zeta=0.05) == True:
                            l = cart_distance((x1, y1), (x2, y2))
                            # if room == obs_room:
                                # print('cart_distance', ':', l)
                            nwa_dict[floor][room][wall]['ext_perim'] = l
                            nwa_dict[floor][room]['ext_perim'] += l
                        else:
                            string = string.replace('is colinear', 'is NOT colinear')
                            
                        # if room == obs_room:
                            # print(string)
                    
                    # if room == obs_room:
                        # print("nwa_dict[" + str(floor) + "][" + room + "]['ext_perim']", ':')
                        # pprint.pprint(nwa_dict[floor][room]['ext_perim'])
                        
                        # if (x1 == x3 and y1 == y3 and x2 == x4 and y2 == y4) or (x1 == x4 and y1 == y4 and x2 == x3 and y2 == y3):
                            # print('nwa_dict[floor][room]', ':')
                            # pprint.pprint(nwa_dict[floor][room])
                            # print(exploded_wall_dict[floor][ex_wall]['type'])
                            # nwa_dict[floor][room][wall]['type'] = exploded_wall_dict[floor][ex_wall]['type']
                
            # print('point_list', ':', point_list)
            # l = np.array(point_list)
            # datapoints = l.T
            # plt.scatter(datapoints[0], datapoints[1])
            # plt.plot(datapoints[0], datapoints[1])
        
            # plt.gca().set_aspect('equal', adjustable='box')
            # plt.show()
        
        # print('nwa_dict["10"][obs_room]', ':')
        # pprint.pprint(nwa_dict["10"][obs_room])
        # print('nwa_dict["10"]', ':')
        # pprint.pprint(nwa_dict["10"])
        # print('nwa_dict', ':')
        # pprint.pprint(nwa_dict)
        # print('exploded_wall_dict["10"]', ':')
        # pprint.pprint(exploded_wall_dict["10"])
        # print('exploded_wall_dict', ':')
        # pprint.pprint(exploded_wall_dict)
        
        
        
        
        
        # *****************************
        
        # Is this used for anything?:
        
        wall_dict = {}
        for floor in nwa_dict:
            for room in nwa_dict[floor]:
                for wall in nwa_dict[floor][room]:
                    if wall != 'ext_perim':
                        wall_dict[wall] = {}
                        wall_dict[wall]['value'] = {}
                        # print(list(nwa_dict[floor][room][wall].keys()))
                        wall_dict[wall]['value']['total_area'] = nwa_dict[floor][room][wall]['net_a']
            
        # print('est_dict', ':')
        # pprint.pprint(est_dict)        
        
        for item in est_dict:
            if item in wall_dict.keys():
                for v in est_dict[item]:
                    wall_dict[item]['value'][v] = est_dict[item][v]
        
        # print('wall_dict', ':')
        # pprint.pprint(wall_dict)
        
        # *****************************
        # *****************************
        # *****************************
        
        # At this point read in template_DEAP.xlsx
        
        # As was originally intended
        
        # Maybe we should read XL first, then XML?
        
        # *****************************
        
        # read template "template_ber.xlsx" from cloud (Azure for now, later SharePoint)
        local_xl_fp = Azure_2_Local(file_name = "template_ber.xlsx")
        # local_xl_fp = SharePoint_2_Local(url) # once we get access sorted
        
        # create dictionaries of required table/tab contents
        output_dict, lookup_dict = XL_2_dict_new(local_xl_fp) # is lookup_dict being returned here? we have a better way yes?
        # print('output_dict', ':')
        # pprint.pprint(output_dict)
        if not isinstance(output_dict, dict):
            print('not a dict:', output_dict)
            raise Exception(output_dict)
        
        
        
        
        # *****************************
        
        # list of uids of true_floors
        true_floors = []
        for tf in lookup_dict['Object Reference']['Floor Reference']['True']:
            if str(tf) in xml_ref_dict.keys():
                true_floors.append(xml_ref_dict[str(tf)])
        print('true_floors', ':', true_floors)
        xml_ref_dict['true_floors'] = true_floors
        
        # *****************************
        
        
        json_dict = JSON_2_dict(project_id
                                , forms_data=forms_data
                                , xml_ref_dict=xml_ref_dict
                                , xl_ref_dict=lookup_dict['Object Reference']
                                , wall_dict=wall_dict # from xml
                                , wo=wo
                                , colours_dict=colours_dict
                                ) # does this need to be project-specific?
        # print('json_dict', ':')
        # pprint.pprint(json_dict)
        if not isinstance(json_dict, dict):
            raise Exception(json_dict)
        
        
        
        # We now have all the info we need to create certain floor/roof/window/door dicts...
        # actually they are already included in json_dict
        # add them later to the output dict after we've created it from the INPUT TEMPLATE FILE
        # print('json_dict["count_dict"]', ':')
        # pprint.pprint(json_dict['count_dict'])
        
        
        # print("json_dict['count_dict']", ':')
        # pprint.pprint(json_dict['count_dict'])
        
        colours_dict_2 = {}
        for c in colours_dict:
            colours_dict_2[c] = {}
            colours_dict_2[c]['value'] = {}
            for floor_uid in colours_dict[c]['value']:
                floor_name = json_dict['floor_name_dict'][floor_uid]
                colours_dict_2[c]['value'][floor_name] = colours_dict[c]['value'][floor_uid]
        
        colours_dict = colours_dict_2
        
        # print('colours_dict', ':')
        # pprint.pprint(colours_dict)
        
        
        # *****************************
        
        # print('json_dict["floor_type_dict"]', ':')
        # pprint.pprint(json_dict['floor_type_dict'])
        # print('nwa_dict', ':')
        # pprint.pprint(nwa_dict)
        
        for room_uid in json_dict["floor_type_dict"]:
            ff = json_dict["floor_type_dict"][room_uid]['value']['floor_name']
            tf = get_true_floor(ff)
            json_dict["floor_type_dict"][room_uid]['value']['True Floor'] = tf
            
            for floor in nwa_dict:
                for room in nwa_dict[floor]:
                    if room_uid in room:
                        if 'ext_perim' in nwa_dict[floor][room].keys():
                            json_dict["floor_type_dict"][room_uid]['value']['perimeter'] = round(nwa_dict[floor][room]['ext_perim'], 2)
        
        
        
        
        
        
        # ************** WALLS: ADD XML DATA TO JSON ***************
        
        json_dict['wall_type_dict'] = wall_total_surface_new(json_dict['wall_type_dict'], est_dict)
        
        json_dict['wall_type_dict']['Wall Type 1']['value']['total_surface'] = json_dict['wall_type_dict']['Wall Type 1']['total_surface']
        
        
        
        # *****************************
        print('************** PROJECT FILES ***************')
        
        ofl_filelist = []
        ofl_filelist = get_project_files(project_id, plan_name = project_name) # ofl_filelist is part of this function's output
        
        if not isinstance(ofl_filelist, list):
            raise Exception(ofl_filelist)
        
        
        # print('finished getting project files')
        print('warning: did not get project files')
        
        
        
        
        # *****************************
        
        print('************** THERMAL MASS CALC ***************')
        tm_count_dict = {}
        
        if '7. Thermal Mass P1' in output_dict.keys():
            tm_count_dict['Number of Light Elements'] = 0
            tm_count_dict['Number of Medium Elements'] = 0
            tm_count_dict['Number of Heavy Elements'] = 0
            
            masses = []
            for mass_cat in ['Ground Floor Mass', 'External Wall Mass', 'Separating Wall Mass', 'Internal Partition Mass']:
                if mass_cat in json_dict.keys():
                    masses.append(json_dict[mass_cat])
                # else:
                    # print(mass_cat + ' not in json_dict.keys()')
                    # print(list(json_dict.keys()))
            
            # print('masses', ':', masses)
            for m in masses:
                # print(m)
                # print('masses[m]', ':', masses[m])
                for n in m:
                    # print(n)
                    val = m[n]
                field = f'Number of {val} Elements'
                tm_count_dict[field] += 1
            
            tm_count_dict['lmh'] = str(tm_count_dict['Number of Light Elements']) + str(tm_count_dict['Number of Medium Elements']) + str(tm_count_dict['Number of Heavy Elements'])
            
            if tm_count_dict['lmh'] in lookup_dict['7.1 Referance Table'].keys():
                tm_count_dict['Overall Thermal Mass Category'] = lookup_dict['7.1 Referance Table'][tm_count_dict['lmh']]['field_req']
            else:
                tm_count_dict['Overall Thermal Mass Category'] = 'not found (' + str(tm_count_dict['lmh']) + ')'
        # print('tm_count_dict', ':')
        # pprint.pprint(tm_count_dict)
        
        for e in tm_count_dict:
            json_dict[e] = str(tm_count_dict[e])
            # print(e, ':', json_dict[e])
        
        
        # *****************************
        # print('storey_height_dict', ':')
        # pprint.pprint(storey_height_dict)
        # print("json_dict['storey_height_dict']", ':')
        # pprint.pprint(json_dict['storey_height_dict'])
        
        for e in storey_height_dict:
            if e in json_dict['storey_height_dict'].keys():
                if storey_height_dict[e]['value']['floor_type'] in ['20', '21', '22', '23', '24']: # see table "Floor Reference" 
                    del json_dict['storey_height_dict'][e]
                else:
                    for f in storey_height_dict[e]['value']:
                        json_dict['storey_height_dict'][e]['value'][f] = storey_height_dict[e]['value'][f]
        
        
        
        
        
        
        
        # print("'output_dict'", ':')
        # pprint.pprint(output_dict)
        
        print('************** POPULATING OUTPUT DICT ***************')
        
        for door_group in json_dict['door_summary_dict']:
            output_dict['5.4 Door Summary Table'][door_group] = json_dict['door_summary_dict'][door_group]
        
        for window_group in json_dict['window_summary_dict']:
            output_dict['5.1 Windows Summary Table'][window_group] = json_dict['window_summary_dict'][window_group]
        
        for bulb_group in json_dict['bulb_summary_dict']:
            output_dict['11.1 Lighting Schedule'][bulb_group] = json_dict['bulb_summary_dict'][bulb_group]
        
        for group in json_dict['roof_summary_dict']:
            output_dict['3.5 Roof Type Summary Table'][group] = json_dict['roof_summary_dict'][group]
            
        for window in json_dict['window_dict']:
            output_dict['5.2 Window Schedule Table'][window] = json_dict['window_dict'][window]
        
        for door in json_dict['door_dict']:
            output_dict['5.5 Door Schedule Table'][door] = json_dict['door_dict'][door]
        
        # for bulb in json_dict['bulb_dict']:
            # output_dict['11.1 Lighting Schedule'][bulb] = json_dict['bulb_dict'][bulb]
        
        json_dict['attic_hatch_dict'] = {}
        for attic_hatch in json_dict['attic_hatch_dict']:
            output_dict['8.1 Attic Hatches'][attic_hatch] = json_dict['attic_hatch_dict'][attic_hatch]
        
        for vent in json_dict['vent_dict']:
            output_dict['8.2 Ventilation Items'][vent] = json_dict['vent_dict'][vent]
        
        # for floor in json_dict['floor_dict']:
            # output_dict['2.3 Floor Schedule Table'][floor] = json_dict['floor_dict'][floor]
        
        # print("json_dict['heating_dict']", ':')
        # pprint.pprint(json_dict['heating_dict'])
        
        for floor_type in json_dict['floor_type_dict']:
            if json_dict['floor_type_dict'][floor_type]['value']['is this floor being used?'] == True:
                output_dict['2.3 Floor Schedule Table'][floor_type] = json_dict['floor_type_dict'][floor_type]
        
        for roof in json_dict['roof_dict']:
            output_dict['3.4 Roof Type Schedule Table'][roof] = json_dict['roof_dict'][roof]
        
        for wall_type in json_dict['wall_type_dict']:
            if 'total_surface' in json_dict['wall_type_dict'][wall_type].keys():
                if json_dict['wall_type_dict'][wall_type]['total_surface'] != '' or 'Semi-Exposed' not in json_dict['wall_type_dict'][wall_type]:
                    output_dict['4.3 Wall Summary Table'][wall_type] = json_dict['wall_type_dict'][wall_type]
        
        for colour in colours_dict:
            output_dict['6. Colour Area Table P1'][colour] = colours_dict[colour]
        
        for h in json_dict['heating_dict']:
            if 'Heat Source Type on DEAP' in json_dict['heating_dict'][h]['value'].keys():
                if json_dict['heating_dict'][h]['value']['Heat Source Type on DEAP'] in ['Primary', 'Secondary']:
                    output_dict['9.3 Space Heating Category'][h] = json_dict['heating_dict'][h]
        
        
        
        

        
        json_dict['storey_height_dict']['floors'] = {}
        json_dict['storey_height_dict']['rooms'] = {}
        # la_dict = {}
        la_max = 0
        for e in json_dict['storey_height_dict']:
            # print("e", ':', e)
            # print("json_dict['storey_height_dict'][e]", ':')
            # pprint.pprint(json_dict['storey_height_dict'][e])
            
            if 'value' not in json_dict['storey_height_dict'][e].keys():
                continue
            
            if 'thermal_envelope' not in json_dict['storey_height_dict'][e]['value'].keys(): # hasn't been determined from XML
                if json_dict['storey_height_dict'][e]['value']['name'] in ex_thermal_envelope:
                    json_dict['storey_height_dict'][e]['value']['thermal_envelope'] = 0
                else:
                    json_dict['storey_height_dict'][e]['value']['thermal_envelope'] = 1 # this includes renamed rooms 
                    
                    
            if 'room_type' not in json_dict['storey_height_dict'][e]['value'].keys(): # i.e. floors only
                json_dict['storey_height_dict']['floors'][e] = json_dict['storey_height_dict'][e]
            else:
                json_dict['storey_height_dict']['rooms'][e] = json_dict['storey_height_dict'][e]
                if json_dict['storey_height_dict'][e]['value']['room_type'] == 'Living Room':
                    ft = json_dict['storey_height_dict'][e]['value']['floor_type']
                    if ft in ['-2', '-1', '0', '1', '2', '3', '4', '5', '6', '7', '8']:
                        la = json_dict['storey_height_dict'][e]['value']['area']
                        if la > la_max:
                            la_max = la
        
        json_dict['largest_living_area'] = str(la_max)
        
        # print("json_dict['largest_living_area']", ':', json_dict['largest_living_area'])
        
        
        # Need to sum room volumes by floor_type
        fv_dict = {}
        fa_dict = {}
        for room in json_dict['storey_height_dict']['rooms']:
            
            # print("room", ':', room)
            # print("json_dict['storey_height_dict']['rooms'][room]", ':')
            # pprint.pprint(json_dict['storey_height_dict']['rooms'][room])
            
            
            if json_dict['storey_height_dict']['rooms'][room]['value']['thermal_envelope'] == '0':
                continue
            if 'floor_type' in json_dict['storey_height_dict']['rooms'][room]['value']:
                f = json_dict['storey_height_dict']['rooms'][room]['value']['floor_type']
                if f not in fa_dict.keys():
                    fa_dict[f] = 0
                if f not in fv_dict.keys():
                    fv_dict[f] = 0
                a = json_dict['storey_height_dict']['rooms'][room]['value']['area']
                v = json_dict['storey_height_dict']['rooms'][room]['value']['volume']
                fa_dict[f] += a
                fv_dict[f] += v
        
        # print('fv_dict', ':')
        # pprint.pprint(fv_dict)
        # print('fa_dict', ':')
        # pprint.pprint(fa_dict)
        
        # *************** CALCULATE CEILING HEIGHT ***************
        
        calc_floors = [] # list of floors for which calc is necessary (output will include a table for each of these)
        ch_dict = {}
        for fv in fv_dict:
            for floor in json_dict['storey_height_dict']['floors']:
                if 'floor_type' in json_dict['storey_height_dict']['floors'][floor]['value'].keys():
                    if json_dict['storey_height_dict']['floors'][floor]['value']['floor_type'] == fv:
                        json_dict['storey_height_dict']['floors'][floor]['value']['volume'] = round(fv_dict[fv], 2)
                        if json_dict['storey_height_dict']['floors'][floor]['value']['use_floor_level_height'] == '0':
                            calc_floors.append(fv)
                            v = json_dict['storey_height_dict']['floors'][floor]['value']['volume']
                            # a = json_dict['storey_height_dict']['floors'][floor]['value']['area']
                            a = fa_dict[fv]
                            ch = round(v / a, 2)
                            json_dict['storey_height_dict']['floors'][floor]['value']['ceiling_height'] = ch
                            # print('fv', ':', fv)
                            # print('v', ':', v)
                            # print('a', ':', a)
                            # print('ch', ':', ch)
                            ch_dict[fv] = ch
                            # also need to add this value to rooms dict
                            
                            
                        else:
                            json_dict['storey_height_dict']['floors'][floor]['value']['ceiling_height'] = json_dict['storey_height_dict']['floors'][floor]['value']['height']
        # print('ch_dict', ':', ch_dict)
        for room in json_dict['storey_height_dict']['rooms']:
            ft = json_dict['storey_height_dict']['rooms'][room]['value']['floor_type']
            # print('ft', ':', ft)
            if ft in ch_dict.keys():
                json_dict['storey_height_dict']['rooms'][room]['value']['Weighted Average Ceiling Height'] = ch_dict[ft]
                # print("json_dict['storey_height_dict'][room]", ':')
                # pprint.pprint(json_dict['storey_height_dict'][room])
        
        output_dict['2 Building Average Storey (Floors)'] = dict(output_dict['2 Building Average Storey'])
        
        
        # print(len(json_dict['storey_height_dict']['floors']))
        fn = 0
        for e in json_dict['storey_height_dict']['floors']:
            ft = json_dict['storey_height_dict']['floors'][e]['value']['floor_type']
            if ft in ['-2', '-1', '0', '1', '2', '3', '4', '5', '6', '7', '8']:
                fn += 1
                output_dict['2 Building Average Storey (Floors)'][e] = json_dict['storey_height_dict']['floors'][e]
        json_dict['no_of_stories'] = str(fn)
        
        
        
        
        floors = []
        for e in json_dict['storey_height_dict']['rooms']:
            # print(json_dict['storey_height_dict']['rooms'][e]['value']['floor_type'])
            ft = json_dict['storey_height_dict']['rooms'][e]['value']['floor_type']
            if ft not in floors:
                floors.append(ft)
            # print(ft)
        print('floors', ':', floors)
        print('calc_floors', ':', calc_floors)
        
        for ft in calc_floors:
            output_dict['2 Building Average Storey (Rooms - Floor ' + ft + ')'] = {}
        
        
        
        # print("json_dict['storey_height_dict']['floors']", ':')
        # pprint.pprint(json_dict['storey_height_dict']['floors'])
        
        for e in json_dict['storey_height_dict']['rooms']:
            ft = json_dict['storey_height_dict']['rooms'][e]['value']['floor_type']
            # print('ft', ':', ft)
            if ft in calc_floors:
                # if ft in ['-2', '-1', '0', '1', '2', '3', '4', '5', '6', '7', '8']:
                # print("json_dict['storey_height_dict']['rooms'][e]", ':')
                # pprint.pprint(json_dict['storey_height_dict']['rooms'][e])
                if 'ceiling_height' not in json_dict['storey_height_dict']['rooms'][e].keys():
                    json_dict['storey_height_dict']['rooms'][e]['value']['ceiling_height'] = json_dict['storey_height_dict']['rooms'][e]['value']['height']
                output_dict['2 Building Average Storey (Rooms - Floor ' + ft + ')'][e] = json_dict['storey_height_dict']['rooms'][e]
                
                
        
        
        
        
        
        
        
        
        
        
        
        if 'Chimney' in json_dict['count_dict'].keys():
            json_dict['count_dict']['Number of Chimneys'] = json_dict['count_dict']['Chimney']
        if 'Flue' in json_dict['count_dict'].keys():
            json_dict['count_dict']['Number of Flues'] = json_dict['count_dict']['Flue']
        if 'Attic Hatch Draughtproofed' in json_dict['count_dict'].keys():
            json_dict['count_dict']['Attic Hatches Draught stripped'] = json_dict['count_dict']['Attic Hatch Draughtproofed']
        if 'Attic Hatch Not Draughtproofed' in json_dict['count_dict'].keys():
            json_dict['count_dict']['Attic Hatches Not Draught stripped'] = json_dict['count_dict']['Attic Hatch Not Draughtproofed']
        
        
        print("json_dict['count_dict']", ':')
        pprint.pprint(json_dict['count_dict'])
        
        
        # json_dict['Number of LED/CFL bulbs'] = json_dict['count_dict']['LED/CFL']
        # json_dict['Number of Halogen Lamp bulbs'] = json_dict['count_dict']['Halogen Lamp']
        # json_dict['Number of Halogen Lamp Low Voltage bulbs'] = json_dict['count_dict']['Halogen LV']
        # json_dict['Number of Incandescent/Unknown bulbs'] = json_dict['count_dict']['Incandescent']
        # json_dict['Number of Linear Fluorescent bulbs'] = json_dict['count_dict']['Linear Fluorescent']
        
        json_dict['LED/CFL'] = json_dict['count_dict']['LED/CFL']
        json_dict['Halogen Lamp'] = json_dict['count_dict']['Halogen Lamp']
        json_dict['Halogen LV'] = json_dict['count_dict']['Halogen LV']
        json_dict['Incandescent'] = json_dict['count_dict']['Incandescent']
        json_dict['Linear Fluorescent'] = json_dict['count_dict']['Linear Fluorescent']
        
        print("json_dict['bulb_dict']", ';')
        pprint.pprint(json_dict['bulb_dict'])
        
        
        
        
        
        # *****************************
        
        # populate the dicts 
        print('output_dict', ':')
        for sheet_name in output_dict:
            print(sheet_name)
            # pprint.pprint(output_dict[sheet_name])
            # for record in output_dict[sheet_name]:
            for field in output_dict[sheet_name]:
                # if sheet_name == "7. Thermal Mass P1":
                if sheet_name == "11. Lighting P1":
                    print('field_req', ':', field_req)
                if not isinstance(output_dict[sheet_name][field], dict):
                    continue
                if 'field_req' not in output_dict[sheet_name][field].keys():
                    # if sheet_name == "7. Thermal Mass P1":
                        # print('field_req', 'not found in output_dict[sheet_name][field].keys()')
                        # print(list(output_dict[sheet_name][field].keys()))
                    continue
                field_req = output_dict[sheet_name][field]['field_req']
                
                # first check if it's Exact Text (Forms question):
                if field_req in json_dict.keys():
                    # print('Exact Text')
                    output_dict[sheet_name][field]['value'] = json_dict[field_req]
                
                # then check if it's a variable name from xml_val_dict
                elif field_req in xml_val_dict.keys():
                    output_dict[sheet_name][field]['value'] = xml_val_dict[field_req]
                
                # lookup_table
                elif isinstance(field_req, str) and field_req[0:6] == "lookup":
                    lu = field_req.split("|")
                    # print('eval(lu[2])', ':', eval(lu[2]))
                    
                    output_dict[sheet_name][field]['value'] = str(lookup_dict[lu[1]][eval(lu[2])]['field_req'])
                
                
                elif field_req in json_dict['count_dict'].keys():
                    print('Count Dict')
                    output_dict[sheet_name][field]['value'] = str(json_dict['count_dict'][field_req])
                
                # logic is uncharted
                else:
                    output_dict[sheet_name][field]['value'] = output_dict[sheet_name][field]['default_val'] if 'default_val' in output_dict[sheet_name][field].keys() else "NOT FOUND"


        # *****************************
        
        # use output_dict to populate (a copy of) the Excel template and save it to be sent as an email attachment
        # populate_template(project_name, template_name, data_dict = {})
        
        # *****************************
        
        # output_dict_DEAP = output_dict.copy()
        output_dict_DEAP = {}
        
        for field in output_dict:
            output_dict_DEAP[field] = output_dict[field]
        
        
        d = '1. Survey Details P1'
        
        efs = ['Age Band: Extension 1', 'Year of Construction: Extension 1', 'Age Band: Extension 2', 'Year of Construction: Extension 2', 'Age Band: Extension 3', 'Year of Construction: Extension 3', 'Age Band: Extension 4', 'Year of Construction: Extension 4', 'Age Band: Extension 5', 'Year of Construction: Extension 5']
        
        for ef in efs:
            if output_dict[d][ef]['value'] == '':
                del output_dict[d][ef]
            else:
                if output_dict[d][ef]['value']['value'] == '':
                    del output_dict[d][ef]
                # print(ef, ':', output_dict[d][ef])
        
        d = '7. Thermal Mass P1'
        
        # print(d)
        # pprint.pprint(output_dict[d])
        
        efs = ['Comments on Ground Floors', 'Comments on External Walls', 'Comments on Separating Walls', 'Comments on Internal Partitions', 'Acceptable Construction Details 2008, 2011 and 2019 TGDL Factor for Thermal Bridging [Wm²K]', 'Building Regulation 2005 TGDL Factor for Thermal Bridging [Wm²K]', 'User Defined Factor for Thermal Bridging [Wm²K]']
        
        for ef in efs:
            if output_dict[d][ef]['value'] == '':
                del output_dict[d][ef]
            else:
                if 'value' in output_dict[d][ef]['value'].keys():
                    if output_dict[d][ef]['value']['value'] == '':
                        del output_dict[d][ef]
            
        d = '8. Ventilation P1'
        
        print(d)
        pprint.pprint(output_dict[d])
        
        
        
        
        
        # print(list(output_dict[d].keys()))
        
        efs = list(output_dict[d].keys())
        
        # efs = ['Positive input ventilation from outside Specific Fan Power [W/[l/s]]'
                # , 'Positive input ventilation from outside Default Specific Fan Power [W/[l/s]]'
                # , 'Whole-house extract ventilation Specific Fan Power [W/[l/s]]'
                # , 'Whole-house extract ventilation Default Specific Fan Power [W/[l/s]]'
                # , 'Balanced whole-house mechanical ventilation, no heat recovery Specific Fan Power [W/[l/s]]'
                # , 'Balanced whole-house mechanical ventilation, no heat recovery Default Specific Fan Power [W/[l/s]]'
                # , 'Balanced whole-house mechanical ventilation with heat recovery Specific Fan Power [W/[l/s]]'
                # , 'Balanced whole-house mechanical ventilation with heat recovery Default Specific Fan Power [W/[l/s]]'
                # , 'Balanced whole-house mechanical ventilation with heat recovery Heat Exchanger Efficiency [%]'
                # , 'Balanced whole-house mechanical ventilation with heat recovery Default Heat Exchanger Efficiency [%]'
                # , 'Is there uninsulated ducting on MVHR system outside dwelling envelope? (Non-Default)'
                # , 'Is there uninsulated ducting on MVHR system outside dwelling envelope?'
                # , 'Exhaust Air Heat Pump Specific Fan Power [W/[l/s]]'
                # , 'Exhaust Air Heat Pump Default Specific Fan Power [W/[l/s]]'
                # , 'Exhaust Air Flow Rate [m3/h]'
                # , 'Default Exhaust Air Flow Rate [m3/h]'
                # , 'Manufacturer/Model'
                # , 'Manufactur Model'
                # , 'Is the ventilation ducting flexible/rigid/both?'
                # ]
        
        for ef in efs:
            if ef in output_dict[d].keys():
                # print("output_dict[d][ef]['value']")
                # pprint.pprint(output_dict[d][ef]['value'])
                if output_dict[d][ef]['value'] == '':
                    del output_dict[d][ef]
                else:
                    if isinstance(output_dict[d][ef]['value'], dict):
                        if 'value' in output_dict[d][ef]['value'].keys():
                            if output_dict[d][ef]['value']['value'] == '':
                                del output_dict[d][ef]
                
        
        
        
        d = '11. Lighting P1'
        
        # print(d)
        # pprint.pprint(output_dict[d])
        
        efs = ['Number of LED/CFLs'
                , 'LED/CFL Efficiency [lm/W]'
                , 'LED/CFL Power [W]'
                , 'Number of Linear Florescents'
                , 'Linear Florescent Efficiency [lm/W]'
                , 'Linear Florescent Power [W]'
                , 'Number of Incandescents'
                , 'Incandescent Efficiency [lm/W]'
                , 'Incandescent Power [W]'
                , "Number of Halogen LV's"
                , 'Halogen LV Efficiency [lm/W]'
                , 'Halogen LV Power [W]'
                , 'Number of Halogen Lamps'
                , 'Halogen Lamp Efficiency [lm/W]'
                , 'Halogen Lamp Power [W]'
                ]
        
        for ef in efs:
            if output_dict[d][ef]['value'] == '':
                del output_dict[d][ef]
            else:
                if 'value' in output_dict[d][ef]['value'].keys():
                    if output_dict[d][ef]['value']['value'] == '':
                        del output_dict[d][ef]
            
        d = '2 Building Average Storey (Floors)'
        output_dict[d]['headers'] = ['uid', 'name', 'use_floor_level_height', 'ceiling_height']
        
        for d in list(output_dict.keys()):
            if '2 Building Average Storey (Rooms' in d:
                output_dict[d]['headers'] = ['uid', 'area', 'ceiling_height', 'volume', 'room_type', 'thermal_envelope', 'Weighted Average Ceiling Height']
        
        
        # *****************************
        
        
        
        d = '11. Lighting P1'
        
        
        for f in output_dict[d]:
            output_dict[d][f]['value'] = str(output_dict[d][f]['value'])
        
        # print(d)
        # pprint.pprint(output_dict[d])
        
        
        # *****************************
        
        # use output_dict to generate this function's "output" HTML to serve as the body of the return email
        if output == '': # otherwise might contain error details from a function? Need to revisit this
            styling = "border=\"1\""
            
            output_table_list = [
                                '1. Survey Details P1'
                                , '2 Building Average Storey (Floors)'
                                ]
            
            for section in output_dict:
                # print('output_dict section', ':', section)
                if '2 Building Average Storey (Rooms - Floor' in section:
                    output_table_list.append(section)
                    
            
            output_table_list_2 = [
                                '2.3 Floor Schedule Table'
                                , '3.4 Roof Type Schedule Table'
                                , '3.5 Roof Type Summary Table'
                                , '4.3 Wall Summary Table'
                                , '5.5 Door Schedule Table'
                                # , '5.4 Door Summary Table'
                                , '5.2 Window Schedule Table'
                                , '5.1 Windows Summary Table'
                                , '6. Colour Area Table P1'
                                , '7. Thermal Mass P1'
                                , '8. Ventilation P1'
                                , '8.1 Attic Hatches'
                                , '8.2 Ventilation Items'
                                , '11. Lighting P1'
                                , '11.1 Lighting Schedule'
                                # 9. Space Heating P4
                                # 9.3 Space Heating Category
                                # 9.4 Heating System Controls
                                # 9.4 Heating System Controls (2)
                                # 9.4 Pumps and Fans
                                ]
                                
            
            output_table_list = output_table_list + output_table_list_2
            
            
            for section in output_table_list:
                print('output_table_list section', ':', section)
                # pprint.pprint(output_dict[section])
                
                if 'headers' in output_dict[section].keys():
                    headers = output_dict[section]['headers']
                    del output_dict[section]['headers']
                    order_list = [field for field in output_dict[section]]
                    # order_list = [field for field in headers]
                else:
                    headers = ['name', 'value']
                    order_list = [field for field in output_dict[section]]
                # print('headers', ':', headers)
                # print('order_list', ':', order_list)
                
                
                
                
                
                # if section in ['2 Building Average Storey (Floors)', '2.3 Floor Schedule Table', '3.5 Roof Type Summary Table', '5.2 Window Schedule Table', '5.5 Door Schedule Table', '8.1 Ventilation Items', '11.1 Lighting Schedule']:
                    # headers_only = True
                # else:
                    # headers_only = False
                headers_only = True
                
                if '2 Building Average Storey (Rooms' in section:
                    headers_only = True
                
                colour_table = False
                if section == '6. Colour Area Table P1':
                    colour_table = True
                    order_list = []
                
                if section == '8. Ventilation P1':
                    # order_list = []
                    ovm = json_dict["Is there another ventilation method other than Natural Ventilation?"]['value']
                    print("Is there another ventilation method other than Natural Ventilation?")
                    print(ovm)
                    if ovm == False:
                        json_dict["Ventilation Type"]['value'] = "Natural Ventilation"
                    else:
                        json_dict["Ventilation Type"]['value'] = "Other"
                
                print('json_dict["Ventilation Type"]', ':', json_dict["Ventilation Type"])
                
                section_output = f"""\
                                <h1>{section}</h1> \
                                {create_table_new(output_dict[section], headers, styling=styling, do_not_sum=['All'], order_list=order_list, title=section, headers_only=headers_only, colour_table=colour_table)} \
                                </div>"""
                output = output + section_output
                
            file_list = f"""\
                <h1>File List</h1> \
                {create_table_text(output_dict, headers = ['name'], styling=styling, do_not_sum=['All'], order_list = ofl_filelist)} \
                </div>"""
                
            output = output + file_list
                
            output = output + "</div>"
        
    except:
        output = traceback.format_exc()
        print(output)
    
    return output




def wall_total_surface_new(wall_type_dict, est_dict):
    try:
        for wall_type in wall_type_dict:
            if wall_type in est_dict.keys():
                wall_type_dict[wall_type]['total_surface'] = est_dict[wall_type]['total_surface']
                wall_type_dict[wall_type]['value']['total_surface'] = est_dict[wall_type]['total_surface']
            
        output = wall_type_dict
    finally:
    
        return output



def create_table_new(data_dict
                    , headers  = ['name', 'value'] # in this case headers should come from dict
                    , do_not_sum : list[str] = []
                    , styling: str = ""
                    , order_list = []
                    , title = ''
                    , headers_only = False
                    , colour_table = False
                    ) -> str:
    try:
        
        # dict is output_dict[section]
        # top-level entry is currently Field Name
        # To allow multi-column need to insert new top level if each new column is a record?
        # What if each new column is an attribute? Need value_list?
        # headers depends on whether records are columns or rows
            # columns: headers = order_list
            # rows: headers = value_list
        # identify from dict top-level?
        
        
        d = '11. Lighting P1'
        print('creating table for title', ':', title)
        # if title == d:
            # print('data_dict', ':')
            # pprint.pprint(data_dict)
        # print('headers_only', ':', headers_only)
        # print('headers', ':', headers)
        # print('order_list', ':', order_list)
        
        # print('this is the data_dict we need to use to create a multicol table:')
        # pprint.pprint(data_dict)
        
        output = f'<table {styling}><tr>'
        

        
        
        
        if len(order_list) != 0:
            for item in order_list:
                # if item not in data_dict.keys():
                    # continue
                if not isinstance(data_dict[item], dict):
                    continue
                # print('item', ':', item)
                # for r, record in enumerate(data_dict):
                    # print(data_dict[record][item].keys())
                if 'value' in list(data_dict[item].keys()):
                    # print('type', ':', type(data_dict[record][item]['value']))
                    e = data_dict[item]['value']
                    if headers_only == False: # headers will be appended for any additional values present
                        if isinstance(e, dict):
                            for key in data_dict[item]['value'].keys():
                                if key not in headers:
                                    headers.append(key)
            
                        
        replace_header = ''
        # print('headers', ':', headers)
        if len(headers) == 3:
            print('headers', ':', headers)
            replace_header = headers[2]
            print('replace_header', ':', replace_header)
            headers = headers[:-1]
            # print('headers', ':', headers)
        
        for header in headers:
            output += f'<th>{header}</th>'
        output += '</tr>'
        
        
        
        if len(order_list) != 0:
            for item in order_list:
                # if item not in data_dict.keys():
                    # continue
                if not isinstance(data_dict[item], dict):
                    continue
                # add the row (populate first column): 
                if item.isupper():
                    output += f'<tr><td><strong>{item}</strong></td>'
                else:
                    output += f'<tr><td>{item}</td>'
                # add the cols:
                # for r, record in enumerate(data_dict):
                    # for elem in data_dict[record][item]:
                if 'value' not in data_dict[item].keys():
                    continue
                
                # if title == d:
                    # print(data_dict[item]['value'])
                
                if isinstance(data_dict[item]['value'], str):
                    # if title == d:
                        # print('String')
                    if data_dict[item]['value'] != '':
                        v = data_dict[item]['value']
                        data_dict[item]['value'] = {}
                        data_dict[item]['value']['value'] = v # convert the string to a single-entry dict
                
                if isinstance(data_dict[item]['value'], dict):
                    # if title == d:
                        # print('Dict')
                    if replace_header != '':
                        if replace_header in data_dict[item]['value'].keys():
                            data_dict[item]['value']['value'] = data_dict[item]['value'][replace_header]
                    for header in headers[1:]:
                        if header in data_dict[item]['value'].keys():
                            value = data_dict[item]['value'][header]
                        else:
                            value = ' '
                        # if title == d:
                            # print('value', ';', value)
                        output += f'<td>{value}</td>'
                        

                        # output(val, ':', data_dict[item][elem][m][val])
                        # value = data_dict[item]["value"] if (item in data_dict.keys() and "value" in dict[item].keys()) else '' # already have blank values covered? If not then do we need a warning at this point?
                        # value = data_dict[item][elem][m][val] if (m in data_dict[item][elem].keys() and val in data_dict[item].keys()) else '' # already have blank values covered? If not then do we need a warning at this point?
                
                    # substitute boolean values for strings (should this be done elsewhere?):
                    # if (type(value) == bool and value == True):
                        # value = "Yes"
                    # if (type(value) == bool and value == False):
                        # value = "No"
                    # output += f'<td>{value}</td>'
                    # print(output)
                # print(item, value)
        else:
            for i, key in enumerate(data_dict):
                # print(key, data_dict[key])
                if key.isupper():
                    output += f'<tr><td><strong>{key}</strong></td>'
                elif colour_table:
                    print(key[:len(key)-2])
                    print(data_dict[key])
                    output += f'<tr><td><font color={key[:len(key)-2]}><b>Colour {i}</b></font></td>'
                    data_dict[key] = data_dict[key]['value']
                else:
                    output += f'<tr><td>{key}</td>'
                if isinstance(data_dict[key], dict):
                    if 'value' in data_dict[key].keys():
                        data_dict[key] = data_dict[key]['value']
                
                output += f'<td>{data_dict[key]}</td>'

        output += '</table>'
        
        output = output.replace('\u03bb', '&#955')
        output = output.replace('\u00b2', '&#178')
    except:
        output = traceback.format_exc()
        print(output)
        raise

    finally:
        return output
    return output




def populate_template(project_name, template_name, data_dict = {}):
    try:
        if template_name in ['template_ber', 'template']:
            filename = json_val_dict['plan_name'] + '.xlsx'
        if template_name == 'template_mrc':
            filename = json_val_dict['plan_name'] + ' Major Renovation calculation.xlsx'
        return_filename = filename

        account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
        default_credential = DefaultAzureCredential()
        output = ''
        container_name = 'attachment'
        local_path = "/tmp"
        instance_file_path = os.path.join(local_path, filename)
        
            
        # v = {
            # 'plan_name': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'D2'}
            # , 'Thermal Envelope - Heat loss walls, windows and doors': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E6'}
            # , 'Thermal Envelope - Heat loss floor area': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E7'}
            # , 'Thermal Envelope - Heat loss roof area': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E8'}
            # , 'THERMAL ENVELOPE OF BUILDING AREA': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E9'}
            # , 'Heat loss Wall Area recommended for EWI and IWI': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E12'}
            # , 'New Windows being recommended for replacement': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E13'}
            # , 'TOTAL SURFACE AREA FOR MAJOR RENOVATION WORKS': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E14'}
            # , 'Total surface area for MR works / Thermal Envelope': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E17'}
            # , 'WARMER HOMES MAJOR RENOVATION RESULT': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E18'}
            # }
        
        for field in data_dict:
            print('field', ':', field)

            if type(field['Value']) == str and '<BR>' in v[field]['Value']:
                v[field]['Value'] = v[field]['Value'].replace('<BR>', '\n')


        # Create the BlobServiceClient object
        blob_service_client = BlobServiceClient(account_url, credential=default_credential)
        container_client = blob_service_client.get_container_client(container= container_name) 


        xfile = openpyxl.load_workbook(instance_file_path)
        
        for field in v:
            # print(field)
            sheet = xfile[v[field]['Tab']]
            sheet[v[field]['Cell']] = v[field]['Value']



        xfile.save(instance_file_path)


        instance_file_path = os.path.join(local_path, filename)
        with open(file=instance_file_path, mode="rb") as upload_file:
            blob_client = blob_service_client.get_blob_client(container=container_name, blob=instance_file_path)
            blob_client.upload_blob(upload_file, overwrite=True)
    

        
        print('created')
        
        
        # if template in ['template', 'template_mrc']:
        output = copy_from_container(json_val_dict['plan_name'], filename)
        
        
    except:
        output = traceback.format_exc()
        print(output)

    finally:
        return output, return_filename





app = func.FunctionApp()
@app.function_name(name="MagicplanTrigger")
@app.route(route="magicplan", auth_level=func.AuthLevel.ANONYMOUS)


def test_function(req: func.HttpRequest) -> func.HttpResponse:
    try:

        form = dict(req.form)
        json_data = distributor_function(form)




        sc = 200    # OK

    except Exception as ex:
        output = str(ex)
        output = traceback.format_exc()
        sc = 503    # Service Unavailable
        

    finally:        
        try:
            # azure_upload(json_data)
            account_url = os.environ['AZ_STR_URL']
            # account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
            default_credential = DefaultAzureCredential()
            blob_service_client = BlobServiceClient(account_url, credential=default_credential)
            
            container_name = os.environ['AZ_CNTR_ST']
            # container_name = "magicplan-container"
            container_client = blob_service_client.get_container_client(container_name)
            if not container_client.exists():
                container_client = blob_service_client.create_container(container_name)
            
            local_file_name = str(uuid.uuid4()) + '.json'
            blob_client = blob_service_client.get_blob_client(container=container_name, blob=local_file_name)
            blob_client.upload_blob(json_data)
            
            # container_name = "project-files"
            # container_client = blob_service_client.get_container_client(container_name)
            # if not container_client.exists():
                # container_client = blob_service_client.create_container(container_name)
            
            
            return_body = '0'
            
        except Exception as ex:
            output = str(ex)
            output = traceback.format_exc()
            sc = 500     # Internal Server Error
            # return_body = output
            json_data = json.dumps({
                'email' : email,
                'name'  : project_name, 
                'table' : output
            })
            azure_upload(json_data)
        return func.HttpResponse(status_code=sc, body=return_body)



def azure_upload(file_data, container_name = 'attachment'):
    # account_url = os.environ['AZ_STR_URL']
    account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
    
    default_credential = DefaultAzureCredential()
    blob_service_client = BlobServiceClient(account_url, credential=default_credential)
    
    # container_name = os.environ['AZ_CNTR_ST']
    container_client = blob_service_client.get_container_client(container_name)
    if not container_client.exists():
        container_client = blob_service_client.create_container(container_name)
    
    local_file_name = str(uuid.uuid4()) + '.json'
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=local_file_name)
    blob_client.upload_blob(file_data)
    
    return




def populate_template_new(json_val_dict, template):
    try:

        account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
        default_credential = DefaultAzureCredential()
        output = ''
        return_filename = ''
        
        if template == 'template_deap':
            v = {
                'plan_name': { 'Value': '' , 'Tab': 'Sheet1' , 'Cell': 'D2'}
                }
            
        
        if template == 'template_ber':
            filename = json_val_dict['plan_name'] + '.xlsx'
            container_name = 'attachment'
            local_path = "/tmp"
            print('local_path', ':', local_path)
            instance_file_path = os.path.join(local_path, filename)
            print('instance_file_path', ':', instance_file_path)
            v = {
            'plan_name': { 'Value': '' , 'Tab': 'Sheet1' , 'Cell': 'D2'}
                }
        
        if template == 'template':
            filename = json_val_dict['plan_name'] + '.xlsx'
            container_name = 'attachment'
            local_path = "/tmp"
            print('local_path', ':', local_path)
            instance_file_path = os.path.join(local_path, filename)
            print('instance_file_path', ':', instance_file_path)
            v = {
                'Applicant Name': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C4'}
                , 'Application ID': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E4'}
                , 'Client Address': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C6'}
                , 'MPRN': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E6'}
                , 'Surveyor': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C8'}
                , 'Survey Date': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E8'}
                , 'Dwelling Type*': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C10'}
                , 'Gross floor area (m2) *': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E10'}
                , 'Dwelling Age*': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C12'}
                , 'Number of Storeys *': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E12'}
                , 'Age Extension 1': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C14'}
                , 'Room in Roof': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E14'}
                , 'Age Extension 2': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C16'}
                , 'No. Single Glazed Windows *': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E16'}
                , 'Asbestos Suspected': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C18'}
                , 'No. Double Glazed Windows *': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E18'}
                , 'Asbestos Details': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C20'}
                , 'Lot *': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C22'}
                , 'Property Height (m)*': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E22'}
                , 'Eircode': { 'Value': '' , 'Tab': 'General' , 'Cell': 'C24'}
                , 'Internet Available': { 'Value': '' , 'Tab': 'General' , 'Cell': 'E24'}


                , 'Roof 1 Type*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E25'}
                , 'Sloped Ceiling Roof 1*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H25'}
                , 'Other Details Roof 1*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E27'}
                , 'Roof 1 greater than 2/3 floor area*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H27'}
                , 'Roof 1 Pitch (degrees)*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H29'}
                , 'Roof Type 1 Insulation Exists*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E33'}
                , 'Can Roof Type 1 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E35'}
                , 'Roof 1 Thickness (mm)*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E37'}
                , 'Roof 1 Insulation Type*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H33'}
                #N/A
                , 'Required per standards (mm2) *': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E42'}
                , 'Existing (mm2)*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H42'}
                , 'Area of Roof Type 1 with fixed flooring (m2)*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E46'}
                , 'Folding/stair ladder in Roof Type 1*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E48'}
                , 'Fixed light in Roof Type 1*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E50'}
                , 'Downlighters in Roof Type 1*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H46'}
                , 'High power cable in Roof Type 1 (6sq/10sq or higher)*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'H48'}
                , 'Roof 2 Type': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E25'}
                , 'Other Details Roof 2*': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E27'}
                , 'Sloped Ceiling Roof 2*': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H25'}
                , 'Roof 2 greater than 2/3 floor area*': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H27'}
                , 'Roof 2 Pitch (degrees)*': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H29'}
                , 'Roof 2 Insulation Exists*': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E33'}
                , 'Can Roof Type 2 Insulation Thickness be Measured?': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E35'}
                , 'Roof 2 Thickness (mm)': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E37'}
                , 'Roof 2 Insulation Type': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H33'}
                #N/A
                , 'Roof 2 Required per standards (mm2) *': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E42'}
                , 'Roof 2 Existing (mm2) *': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H42'}
                , 'Area of Roof Type 2 with fixed flooring (m2)': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E46'}
                , 'Folding/stair ladder in Roof Type 2': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E48'}
                , 'Fixed light in Roof Type 2': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'E50'}
                , 'Downlighters in Roof Type 2': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H46'}
                , 'High power cable in Roof Type 2 (6sq/10sq or higher)': { 'Value': '' , 'Tab': 'Roof Type 2' , 'Cell': 'H48'}
                , 'Roof 3 Type': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E25'}
                , 'Other Details Roof 3*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E27'}
                , 'Sloped Ceiling Roof 3*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H25'}
                , 'Roof 3 greater than 2/3 floor area*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H27'}
                , 'Roof 3 Pitch (degrees)*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H29'}
                , 'Roof 3 Insulation Exists*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E33'}
                , 'Can Roof Type 3 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E35'}
                , 'Roof 3 Thickness (mm)*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E37'}
                , 'Roof 3 Insulation Type*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H33'}
                #N/A
                , 'Roof 3 Required per standards (mm2) *': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E42'}
                , 'Roof 3 Existing (mm2) *': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H42'}
                , 'Area of Roof Type 3 with fixed flooring (m2)*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E46'}
                , 'Folding/stair ladder in Roof Type 3*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E48'}
                , 'Fixed light in Roof Type 3*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'E50'}
                , 'Downlighters in Roof Type 3*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H46'}
                , 'High power cable in Roof Type 3 (6sq/10sq or higher)*': { 'Value': '' , 'Tab': 'Roof Type 3' , 'Cell': 'H48'}
                , 'Roof 4 Type': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E25'}
                , 'Other Details Roof 4*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E27'}
                , 'Sloped Ceiling Roof 4*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H25'}
                , 'Roof 4 greater than 2/3 floor area': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H27'}
                , 'Roof 4 Pitch (degrees)*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H29'}
                , 'Roof Type 4 Insulation Exists*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E33'}
                , 'Can Roof Type 4 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E35'}
                , 'Roof 4 Thickness (mm)': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E37'}
                , 'Roof 4 Insulation Type*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H33'}
                #N/A
                , 'Roof 4 Required per standards (mm2) *': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E42'}
                , 'Roof 4 Existing (mm2) *': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H42'}
                , 'Area of Roof Type 4 with fixed flooring (m2)*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E46'}
                , 'Folding/stair ladder in Roof Type 4*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E48'}
                , 'Fixed light in Roof Type 4*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'E50'}
                , 'Downlighters in Roof Type 4*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H46'}
                , 'High power cable in Roof Type 4 (6sq/10sq or higher)*': { 'Value': '' , 'Tab': 'Roof Type 4' , 'Cell': 'H48'}
                , 'Suitable for Insulation *': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E54'}
                , 'Roof not suitable details*': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E56'}
                , 'Notes (Roof)': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'E59'}
                , 'Internal Wall Insulation: Sloped or flat (horizontal) surface': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I67', 'Default': ''}
                , 'Attic (Loft) Insulation 100 mm top-up': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I68', 'Default': ''}
                , 'Attic (Loft) Insulation 150 mm top-up': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I69', 'Default': ''}
                , 'Attic (Loft) Insulation 200 mm top-up': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I70', 'Default': ''}
                , 'Attic (Loft) Insulation 250 mm top up': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I71', 'Default': ''}
                , 'Attic (Loft) Insulation 300 mm': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I72', 'Default': ''}
                
                , 'Attic Storage (5m2)': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I76', 'Default': ''}
                , 'Replacement of CWST/F&E tank': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I78', 'Default': ''}
                , 'Additional Roof Ventilation (High Level)': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I84', 'Default': ''}
                , 'Additional Roof Ventilation (Low Level)': { 'Value': '' , 'Tab': 'Roof' , 'Cell': 'I85', 'Default': ''}

                
                
                
                

                , 'Wall Type 1*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E27'}
                , 'Wall 1 wall thickness (mm)*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E29'}
                , 'Wall 1 Insulation Present?*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E33'}
                , 'Wall 1 Insulation Type*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E35'}
                , 'Wall 1 Fill Type*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E37'}
                , 'Wall 1 Residual Cavity Width (mm)*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'K37'}
                , 'Can Wall type 1 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'K33'}
                , "If 'Yes' enter Wall type 1 insulation thickness (mm)*": { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'K35'}
                , 'Wall Type 2': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'E27'}
                , 'Wall 2 wall thickness (mm)*': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'E29'}
                , 'Wall 2 Insulation Present?*': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'E33'}
                , 'Wall 2 Insulation Type*': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'E35'}
                , 'Wall 2 Fill Type*': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'E37'}
                , 'Wall 2 Residual Cavity Width (mm)*': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'K37'}
                , 'Can Wall type 2 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'K33'}
                , "If 'Yes' enter Wall type 2 insulation thickness (mm)*": { 'Value': '' , 'Tab': 'Wall Type 2' , 'Cell': 'K35'}
                , 'Wall Type 3': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'E27'}
                , 'Wall 3 wall thickness (mm)*': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'E29'}
                , 'Wall 3 Insulation Present?*': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'E33'}
                , 'Wall 3 Insulation Type*': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'E35'}
                , 'Wall 3 Fill Type*': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'E37'}
                , 'Wall 3 Residual Cavity Width (mm)*': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'K37'}
                , 'Can Wall type 3 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'K33'}
                , "If 'Yes' enter Wall type 3 insulation thickness (mm)*": { 'Value': '' , 'Tab': 'Wall Type 3' , 'Cell': 'K35'}
                , 'Wall Type 4': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'E27'}
                , 'Wall 4 wall thickness (mm)*': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'E29'}
                , 'Wall 4 Insulation Present?*': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'E33'}
                , 'Wall 4 Insulation Type*': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'E35'}
                , 'Wall 4 Fill Type*': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'E37'}
                , 'Wall 4 Residual Cavity Width (mm)*': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'K37'}
                , 'Can Wall type 4 Insulation Thickness be Measured?*': { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'K33'}
                , "If 'Yes' enter Wall type 4 insulation thickness (mm)*": { 'Value': '' , 'Tab': 'Wall Type 4' , 'Cell': 'K35'}
                , 'Is the property suitable for wall insulation? *': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E41'}
                , 'No wall insulation details *': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E43'}
                , 'Notes (Walls)': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E54'}
                , 'EWI/IWI > 25% *': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E47'}
                , 'Suitable for Draught Proofing': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E49'}
                , 'Not suitable details Draughtproofing*': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E51'}
                , 'Draught Proofing (<= 20m installed)': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L68', 'Default': ''}
                , 'Draught Proofing (> 20m installed)': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L69', 'Default': ''}
                , 'MEV 15l/s Bathroom': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L74', 'Default': ''}
                , 'MEV 30l/s Utility': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L75', 'Default': ''}
                , 'MEV 60l/s Kitchen': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L76', 'Default': ''}
                , 'New Permanent Vent': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L80', 'Default': ''}
                , 'New Background Vent': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L81', 'Default': ''}
                , 'Duct Cooker Hood': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L82', 'Default': ''}
                , 'Cavity Wall Insulation Bonded Bead': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F86', 'Default': ''}
                , 'Loose Fibre Extraction': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F87', 'Default': ''}
                , 'External Wall Insulation: Less than 60m2': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F88', 'Default': ''}
                , 'External Wall Insulation: 60m2 to 85m2': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F89', 'Default': ''}
                , 'External Wall Insulation: Greater than 85m2': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F90', 'Default': ''}
                , 'ESB alteration': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F91', 'Default': ''}
                , 'GNI meter alteration': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F92', 'Default': ''}
                , 'New Gas Connection': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F93', 'Default': ''}
                , 'RGI Meter_No Heating': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F94', 'Default': ''}
                , 'Internal Wall Insulation: Vertical Surface': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F99', 'Default': ''}
                , 'External wall insulation and CWI: less than 60m2': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F104', 'Default': ''}
                , 'External wall insulation and CWI: 60m2 to 85m2': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F105', 'Default': ''}
                , 'External wall insulation and CWI: greater than 85m2': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'F106', 'Default': ''}
                , 'replace_window_area': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'L113', 'Default': ''}
                , 'Notes (Windows and Doors)': { 'Value': '' , 'Tab': 'Wall' , 'Cell': 'E117', 'Default': ''}
                
                
                
                , 'Heating System *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E25'}
                , 'Qualifying Boiler': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E27'}
                , 'Major Renovation': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E29'}
                , 'System Age *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E31'}
                , 'Fully Working *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E33'}
                , 'Requires Service *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E37'}
                , 'Other Primary Heating Details *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'I25'}
                , 'Not Working Details Primary Heating *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'I33'}
                , 'Requires Service Details Primary Heating *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'I37'}
                , 'Hot Water System Exists *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E44'}
                , 'HWS': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'G44'}
                , 'Other HW Details *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'I49'}
                , 'Hot Water Cylinder*': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E55'}
                , 'Insulation *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E57'}
                , 'Condition of Lagging Jacket *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'H55'}
                , 'HWC Controls *': { 'Value': 'Cylinder Thermostat Controls *' , 'Tab': 'Heating' , 'Cell': 'H57'}
                , 'Heating Systems Controls *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E62'}
                , 'Partial Details *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E64'}
                , 'Programmer / Timeclock *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'H62'}
                , 'Room Thermostat Number *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'H64'}
                , 'Rads Number *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K62'}
                , 'TRVs Number *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K64'}
                , 'Suitable for Heating Measures *': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E68'}
                , 'Not suitable details*': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E70'}
                , 'Notes (Heating)': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'E74'}
                
                
                , 'Basic gas heating system': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K81', 'Default': ''}
                , 'Basic oil heating system': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K85', 'Default': ''}
                , 'Full gas heating system installation': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K89', 'Default': ''}
                , 'Full oil heating system installation': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K114', 'Default': ''}
                , 'Gas boiler and controls (Basic & controls pack)': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K118', 'Default': ''}
                , 'Hot Water Cylinder Jacket': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K127', 'Default': ''}
                , 'Oil boiler and controls (Basic & controls pack)': { 'Value': '' , 'Tab': 'Heating' , 'Cell': 'K167', 'Default': ''}
                
                
                
                , 'Secondary Heating System': { 'Value': '' , 'Tab': 'Heating - Secondary' , 'Cell': 'E25'}
                , 'Secondary System Age *': { 'Value': '' , 'Tab': 'Heating - Secondary' , 'Cell': 'E27'}
                , 'Secondary System Fully Working *': { 'Value': '' , 'Tab': 'Heating - Secondary' , 'Cell': 'E29'}
                , 'Secondary System Requires Service *': { 'Value': '' , 'Tab': 'Heating - Secondary' , 'Cell': 'E33'}
                , 'Not Working Details Secondary Heating *': { 'Value': '' , 'Tab': 'Heating - Secondary' , 'Cell': 'I29'}
                , 'Requires Service Details Secondary Heating *': { 'Value': '' , 'Tab': 'Heating - Secondary' , 'Cell': 'I33'}

                , 'Number of habitable rooms in the property': { 'Value': '' , 'Tab': 'Mechanical Ventilation Systems' , 'Cell': 'D55'}
                , 'Number of wet rooms in the property': { 'Value': '' , 'Tab': 'Mechanical Ventilation Systems' , 'Cell': 'D57'}
                , 'No. of habitable/wet rooms w/ open flued appliance': { 'Value': '' , 'Tab': 'Mechanical Ventilation Systems' , 'Cell': 'D59'}
                , 'LED Bulbs: supply only (4 no.)': { 'Value': '' , 'Tab': 'Mechanical Ventilation Systems' , 'Cell': 'E31', 'Default': ''}
                , 'Air-tightness test recommended?': { 'Value': '' , 'Tab': 'Mechanical Ventilation Systems' , 'Cell': 'E27', 'Default': ''}

                , 'Adequate Access*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'D25'}
                , 'No Access Details*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'I25'}
                , 'Cherry Picker Required*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'D27'}
                , 'Cherry Picker Required Details*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'I27'}
                , 'Mould/Mildew identified by surveyor; or reported by the applicant*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'D29'}
                , 'Mould/Mildew Details*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'I29'}
                , 'As confirmed by homeowner; property is a protected structure*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'D31'}
                , 'Protected Structure Details*': { 'Value': '' , 'Tab': 'Supplementary' , 'Cell': 'I31'}

                # , 'Surveyor Signature': { 'Value': '' , 'Tab': 'Declarations' , 'Cell': 'D27'}
                # , 'Surveyor Signature Date': { 'Value': '' , 'Tab': 'Declarations' , 'Cell': 'L27'}
                # , 'Customer Signature': { 'Value': '' , 'Tab': 'Declarations' , 'Cell': 'D29'}
                # , 'Customer Signature Date': { 'Value': '' , 'Tab': 'Declarations' , 'Cell': 'L29'}

                }
            
        if template == 'template_mrc':
            filename = json_val_dict['plan_name'] + ' Major Renovation calculation.xlsx'
            return_filename = filename
            container_name = 'attachment'
            local_path = "/tmp"
            print('local_path', ':', local_path)
            instance_file_path = os.path.join(local_path, filename)
            
            v = {
                'plan_name': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'D2'}
                , 'Thermal Envelope - Heat loss walls, windows and doors': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E6'}
                , 'Thermal Envelope - Heat loss floor area': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E7'}
                , 'Thermal Envelope - Heat loss roof area': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E8'}
                , 'THERMAL ENVELOPE OF BUILDING AREA': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E9'}
                , 'Heat loss Wall Area recommended for EWI and IWI': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E12'}
                , 'New Windows being recommended for replacement': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E13'}
                , 'TOTAL SURFACE AREA FOR MAJOR RENOVATION WORKS': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E14'}
                , 'Total surface area for MR works / Thermal Envelope': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E17'}
                , 'WARMER HOMES MAJOR RENOVATION RESULT': { 'Value': '' , 'Tab': 'Results' , 'Cell': 'E18'}
                }
        
        for field in v:
            # print('field', ':', field)
            if field in json_val_dict.keys():
                v[field]['Value'] = json_val_dict[field]
                # print("v[field]['Value']", ':', v[field]['Value'])
                
                if v[field]['Value'] in ['', 'N/A', 0]:
                    if "Default" in v[field].keys():
                        v[field]['Value'] = v[field]['Default']
                
                if type(v[field]['Value']) == str and '<BR>' in v[field]['Value']:
                    v[field]['Value'] = v[field]['Value'].replace('<BR>', '\n')
                # print(field, ':', v[field]['Value'])

        # print(v)


        # Create the BlobServiceClient object
        blob_service_client = BlobServiceClient(account_url, credential=default_credential)
        container_client = blob_service_client.get_container_client(container= container_name) 

        # if not os.path.exists(local_path):
            # print('about to create local_path', ':', local_path)
        try:
            error = ''
            os.mkdir(local_path)
        except Exception as ex:
            error = traceback.format_exc()
            # print('error', ':', error)
            
        
        
        
        print("\nDownloading blob to \n\t" + instance_file_path)
        
        
        template_file_name = template + '.xlsx'
        
        if template == 'template_ber':
            template_file_name = 'template_plc' + '.xlsx'
        
        with open(file=instance_file_path, mode="wb") as download_file:
            download_file.write(container_client.download_blob(template_file_name).readall())

        xfile = openpyxl.load_workbook(instance_file_path)
        
        for field in v:
            # print(field)
            sheet = xfile[v[field]['Tab']]
            sheet[v[field]['Cell']] = v[field]['Value']



        xfile.save(instance_file_path)


        instance_file_path = os.path.join(local_path, filename)
        with open(file=instance_file_path, mode="rb") as upload_file:
            blob_client = blob_service_client.get_blob_client(container=container_name, blob=filename)
            blob_client.upload_blob(upload_file, overwrite=True)
    

        
        print('created')
        
        
        # if template in ['template', 'template_mrc']:
        output = copy_from_container(json_val_dict['plan_name'], filename)
        
        
    except Exception as ex:
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # output = "Line " + str(exc_tb.tb_lineno) + ": " + exc_type 
        
        # output = str(ex)
        output = traceback.format_exc()
        # LOGGER.info('Exception : ' + str(traceback.format_exc()))
        
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print(exc_type, fname, exc_tb.tb_lineno)
        print(output)

    finally:
        return output, return_filename


def copy_from_container(plan_name
                        , filename
                        , container_from = 'attachment'
                        , local_path_from = "/tmp"
                        ):

    try:
        output = ''
        account_url = "https://ksnmagicplanfunc3e54b9.blob.core.windows.net"
        default_credential = DefaultAzureCredential()

        # filename = json_val_dict['plan_name'] + '.xlsx'
        # plan_name = 'WH571501 QA'
        # filename = plan_name + ' Major Renovation calculation' + '.xlsx'
        # container_name = 'attachment'
        
        container_to = 'project-files'
        local_path_to = plan_name

        # instance_file_path = os.path.join(local_path, filename)
        # print('instance_file_path', ':', instance_file_path)
        # instance_file_path_from = os.path.join(local_path_from, filename)
        instance_file_path_from = filename
        print('instance_file_path_from', ':', instance_file_path_from)
        instance_file_path_to = os.path.join(local_path_to, filename)
        print('instance_file_path_to', ':', instance_file_path_to)



        # Create the BlobServiceClient object
        blob_service_client = BlobServiceClient(account_url, credential=default_credential)
        container_client = blob_service_client.get_container_client(container= container_from) 
        
        
        file_content = container_client.download_blob(instance_file_path_from).readall()
        print('got file_content')

        blob_client = blob_service_client.get_blob_client(container=container_to, blob=instance_file_path_to)
        blob_client.upload_blob(file_content, overwrite=True)

        # with open(file=instance_file_path_to, mode="rb") as upload_file:
            # blob_client = blob_service_client.get_blob_client(container=container_to, blob=instance_file_path_from)
            # blob_client.upload_blob(upload_file, overwrite=True)

    except:
        output = traceback.format_exc()

    finally:
        # print('output', ':', output)
        return output
        

def lot(output_dict):
    try:
        Lot = 'S' # No works recommended
        
        recommended_works_dict = {
            "Internal Wall Insulation: Sloped or flat (horizontal) surface": "a"
            , "Attic (Loft) Insulation 100 mm top-up": "a"
            , "Attic (Loft) Insulation 150 mm top-up": "a"
            , "Attic (Loft) Insulation 200 mm top-up": "a"
            , "Attic (Loft) Insulation 250 mm top up": "a"
            , "Attic (Loft) Insulation 300 mm": "a"
            , "Cavity Wall Insulation Bonded Bead": "c"
            , "Internal Wall Insulation: Vertical Surface": "I"
            , "Basic gas heating system": "H"
            , "Basic oil heating system": "H"
            , "Full gas heating system installation": "H"
            , "Full oil heating system installation": "H"
            , "Gas boiler and controls (Basic & controls pack)": "H"
            , "Oil boiler and controls (Basic & controls pack)": "H"
            , "External Wall Insulation: Less than 60m2": "E"
            , "External Wall Insulation: 60m2 to 85m2": "E"
            , "External Wall Insulation: Greater than 85m2": "E"
            , "External wall insulation and CWI: less than 60m2": "E"
            , "External wall insulation and CWI: 60m2 to 85m2": "E"
            , "External wall insulation and CWI: greater than 85m2": "E"
            , "Window (same m2 rate will apply to windows with certified trickle vents)": "w"
        }
        Lot_upper = ''
        Lot_lower = ''
        for k in recommended_works_dict:
            v = recommended_works_dict[k]
            if k in output_dict.keys() and output_dict[k] not in ['N/A', 0, ""] and v not in (Lot_upper + Lot_lower):
                # print('adding: ', output_dict[k])
                if v.isupper():
                    Lot_upper += v
                else:
                    Lot_lower += v
        
        Lot_upper = Lot_upper.replace("IE", "E")
        Lot_upper = Lot_upper.replace("IH", "H")
        if Lot_upper == '':
            Lot_upper = 'S'
        else:
            Lot_lower = Lot_lower.replace('a', '')
            Lot_lower = Lot_lower.replace('c', '')
        Lot = Lot_upper + Lot_lower
        
        
        # print(Lot)
        
        # IEacw
        
        valid_lots = ['Sa', 'Sc', 'Sac', 'Scw', 'Sacw', 'I', 'Iw', 'E', 'Ew', 'H', 'Hw', 'HE', 'HEw', 'S']
        if Lot not in valid_lots:
            # print('invalid')
            Lot = Lot + ' (invalid)'
            print(list(valid_lots))

        

        # No longer in use: Sw, a, b, c
    except:
        Lot = traceback.format_exc()

    
    return Lot


def exterior_walls(root):
    try:
        print('ex_wa')
        
        ext_wall_area_gross = 0
        plan_name = root.get('name')
        interior_wall_width = root.get('interiorWallWidth') # always available?
        exteriorWallWidth = float(root.get('exteriorWallWidth')) # always available?
        extern_width_offset = interior_wall_width * 4
        extern_perim = 0
        exploded_wall_dict = {}
        
        floors = root.findall('interiorRoomPoints/floor')
        # floors = root.findall('floor')
        # print('no of floors', ':', len(floors))
        for floor in floors:
            floor_type = floor.get('floorType')
            # print('floor_type', ':', floor_type)
            if floor_type not in ['10', '11', '12', '13']:
                continue
            ft = floor_type
            exterior_walls = [] # {} 
            walls = floor.findall('exploded/wall')
            # print('no of walls', ':', len(walls))
            
            exploded_wall_dict[ft] = {}
            for i, wall in enumerate(walls):
                exploded_wall_dict[ft][i] = {}
                w_type = wall.find('type').text
                points = wall.findall('point') 
                p1, p2, *rest = points
                x1 = float(p1.get('x'))
                y1 = -float(p1.get('y'))
                x2 = float(p2.get('x'))
                y2 = -float(p2.get('y'))
                length = cart_distance((x1, y1), (x2, y2)) - (0.25 * exteriorWallWidth)
                # print(length)
                # wall_height = (float(p1.get('height')) + float(p2.get('height'))) / 2
                if floor_type == '10':
                    wall_height = 2.4
                if floor_type == '11':
                    wall_height = 2.7
                if floor_type in ['12', '13']:
                    wall_height = 2
                    
                area = wall_height * length
                # print('length ' + str(i) , ':', length)
                # print('wall_height ' + str(i) , ':', wall_height)
                # print('area ' + str(i) , ':', area)
                
                # (x1, y1), (x2, y2)
                
                
                exploded_wall_dict[ft][i]['x1'] = x1
                exploded_wall_dict[ft][i]['y1'] = y1
                exploded_wall_dict[ft][i]['x2'] = x2
                exploded_wall_dict[ft][i]['y2'] = y2
                exploded_wall_dict[ft][i]['type'] = w_type
                exploded_wall_dict[ft][i]['length'] = round(length, 2)
                
                if w_type == 'exterior':
                    ext_wall_area_gross += wall_height * length
                    extern_perim += length
                    # print(w_type, x1, y1, x2, y2, wall_height)
                    # print(w)
                    # exterior_walls.append(w)
            # print(exterior_walls)
            
        
        
                # extern_perim -= extern_width_offset
                # floors_perims.append(extern_perim)
                # wall_height = wall_height/nwalls if nwalls != 0 else wall_height
                # floors_heights.append(wall_height)
                

                # ext_wall_area_gross -= wall_types['Party Wall Area'][floor_index_adj] if 'Party Wall Area' in wall_types else 0
                # ext_wall_area_gross -= wall_types['Internal Wall Area'][floor_index_adj] if 'Internal Wall Area' in wall_types else 0

                # walls_area_gross.append(wall_area_gross)
            # print('ext_wall_area_gross', ':', str(ext_wall_area_gross))
            # print('extern_perim', ':', str(extern_perim))
    finally:
        return ext_wall_area_gross, exploded_wall_dict
